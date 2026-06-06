import os
import re
import html
import asyncio
import openpyxl
import logging
import pandas as pd
from uuid import uuid4
from datetime import datetime, timezone, timedelta, time
from pathlib import Path
import shutil
import requests
from typing import Optional
from fastapi import FastAPI, Request
import telegram.ext._jobqueue as tg_jobqueue
from telegram.error import BadRequest
from telegram import InputMediaPhoto, InputMediaVideo, InputMediaDocument
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, constants
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

# -----------------------------------------------------------
# 1) سجلات GO للاقتراحات والنقاشات
# -----------------------------------------------------------

suggestion_records = {}  # جميع اقتراحات المستخدمين
SUGGESTION_TICKET_COUNTER = 0  # عداد تذاكر مركز الدعم الفني (يزيد مع كل استفسار جديد)
SUGGESTION_REPLIES: dict[str, str] = {} 

team_threads: dict[int, dict] = {}  # نقاشات فريق GO الداخلية
TEAM_THREAD_COUNTER = 0
# عدّاد استخدام GO في الذاكرة فقط (بدون كتابة مباشرة على Excel)
GLOBAL_GO_COUNTER = 0

SUPPORT_LOCK_TTL_MIN = 10  # مدة القفل بالدقائق

def _now_dt():
    return datetime.now()

def _parse_dt(val):
    try:
        return datetime.fromisoformat(val)
    except Exception:
        return None

def _lock_expired(record):
    locked_at = record.get("locked_at")
    if not locked_at:
        return True
    dt = _parse_dt(locked_at)
    if not dt:
        return True
    return (_now_dt() - dt) > timedelta(minutes=SUPPORT_LOCK_TTL_MIN)

def lock_ticket(record, admin_id, admin_name):
    """
    قفل مؤقت فقط قبل أول رد.
    بعد أول رد (وجود replied_by) ما نحتاج قفل لأن نظامك يمنع غير نفس المشرف.
    """
    # ✅ إذا تم الرد مسبقًا، لا نطبق قفل
    if record.get("replied_by"):
        return True, ""

    # إذا في قفل صالح لمشرف آخر
    if record.get("locked_by_id") and not _lock_expired(record):
        if int(record.get("locked_by_id")) != int(admin_id):
            locker = record.get("locked_by_name") or "مشرف آخر"
            return False, f"🔒 التذكرة قيد المعالجة بواسطة: {locker}"
        return True, ""  # نفس المشرف

    # إذا القفل منتهي أو غير موجود، اقفلها
    record["locked_by_id"] = int(admin_id)
    record["locked_by_name"] = admin_name
    record["locked_at"] = _now_dt().isoformat()
    return True, ""

def unlock_ticket(record):
    record.pop("locked_by_id", None)
    record.pop("locked_by_name", None)
    record.pop("locked_at", None)

# -----------------------------------------------------------
# 2) نظام السجلات
# -----------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    force=True
)

# -----------------------------------------------------------
# 3) تصحيح set_application داخل JobQueue لإزالة weakref
# -----------------------------------------------------------

# دالة الكتابة في الخلفية
def write_excel_background(path, df, sheet_name):
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def _patched_set_application(self, application):
    """استبدال weakref بـ lambda للحفاظ على التطبيق دائماً."""
    self._application = lambda: application

tg_jobqueue.JobQueue.set_application = _patched_set_application

# -----------------------------------------------------------
# 4) إعداد التوكن
# -----------------------------------------------------------

API_TOKEN = os.getenv("TELEGRAM_TOKEN")

# -----------------------------------------------------------
# 5) تعريف initial_branches لتفادي NameError
# -----------------------------------------------------------

initial_branches = {
    "CHERY": [],
    "EXEED": [],
    "MG": [],
    "JETOUR": [],
    "JAECOO / OMODA": [],
    "BYD": [],
    "SOUEAST": [],
}
# -----------------------------------------------------------
# 6) تهيئة FastAPI + Telegram Application
# -----------------------------------------------------------

app = FastAPI()
application = Application.builder().token(API_TOKEN).updater(None).build()

# 🔒 قفل واحد لعمليات الكتابة على ملف Excel لمنع التعارض والتلف
EXCEL_LOCK = asyncio.Lock()

# 📁 مجلد النسخ الاحتياطي لملف الإكسل
BACKUP_DIR = Path("backups")
try:
    BACKUP_DIR.mkdir(exist_ok=True)
except Exception as e:
    logging.error(f"[BACKUP] ❌ فشل إنشاء مجلد النسخ الاحتياطي: {e}")

# ✅ PP deep-link toggle (from Render env)
_raw_pp_enabled = (os.getenv("PP_DIRECT_ENABLED") or "").strip().lower()
PP_DIRECT_ENABLED = _raw_pp_enabled in ("1", "true", "yes", "on")

# ✅ target bot username (without @)
PP_BOT_USERNAME = (os.getenv("PP_BOT_USERNAME") or "").strip().lstrip("@")

# ✅ Telegram cached animation file_id لتثبيت إرسال فيديو الترحيب بدون رفعه كل مرة
WELCOME_ANIMATION_FILE_ID = (os.getenv("WELCOME_ANIMATION_FILE_ID") or "").strip()

async def create_excel_backup(reason: str = "manual", context: Optional[ContextTypes.DEFAULT_TYPE] = None, notify_chat_id: Optional[int] = None):
    """إنشاء نسخة احتياطية من ملف bot_data.xlsx داخل مجلد backups"""
    src = Path("bot_data.xlsx")
    if not src.exists():
        logging.warning("[BACKUP] ⚠️ ملف bot_data.xlsx غير موجود – لا يمكن إنشاء نسخة احتياطية.")
        if context and notify_chat_id:
            try:
                await context.bot.send_message(
                    chat_id=notify_chat_id,
                    text="⚠️ لا يوجد ملف بيانات bot_data.xlsx حالياً، لم يتم إنشاء نسخة احتياطية."
                )
            except Exception:
                pass
        return

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    ts = now_saudi.strftime("%Y%m%d_%H%M%S")
    backup_name = f"bot_data_{ts}_{reason}.xlsx"
    backup_path = BACKUP_DIR / backup_name

    try:
        loop = asyncio.get_running_loop()
        # نضمن عدم تعارض أي عملية كتابة أخرى على نفس الملف
        async with EXCEL_LOCK:
            await loop.run_in_executor(None, shutil.copy2, src, backup_path)

        logging.info(f"[BACKUP] ✅ تم إنشاء نسخة احتياطية: {backup_path}")
        # إشعار الشخص الذي طلب النسخ (مثل المشرف في لوحة التحكم)
        if context and notify_chat_id:
            try:
                await context.bot.send_message(
                    chat_id=notify_chat_id,
                    text="✅ تم إنشاء نسخة احتياطية لبيانات النظام بنجاح."
                )
            except Exception:
                pass

        # إرسال النسخة الاحتياطية إلى قناة/قروب النسخ الاحتياطي إن وُجد TG_BACKUP_CHAT_ID
        if context:
            try:
                backup_chat_env = os.getenv("TG_BACKUP_CHAT_ID")
                backup_chat_id = int(backup_chat_env) if backup_chat_env else None
            except Exception:
                backup_chat_id = None

            if backup_chat_id:
                try:
                    with open(backup_path, "rb") as doc:
                        await context.bot.send_document(
                            chat_id=backup_chat_id,
                            document=doc,
                            caption=f"📦 نسخة احتياطية ({reason}) من بيانات نظام GO"
                        )
                except Exception as e2:
                    logging.error(f"[BACKUP] ❌ فشل إرسال النسخة الاحتياطية إلى قناة النسخ: {e2}")
    except Exception as e:
        logging.error(f"[BACKUP] ❌ فشل إنشاء النسخة الاحتياطية: {e}")
        if context and notify_chat_id:
            try:
                await context.bot.send_message(
                    chat_id=notify_chat_id,
                    text="❌ حدث خطأ أثناء إنشاء النسخة الاحتياطية."
                )
            except Exception:
                pass

async def daily_backup_job(context: ContextTypes.DEFAULT_TYPE):
    """نسخ احتياطي يومي تلقائي لملف الإكسل"""
    try:
        # نمرر context حتى يتمكن من الإرسال إلى قناة النسخ الاحتياطي إن وُجد TG_BACKUP_CHAT_ID
        await create_excel_backup(reason="daily", context=context, notify_chat_id=None)
    except Exception as e:
        logging.error(f"[BACKUP] ❌ خطأ أثناء تنفيذ النسخ الاحتياطي اليومي: {e}")



# إصلاح الخطأ: تعريف initial_branches قبل استخدامها
application.bot_data["branches"] = initial_branches

# -----------------------------------------------------------
# 7) قواعد البيانات – DataFrames فارغة حتى يتم التحميل لاحقاً
# -----------------------------------------------------------

df_admins = pd.DataFrame()
df_replies = pd.DataFrame()
df_branches = pd.DataFrame()
df_maintenance = pd.DataFrame()
df_parts = pd.DataFrame()
df_manual = pd.DataFrame()
df_independent = pd.DataFrame()
df_faults = pd.DataFrame()
df_group_logs: pd.DataFrame = pd.DataFrame()

# -----------------------------------------------------------
# 8) متغيرات عامة للنظام
# -----------------------------------------------------------

ALL_USERS = set()
user_sessions = {}

# مستخدمون قاموا بالتقييم (كاش في الذاكرة)
RATED_USERS: set[int] = set()

# كاش لقراءة شيتات الإحصائيات لتقليل القراءة من الإكسل
STATS_CACHE = {"excel_all": None, "loaded_at": None}
STATS_CACHE_TTL = 60  # ثانية

# إحصائيات ثابتة (تعويض سنتين تشغيل)
BASE_STATS = {
    "users": 22992,
    "groups": 14,
    "go_uses": 171950,
}

# تعويض تقييمات سنتين تشغيل (إحصائيات فقط، لا تُكتب في الإكسل)
BASE_RATINGS = {
    "count": 9302,   # 👈 عدّل هذا الرقم: عدد المقيمين الافتراضي القديم
     "avg": 5.0,     # 👈 متوسط التقييم (من 5)
}

# قائمة السيارات لخدمة قطع الغيار الاستهلاكية
unique_cars = []

# رسالة النماذج الغير جاهزة
PLACEHOLDER_TEXT = "هذا الطراز قيد التجهيز من قبل فريق GO"

## -----------------------------------------------------------
# 9) دليل تواصل الوكلاء
# -----------------------------------------------------------

BRAND_CONTACTS = {
    "CHERY": {
        "company": "سنابل الحديثة",
        "phone": "8002440228",
    },
    "EXEED": {
        "companies": [
            {"name": "سنابل الحديثة", "phone": "8002440228"},
            {"name": "التراث العربي للسيارات", "phone": "920035590"}
        ]
    },
    "JETOUR": {
        "company": "التوريدات الوطنية للسيارات",
        "phone": "920051222",
    },
    "MG": {
        "company": "جياد الحديثة للسيارات",
        "phone": "8002440390",
    },
    "BYD": {
        "company": "شركة الفطيم",
        "phone": "8003020006",
    },
    "JAECOO / OMODA": {
        "company": "العربات الفاخرة",
        "phone": "920031973",
    },
    "SOUEAST": {
        "company": "سير الشرق للسيارات",
        "phone": "8003050060",
    },
}

# 🆕 ربط أسماء البراندات (زي ما تجي من الإكسل) بمفتاح الوكيل في BRAND_CONTACTS
DEALER_FOR_BRAND = {
    "CHERY": "CHERY",
    "EXEED": "EXEED",
    "EXCEED": "EXEED",
    "JETOUR": "JETOUR",
    "MG": "MG",
    "BYD": "BYD",
    "JAECOO / OMODA": "JAECOO / OMODA",
    "SOUEAST": "SOUEAST",
}
# ==============================
# ✅ أدوات مساعدة للباك أب عند الإقلاع
# ==============================
BACKUPS_DIR = Path("backups")

# =============================
# Helpers: تحميل الإكسل مع النسخ الاحتياطية
# =============================

def _load_excel_from_path(path: Path) -> dict:
    """تحميل ملف إكسل واحد وإرجاع كل الشيتات في dict."""
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    return pd.read_excel(path, sheet_name=None)


def _load_excel_with_backup() -> dict:
    """
    يحاول:
    1) تحميل bot_data.xlsx من الجذر.
    2) إذا فشل → يحاول آخر نسخة من مجلد BACKUP_DIR.
    3) إذا فشل الكل → يرمي خطأ.
    """
    primary_path = Path("bot_data.xlsx")

    # 1) نحاول الملف الأساسي
    try:
        logging.info("[DATA LOAD] نحاول تحميل bot_data.xlsx الأساسي...")
        return _load_excel_from_path(primary_path)
    except Exception as e:
        logging.error(f"[DATA LOAD] فشل تحميل bot_data.xlsx الأساسي: {e}")

    # 2) نحاول آخر نسخة احتياطية (إن وجدت)
    try:
        if BACKUP_DIR.exists():
            backups = sorted(
                BACKUP_DIR.glob("*.xlsx"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
        else:
            backups = []

        if backups:
            latest = backups[0]
            logging.info(f"[DATA LOAD] نحاول التحميل من آخر نسخة احتياطية: {latest}")
            return _load_excel_from_path(latest)
        else:
            logging.error("[DATA LOAD] لا توجد أي ملفات نسخ احتياطية في BACKUP_DIR.")
    except Exception as e2:
        logging.error(f"[DATA LOAD] فشل تحميل آخر نسخة احتياطية: {e2}")

    # 3) إذا كل شيء فشل
    raise RuntimeError("فشل تحميل بيانات الإكسل من الملف الأساسي أو النسخ الاحتياطية.")

# ================================================================
#  تحميل بيانات Excel مع دعم النسخ الاحتياطية (نسخة منقّحة ونهائية)
# ================================================================
try:
    # 1) تحميل البيانات (أساسي + باك أب)
    excel_data = _load_excel_with_backup()

    # 2) قراءة الشيتات بأمان
    df_admins      = excel_data.get("managers",            pd.DataFrame(columns=["manager_id"]))
    df_replies     = excel_data.get("suggestion_replies",  pd.DataFrame(columns=["key", "reply"]))
    df_branches    = excel_data.get("branches",            pd.DataFrame())
    df_maintenance = excel_data.get("maintenance",         pd.DataFrame())
    df_parts       = excel_data.get("parts",               pd.DataFrame())
    df_manual      = excel_data.get("manual",              pd.DataFrame())
    df_independent = excel_data.get("independent",         pd.DataFrame())
    df_faults      = excel_data.get("faults",              pd.DataFrame())
    df_group_logs  = excel_data.get(
        "group_logs",
        pd.DataFrame(columns=["chat_id", "title", "type", "last_seen_utc"])
    )

    # 3) تحميل المجموعات المسجلة مسبقاً في BROADCAST_GROUPS
    global BROADCAST_GROUPS
    BROADCAST_GROUPS = {}
    if not df_group_logs.empty:
        for _, row in df_group_logs.iterrows():
            try:
                gid   = int(row.get("chat_id"))
                title = str(row.get("title", "غير معروف"))
                gtype = str(row.get("type", "group"))
                BROADCAST_GROUPS[gid] = {"title": title, "type": gtype}
            except Exception as e:
                logging.warning(f"[GROUP_LOG LOAD] فشل قراءة مجموعة: {e}")
    else:
        logging.info("[GROUP_LOG LOAD] شيت المجموعات فارغ.")

    # 4) استخراج قائمة السيارات الفريدة للقطع الاستهلاكية
    try:
        unique_cars = sorted(
            df_parts["Station No"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )
    except Exception as e2:
        logging.error(f"[DATA] فشل بناء unique_cars: {e2}")
        unique_cars = []

    # 5) all_users_log → ALL_USERS
    df_users = excel_data.get("all_users_log", pd.DataFrame(columns=["user_id"]))
    try:
        ALL_USERS = set(
            pd.to_numeric(df_users["user_id"], errors="coerce")
            .dropna()
            .astype(int)
            .tolist()
        )
    except Exception as e:
        logging.error(f"[ALL_USERS] فشل تحميل all_users_log: {e}")
        ALL_USERS = set()

    # 6) تحميل التقييمات RATED_USERS
    try:
        df_ratings_init = excel_data.get("ratings", pd.DataFrame())
        if not df_ratings_init.empty and "user_id" in df_ratings_init.columns:
            RATED_USERS = set(
                pd.to_numeric(df_ratings_init["user_id"], errors="coerce")
                .dropna()
                .astype(int)
                .tolist()
            )
        else:
            RATED_USERS = set()
    except Exception as e:
        logging.warning(f"[RATINGS INIT] فشل تحميل قائمة المقيمين: {e}")
        RATED_USERS = set()

    # 6 مكرر) تحميل عداد GO من شيت bot_stats (لو موجود)
    try:
        df_bot_stats_init = excel_data.get(
            "bot_stats",
            pd.DataFrame(columns=["key", "value"])
        )

        if (
            not df_bot_stats_init.empty
            and "key" in df_bot_stats_init.columns
            and "value" in df_bot_stats_init.columns
        ):
            row = df_bot_stats_init.loc[df_bot_stats_init["key"] == "total_go_uses"]
            if not row.empty:
                # لو فيه قيمة محفوظة نستخدمها بدل 0
                GLOBAL_GO_COUNTER = int(
                    pd.to_numeric(row["value"], errors="coerce").fillna(0).iloc[0]
                )
            else:
                GLOBAL_GO_COUNTER = 0
        else:
            GLOBAL_GO_COUNTER = 0

        logging.info(f"[GO STATS INIT] تم تحميل GLOBAL_GO_COUNTER = {GLOBAL_GO_COUNTER} من bot_stats")

    except Exception as e:
        logging.warning(f"[GO STATS INIT] فشل تحميل عداد GO من bot_stats: {e}")
        GLOBAL_GO_COUNTER = 0

    # 7) قائمة المشرفين AUTHORIZED_USERS
    try:
        if "manager_id" in df_admins.columns:
            AUTHORIZED_USERS = (
                pd.to_numeric(df_admins["manager_id"], errors="coerce")
                .dropna()
                .astype(int)
                .tolist()
            )
        else:
            AUTHORIZED_USERS = []
    except Exception as e:
        logging.error(f"[ADMINS] فشل تحميل قائمة المشرفين: {e}")
        AUTHORIZED_USERS = []

    # 8) الردود الجاهزة SUGGESTION_REPLIES
    if not df_replies.empty and "key" in df_replies.columns and "reply" in df_replies.columns:
        SUGGESTION_REPLIES = dict(zip(df_replies["key"], df_replies["reply"]))
    else:
        SUGGESTION_REPLIES = {}

    # 9) تحميل الفروع branches → مهم لقائمة مراكز الصيانة
    try:
        if not df_branches.empty:
            initial_branches = df_branches.to_dict(orient="records")
        else:
            initial_branches = []
    except Exception as e:
        logging.error(f"[BRANCHES] فشل تحويل شيت الفروع إلى records: {e}")
        initial_branches = []

    # 🔴 هذا السطر هو قلب مشكلة الفروع سابقاً – الآن يشتغل في حالة النجاح الطبيعية
    application.bot_data["branches"] = initial_branches

    # 10) ضمان أن df_group_logs دائماً له نفس الأعمدة
    if df_group_logs is None or df_group_logs.empty:
        df_group_logs = pd.DataFrame(
            columns=["chat_id", "title", "type", "last_seen_utc"]
        )

except Exception as e:
    # 🔥 فشل كامل في التحميل (الملف والباك أب)
    logging.error(f"[DATA LOAD ERROR] ⚠️ فشل قراءة بيانات الإكسل (الأساسي + النسخ الاحتياطية): {e}")

    # نعطي قيم آمنة حتى لا يتعطّل البوت
    df_admins      = pd.DataFrame(columns=["manager_id"])
    df_replies     = pd.DataFrame(columns=["key", "reply"])
    df_branches    = pd.DataFrame()
    df_maintenance = pd.DataFrame()
    df_parts       = pd.DataFrame()
    df_manual      = pd.DataFrame()
    df_independent = pd.DataFrame()
    df_faults      = pd.DataFrame()
    df_group_logs  = pd.DataFrame(columns=["chat_id", "title", "type", "last_seen_utc"])

    unique_cars      = []
    ALL_USERS        = set()
    RATED_USERS      = set()
    AUTHORIZED_USERS = []
    BROADCAST_GROUPS = {}

    SUGGESTION_REPLIES = {}
    initial_branches   = []
    application.bot_data["branches"] = initial_branches


async def show_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    صفحة إحصائيات GO + فتح التقييم في نفس الشاشة (HTML مسموح من تيليجرام)
    """
    query = update.callback_query
    data = query.data or ""
    user = query.from_user

    # استخراج user_id من الكول باك لو متوفر
    user_id = user.id
    if data.startswith("rate_"):
        try:
            user_id = int(data.split("_", 1)[1])
        except Exception:
            pass

    user_name_raw = user.full_name or "الصديق"
    user_name_safe = html.escape(user_name_raw)

    # === المستخدمون ===
    try:
        real_users = len(ALL_USERS)
    except Exception:
        real_users = 0
    total_users = BASE_STATS["users"] + real_users

    # === قراءة كل الشيتات مرة واحدة (مع كاش بسيط) ===
    global STATS_CACHE, STATS_CACHE_TTL

    now_utc = datetime.now(timezone.utc)

    if (
        STATS_CACHE["excel_all"] is None
        or STATS_CACHE["loaded_at"] is None
        or (now_utc - STATS_CACHE["loaded_at"]).total_seconds() > STATS_CACHE_TTL
    ):
        # نحتفظ بنسخة من الكاش القديم قبل المحاولة
        old_excel_all = STATS_CACHE["excel_all"]
        try:
            # ✅ قراءة الإكسل داخل قفل EXCEL_LOCK لتجنب القراءة أثناء الكتابة
            try:
                async with EXCEL_LOCK:
                    new_data = pd.read_excel("bot_data.xlsx", sheet_name=None)
            except Exception:
                new_data = pd.read_excel("bot_data.xlsx", sheet_name=None)

            STATS_CACHE["excel_all"] = new_data
            STATS_CACHE["loaded_at"] = now_utc
        except Exception as e:
            logging.warning(f"[STATS CACHE] فشل قراءة bot_data.xlsx للإحصائيات: {e}")
            # لو ما عندنا كاش قديم نهائياً، نضع قاموس فاضي مرة واحدة
            if old_excel_all is None:
                STATS_CACHE["excel_all"] = {}
                STATS_CACHE["loaded_at"] = now_utc
            else:
                # لو فيه كاش قديم، نرجع له ولا نمسحه
                STATS_CACHE["excel_all"] = old_excel_all

    excel_all = STATS_CACHE["excel_all"]

    # === المجموعات ===
    try:
        df_groups = excel_all.get("group_logs", pd.DataFrame())
        real_groups = df_groups["chat_id"].nunique() if not df_groups.empty else 0

        # ✅ fallback: لو كانت قراءة الإكسل فاشلة/فاضية، احسب من df_group_logs بالذاكرة
        if real_groups == 0:
            try:
                mem_df = globals().get("df_group_logs")
                if mem_df is not None and not mem_df.empty and "chat_id" in mem_df.columns:
                    real_groups = mem_df["chat_id"].nunique()
            except Exception:
                pass
    except Exception:
        real_groups = 0
        # ✅ fallback إضافي داخل except
        try:
            mem_df = globals().get("df_group_logs")
            if mem_df is not None and not mem_df.empty and "chat_id" in mem_df.columns:
                real_groups = mem_df["chat_id"].nunique()
        except Exception:
            pass

    total_groups = BASE_STATS["groups"] + real_groups

    # === مرات استخدام GO (من الذاكرة فقط) ===
    try:
        real_go = int(GLOBAL_GO_COUNTER)
    except Exception:
        real_go = 0

    total_go = BASE_STATS["go_uses"] + real_go

    # === التقييمات (مع BASE_RATINGS) ===
    rating_info = "⭐ لا توجد تقييمات مسجلة حاليًا"
    already_rated = False  # 👈 نستخدمها لتحديد إظهار أزرار التقييم أو إخفائها

    try:
        df_ratings = excel_all.get("ratings", pd.DataFrame())

        real_count = 0
        real_avg = 0.0
        if not df_ratings.empty and "rating" in df_ratings.columns:
            real_count = len(df_ratings)
            real_avg = float(df_ratings["rating"].mean())

        # 👇 التحقق هل هذا المستخدم قيّم سابقًا من الإكسل
        if not df_ratings.empty and "user_id" in df_ratings.columns:
            try:
                df_ratings["user_id"] = pd.to_numeric(df_ratings["user_id"], errors="coerce")
                already_rated = int(user_id) in df_ratings["user_id"].dropna().astype(int).tolist()
            except Exception:
                already_rated = False

        # ربط مع الكاش RATED_USERS
        if user_id in RATED_USERS:
            already_rated = True

        base_count = int(BASE_RATINGS.get("count", 0) or 0)
        base_avg = float(BASE_RATINGS.get("avg", 0.0) or 0.0)

        total_ratings = base_count + real_count
        if total_ratings > 0:
            combined_avg = (
                (base_avg * base_count + real_avg * real_count) / total_ratings
            )
        else:
            combined_avg = 0.0

        total_ratings_display = f"{total_ratings:,}".replace(",", "٬")
        combined_avg = round(combined_avg, 2)

        if total_ratings > 0:
            stars = "⭐" * min(5, int(round(combined_avg)))
            rating_info = (
                "⭐ التقييمات:\n"
                f"عدد المقيمين: <a href=\"tg://user?id=0\">{total_ratings_display}</a>\n"
                f"متوسط التقييم: <a href=\"tg://user?id=0\">{combined_avg}</a> من (5) {stars}"
            )
    except Exception:
        pass

    # === الوقت ===
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    refresh_time = (now_saudi + timedelta(minutes=12)).strftime("%I:%M %p")

    # سطر يوضح حالة تقييم المستخدم
    if already_rated:
        user_rating_line = "✅ <i>لقد قمت بتقييم نظام GO مسبقًا، شكرًا لدعمك.</i>\n\n"
    else:
        user_rating_line = ""

    # === بناء نص الإحصائيات (HTML مسموح) ===
    text = (
        "<b>📊 لوحة إحصائيات نظام الصيانة GO</b>\n"
        f"👤 <i>المستخدم:</i> <code><i>{user_name_safe}</i></code>\n"
        f"{user_rating_line}"
        "<b>📌 الملخص العام</b>\n"
        f"🏡 عدد القروبات المرتبطة بـ GO داخل النظام: <a href=\"tg://user?id=0\">{total_groups}</a>\n"
        f"👥 عدد مستخدمين GO داخل القروبات: <a href=\"tg://user?id=0\">{total_users}</a>\n"
        f"🚀 عدد مرات استدعاء GO داخل القروبات: <a href=\"tg://user?id=0\">{total_go}</a>\n\n"
        f"{rating_info}\n\n"
        "⏳ <code><i>تُحدَّث هذه الأرقام تلقائيًا مع نشاط الاعضاء.</i></code>\n"
        f"<code>{refresh_time} / 🇸🇦</code>\n\n"
        "🔹 <i>فريق GO يشكرك على ثقتك ودعمك المستمر.</i>"
    )

    # 👇 الكيبورد حسب حالة المستخدم:
    # لو قيّم سابقًا → فقط زر الرجوع
    # لو ما قيّم → تظهر أزرار التقييم + الرجوع
    if already_rated:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ الرجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        ])
    else:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("😞 غير راضٍ", callback_data=f"ratingval_1_{user_id}")],
            [InlineKeyboardButton("😐 مقبول", callback_data=f"ratingval_2_{user_id}")],
            [InlineKeyboardButton("😊 جيد", callback_data=f"ratingval_3_{user_id}")],
            [InlineKeyboardButton("😍 ممتاز", callback_data=f"ratingval_4_{user_id}")],
            [InlineKeyboardButton("⬅️ الرجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        ])

    try:
        await query.message.edit_text(
            text=text,
            reply_markup=keyboard,
            parse_mode=constants.ParseMode.HTML,
            disable_web_page_preview=True,
        )
    except BadRequest as e:
        # نتجاهل فقط حالة "Message is not modified"
        if "Message is not modified" in str(e):
            return
        raise

# ✅ 1. تعريف دالة تنظيف الجلسات
async def cleanup_old_sessions(context: ContextTypes.DEFAULT_TYPE, max_age_minutes: int = 15):
    """🧹 يحذف الجلسات القديمة من user_sessions لتقليل الضغط"""
    now = datetime.now(timezone.utc)
    removed = 0

    for user_id in list(user_sessions):
        original_count = len(user_sessions[user_id])
        user_sessions[user_id] = [
            msg for msg in user_sessions[user_id]
            if (now - msg["timestamp"]).total_seconds() < max_age_minutes * 60
        ]
        if not user_sessions[user_id]:
            del user_sessions[user_id]
            removed += original_count

    logging.info(f"[CLEANUP] 🧹 تم تنظيف {removed} رسالة من الجلسات القديمة.")
    return removed

# ================================================================
#  ⚙️ عدادات الإحصائيات: تحديث الذاكرة + حفظ فعلي في Excel
#  - group_logs      → للإحصائيات + الإرسال الجماعي
#  - ALL_USERS       → للإحصائيات + النسخ الاحتياطي
#  - total_go_uses   → عداد استخدام GO في bot_stats
# ================================================================
# 📌 حفظ ALL_USERS في Excel — يُستخدم في الإحصائيات والنسخ الاحتياطي
def _update_all_users_log_sync():
    """
    حفظ ALL_USERS في شيت all_users_log داخل bot_data.xlsx
    """
    global ALL_USERS
    try:
        df_users = pd.DataFrame(sorted(ALL_USERS), columns=["user_id"])

        with pd.ExcelWriter(
            "bot_data.xlsx",
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df_users.to_excel(writer, sheet_name="all_users_log", index=False)

        logging.info(f"[SAVE USERS] ✅ تم حفظ {len(ALL_USERS)} مستخدم في all_users_log")
    except Exception as e:
        logging.error(f"[SAVE USERS] ❌ فشل حفظ all_users_log في Excel: {e}")


async def update_all_users_log_async():
    """
    غلاف async لحفظ المستخدمين:
    - يشغل _update_all_users_log_sync في ثريد مستقل
    - حتى ما يبطّئ /go ولا start
    """
    try:
        loop = asyncio.get_running_loop()
        # استخدام قفل واحد لكل عمليات الكتابة على bot_data.xlsx
        async with EXCEL_LOCK:
            await loop.run_in_executor(None, _update_all_users_log_sync)
    except Exception as e:
        logging.error(f"[SAVE USERS] ❌ خطأ في تشغيل حفظ all_users_log في الخلفية: {e}")
        
# 📌 تحديث group_logs: تعديل الداتا في الذاكرة + حفظ مباشر في Excel
async def update_group_logs(chat_id: int, chat_title: str, context: ContextTypes.DEFAULT_TYPE):
    """
    تسجيل المجموعات في شيت group_logs + تحديث BROADCAST_GROUPS
    بدون تجميد البوت، وبشكل آمن مع النسخ الاحتياطية.
    """
    global df_group_logs, BROADCAST_GROUPS

    # لا نسجل الخاص – نسجل فقط المجموعات (chat_id يكون سالب)
    if chat_id >= 0:
        return

    # نحفظ داخل BROADCAST_GROUPS (مهم جداً للتوصيات)
    BROADCAST_GROUPS[chat_id] = {
        "title": chat_title or "غير معروف",
        "type": "group",
    }

    now_iso = datetime.now(timezone.utc).isoformat()

    # لو الشيت فيه صف سابق لنفس المجموعة -> نحدثه بدل ما نضيف واحد جديد
    if not df_group_logs.empty and (df_group_logs["chat_id"] == chat_id).any():
        mask = df_group_logs["chat_id"] == chat_id
        df_group_logs.loc[mask, "title"] = chat_title or "غير معروف"
        df_group_logs.loc[mask, "type"] = "group"
        df_group_logs.loc[mask, "last_seen_utc"] = now_iso
    else:
        # مجموعة جديدة -> نضيف صف واحد فقط
        new_row = {
            "chat_id": chat_id,
            "title": chat_title or "غير معروف",
            "type": "group",
            "last_seen_utc": now_iso,
        }
        df_group_logs = pd.concat(
            [df_group_logs, pd.DataFrame([new_row])],
            ignore_index=True
        )

    # حفظ للملف بدون تجميد البوت
    try:
        async with EXCEL_LOCK:
            await asyncio.to_thread(
                write_excel_background,
                "bot_data.xlsx",
                df_group_logs,
                "group_logs"
            )
    except Exception as e:
        logging.error(f"[GROUP_LOGS] فشل الحفظ في الخلفية: {e}")

async def register_user(user_id: int):
    """تسجيل مستخدم جديد في شيت all_users_log بشكل آمن وسريع"""
    global ALL_USERS

    if user_id in ALL_USERS:
        return  # مسجل مسبقاً

    ALL_USERS.add(user_id)

    # حفظ البيانات بالخلفية بدون تجميد البوت
    async with EXCEL_LOCK:
        await asyncio.to_thread(
            write_excel_background,
            "bot_data.xlsx",
            pd.DataFrame(sorted(ALL_USERS), columns=["user_id"]),
            "all_users_log"
        )

def _update_go_stats_sync():
    """
    عدّاد استخدام GO:
    - يزيد GLOBAL_GO_COUNTER في الذاكرة
    - يحفظ القيمة في شيت bot_stats داخل bot_data.xlsx
    """
    global GLOBAL_GO_COUNTER

    # 1) نحدّث العداد في الذاكرة
    GLOBAL_GO_COUNTER += 1

    try:
        # 2) نقرأ شيت bot_stats الحالي (لو موجود)
        try:
            df_bot_stats = pd.read_excel("bot_data.xlsx", sheet_name="bot_stats")
        except Exception:
            # لو ما فيه شيت بهالاسم أو أول مرة ننشئه
            df_bot_stats = pd.DataFrame(columns=["key", "value"])

        # نتأكد إن عندنا الأعمدة الرئيسية
        if "key" not in df_bot_stats.columns or "value" not in df_bot_stats.columns:
            df_bot_stats = pd.DataFrame(columns=["key", "value"])

        # 3) نحدّث أو نضيف السطر الخاص بـ total_go_uses
        if df_bot_stats.empty:
            df_bot_stats = pd.DataFrame(
                [{"key": "total_go_uses", "value": GLOBAL_GO_COUNTER}]
            )
        else:
            mask = df_bot_stats["key"] == "total_go_uses"
            if mask.any():
                df_bot_stats.loc[mask, "value"] = GLOBAL_GO_COUNTER
            else:
                df_bot_stats = pd.concat(
                    [
                        df_bot_stats,
                        pd.DataFrame(
                            [{"key": "total_go_uses", "value": GLOBAL_GO_COUNTER}]
                        ),
                    ],
                    ignore_index=True,
                )

        # 4) نحفظ الشيت في ملف bot_data.xlsx
        # هذا يستبدل شيت bot_stats فقط ويترك باقي الشيتات كما هي
        with pd.ExcelWriter(
            "bot_data.xlsx",
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df_bot_stats.to_excel(writer, sheet_name="bot_stats", index=False)

        logging.info(f"[GO STATS] ✅ تم حفظ total_go_uses = {GLOBAL_GO_COUNTER} في bot_stats")

    except Exception as e:
        logging.error(f"[GO STATS] ❌ فشل حفظ عداد GO في bot_stats: {e}")


async def update_go_stats_async():
    """
    غلاف async لزيادة عداد GO:
    - يشغّل _update_go_stats_sync في ثريد مستقل
    - يستخدم EXCEL_LOCK حتى لا يتعارض مع أي كتابة أخرى على bot_data.xlsx
    """
    try:
        loop = asyncio.get_running_loop()
        async with EXCEL_LOCK:
            await loop.run_in_executor(None, _update_go_stats_sync)
    except Exception as e:
        logging.error(f"[GO STATS] فشل تحديث عداد GO: {e}")


# ================================================================
#  ⚙️ health_log أيضًا يبقى في الذاكرة — الكتابة بالجوب لاحقًا
# ================================================================
HEALTH_BUFFER = []

def _write_health_log_sync():
    global HEALTH_BUFFER, ALL_USERS, GLOBAL_GO_COUNTER
    try:
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        HEALTH_BUFFER.append({
            "timestamp": now_saudi.isoformat(timespec="seconds"),
            "total_users": len(ALL_USERS),
            "total_go_uses": GLOBAL_GO_COUNTER,
        })
        logging.info("[HEALTH] buffered heartbeat")
    except Exception as e:
        logging.error(f"[HEALTH] فشل كتابة health_log في الذاكرة: {e}")


async def health_log_job(context: ContextTypes.DEFAULT_TYPE):
    try:
        _write_health_log_sync()
    except Exception as e:
        logging.error(f"[HEALTH LOG] خطأ أثناء تحديث health_log في الذاكرة: {e}")

# 🔁 جوب بسيط يطلب عنوان الخدمة لإبقاء Render مستيقظ
async def keepalive_ping(context: ContextTypes.DEFAULT_TYPE):
    try:
        base_url = os.getenv("RENDER_EXTERNAL_URL") or "https://chery-go-8a2z.onrender.com"

        # لو أحد كتبها بدون بروتوكول
        if not base_url.startswith("http"):
            base_url = "https://" + base_url.lstrip("/")

        # نستخدم ثريد منفصل عشان ما نحجز event loop
        await asyncio.to_thread(
            requests.get,
            base_url,
            timeout=5,
        )
        logging.info(f"[KEEPALIVE] ✅ Ping {base_url}")
    except Exception as e:
        logging.error(f"[KEEPALIVE] ❌ فشل Ping الخدمة: {e}")

def register_message(user_id, message_id, chat_id=None, context=None, skip_delete=False):
    if user_id not in user_sessions:
        user_sessions[user_id] = []

    user_sessions[user_id].append({
        "message_id": message_id,
        "chat_id": chat_id or user_id,
        "timestamp": datetime.now(timezone.utc)
    })

    # ✅ لا تقم بالحذف إذا skip_delete=True
    if not skip_delete and context and hasattr(context, "job_queue") and context.job_queue:
        try:
            context.job_queue.run_once(
                schedule_delete_message,
                timedelta(minutes=15),
                data={
                    "user_id": user_id,
                    "message_id": message_id,
                    "chat_id": chat_id or user_id
                }
            )
        except Exception as e:
            logging.warning(f"[JOB ERROR] فشل في جدولة الحذف التلقائي للرسالة {message_id}: {e}")

async def schedule_delete_message(context: ContextTypes.DEFAULT_TYPE):
    job_data = context.job.data
    chat_id = job_data.get("chat_id")
    message_id = job_data.get("message_id")
    user_id = job_data.get("user_id")

    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
        logging.info(f"[DELETE] 🗑️ تم حذف الرسالة رقم {message_id} للمستخدم {user_id}")
    except Exception:
        logging.warning(f"⚠️ الرسالة {message_id} للمستخدم {user_id} ربما حُذفت مسبقًا أو غير موجودة.")

async def reset_manual_search_state(context: ContextTypes.DEFAULT_TYPE):
    """تصـفير عداد البحث اليدوي (search_attempts) بعد 15 دقيقة من آخر استعلام"""
    job_data = getattr(context, "job", None).data if getattr(context, "job", None) else {}
    user_id = job_data.get("user_id")
    if user_id is None:
        return

    try:
        # user_data على مستوى التطبيق (أكثر أماناً داخل الجوب)
        user_data = context.application.user_data.get(user_id, {})
    except Exception:
        # احتياطاً
        user_data = context.user_data.get(user_id, {})

    if not isinstance(user_data, dict):
        return

    # حذف عداد البحث اليدوي
    user_data.pop("search_attempts", None)

    # إذا ما زالت الحركة parts نلغيها (جلسة بحث يدوي انتهت)
    if user_data.get("action") == "parts":
        user_data.pop("action", None)

    logging.info(f"[CLEANUP] ✅ تصفير عداد البحث اليدوي للمستخدم {user_id}")

async def log_event(update: Update, message: str, level="info"):
    user = update.effective_user
    chat = update.effective_chat
    timestamp = datetime.now(timezone.utc) + timedelta(hours=3)

    log_msg = (
        f"{timestamp:%Y-%m-%d %H:%M:%S} | "
        f"📩 من: [{user.full_name}] | "
        f"🆔 المستخدم: {user.id} | "
        f"📣 المحادثة: {chat.id} | "
        f"📝 {message}"
    )

    if level == "error":
        logging.error(log_msg)
    else:
        logging.info(log_msg)

    # 👇 هذا يضمن ظهور الرسالة في Runtime Logs حتى لو إعدادات اللوق تغيّرت
    print(log_msg)

def get_part_price(row: pd.Series) -> Optional[str]:
    """
    ترجع السعر كنص من الصف اذا كان العمود موجود وغير فارغ
    ندعم عدة أسماء أعمدة محتملة بما فيها Approx Price
    """
    candidate_cols = ["Approx Price", "Price", "price", "السعر", "التكلفة", "Cost", "cost"]
    for col in candidate_cols:
        if col in row:
            value = str(row[col]).strip()
            if value and value.lower() != "nan":
                return value
    return None

def make_back_button(target: str, user_id: int) -> InlineKeyboardButton:
    """
    يبني زر رجوع موحد
    target مثال: main / parts_menu / maintenance_menu / manual_menu ...
    """
    return InlineKeyboardButton("🔙 رجوع", callback_data=f"back:{target}:{user_id}")


async def handle_back(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """معالجة أزرار الرجوع الموحدة من نوع back:target:user_id"""
    query = update.callback_query
    raw = query.data or ""
    parts = raw.split(":")

    if len(parts) < 3:
        await query.answer("❌ زر رجوع غير معروف.", show_alert=True)
        return

    _, target, user_id_str = parts

    try:
        user_id = int(user_id_str)
    except ValueError:
        await query.answer("❌ خطأ في بيانات زر الرجوع.", show_alert=True)
        return

    # تجهيز كيبورد القائمة الرئيسية بشكل آمن
    kb = build_main_menu_keyboard(user_id)
    if isinstance(kb, InlineKeyboardMarkup):
        main_menu_markup = kb
    else:
        main_menu_markup = InlineKeyboardMarkup(kb)

    if target == "main":
        text_main = "فضلا اختار الخدمة المطلوبة 🛠️ :"
        try:
            # لو الرسالة نص نعدلها، لو صورة / ملف نرسل رسالة جديدة
            if query.message and query.message.text:
                msg = await query.edit_message_text(text_main, reply_markup=main_menu_markup)
            else:
                msg = await query.message.reply_text(text_main, reply_markup=main_menu_markup)
        except Exception:
            msg = await query.message.reply_text(text_main, reply_markup=main_menu_markup)

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "⬅️ رجوع الى القائمة الرئيسية (نظام back:main)")
        return

    # باقي الأهداف لاحقاً
    await query.answer("هذا زر رجوع لم يتم تفعيله بعد.", show_alert=True)

def build_main_menu_keyboard(user_id: int) -> InlineKeyboardMarkup:
    keyboard = [
        [InlineKeyboardButton("🔧 استعلامات قطع الغيار", callback_data=f"parts_{user_id}")],
        [InlineKeyboardButton("🚗 استعلامات الصيانة الدورية", callback_data=f"maintenance_{user_id}")],
        [InlineKeyboardButton("📘 استعراض دليل المالك", callback_data=f"manual_{user_id}")],
        [InlineKeyboardButton("🛠️ المتاجر ومراكز الخدمة", callback_data=f"service_{user_id}")],

        # ✅ هنا مكان زر السوق (كما طلبت بالضبط)
        [InlineKeyboardButton("🛒  سوق  قطع  غيار pp", callback_data=f"coming_{user_id}")],

        [InlineKeyboardButton("🔧 الأعطال الشائعة وحلولها", callback_data=f"faults_{user_id}")],
        [InlineKeyboardButton("✉️ مركز الدعم الفني والاستفسارات", callback_data=f"suggestion_{user_id}")],

        # زر الإحصائيات والتقييم
        [InlineKeyboardButton("📊 إحصائيات GO والتقييم", callback_data=f"rate_{user_id}")]
    ]

    # مميزات المشرفين
    if user_id in AUTHORIZED_USERS:
        keyboard.insert(-1, [InlineKeyboardButton("📡 إرسال توصية فنية", callback_data="send_reco")])
        keyboard.insert(-1, [InlineKeyboardButton("🟦 دعوة فريق GO للنقاش", callback_data=f"team_main_{user_id}")])

    return InlineKeyboardMarkup(keyboard)
       
# ✅ دالة البدء async
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.bot_data.get("maintenance_mode"):
        user_name = update.effective_user.full_name
        with open("GO-SS.PNG", "rb") as photo:
            msg = await update.message.reply_photo(
                photo=photo,
                caption=(
                    f"🛠️ مرحبا {user_name}\n\n"
                    "برنامج <b>GO</b> قيد التحديث والصيانة حالياً.\n"
                    "🔄 الرجاء المحاولة لاحقاً."
                ),
                parse_mode="HTML"
            )
        context.job_queue.run_once(
            lambda c: c.bot.delete_message(chat_id=msg.chat_id, message_id=msg.message_id),
            when=30
        )
        return

    user = update.effective_user
    chat = update.effective_chat
    user_id = user.id
    chat_id = chat.id
    user_name = user.full_name

    # حذف رسالة /start أو go الأصلية حتى لا تتكرر
    if update.message:
        try:
            await context.bot.delete_message(chat_id=chat_id, message_id=update.message.message_id)
        except Exception:
            pass

    # ✅ منع المتطفلين من الدخول من الخاص مباشرة بدون جلسة من المجموعة
    if chat.type == "private" and not context.user_data.get(user_id, {}).get("session_valid") and user_id not in AUTHORIZED_USERS:
        text = update.message.text.strip().lower() if update.message else ""
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")
        user_block = f"🧑‍🏫 مرحبا {user_name}"
        delete_block = f"⏳ سيتم حذف هذا التنبيه تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)"

        if text in ["/start", "start", "go", "/go"] and "start=go" not in text:
            alert_message = (
               "📣 يسعدنا اهتمامك بخدمات *نظام الصيانة GO*!\n\n"
               "❌ لا يمكنك بدء الخدمة مباشرة من الخاص.\n"
               "🔐 حفاظًا على الخصوصية، يرجى العودة إلى مجموعتك أو الانضمام إلى المجموعة الرئيسية أدناه وكتابة الأمر (go) هناك.\n\n"
               "[👥 اضغط هنا للانضمام إلى مجموعة CHERY KSA ](https://t.me/CHERYKSA_group)"
            )
        else:
            alert_message = (
                "🚫 عذرًا، لا يمكنك بدء الخدمة بهذه الطريقة.\n"
                "🔐 زر الانطلاق يستعمل لمره واحدة وهو مخصص فقط لمن بدأ الجلسة من المجموعة بنفسه.\n"
                "✳️ يرجى العودة إلى المجموعة وكتابة الأمر (go) يدويًا لبدء الخدمة."
            )

        msg = await update.message.reply_text(
            f"{user_block}\n\n{alert_message}\n\n{delete_block}",
            parse_mode=constants.ParseMode.MARKDOWN,
            disable_web_page_preview=True
        )
        register_message(user_id, msg.message_id, chat_id, context)
        return

    # تنظيف مفاتيح image_opened_ لمنع التعارض في فتح الصور القديمة
    keys_to_remove = [key for key in context.user_data.get(user_id, {}) if key.startswith("image_opened_")]
    for key in keys_to_remove:
        del context.user_data[user_id][key]

    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["manual_sent"] = False

    # ✅ تسجيل المستخدم في all_users_log (تحديث في الخلفية عند أول استخدام)
    global ALL_USERS
    if user_id not in ALL_USERS:
        ALL_USERS.add(user_id)
        try:
            asyncio.create_task(update_all_users_log_async())
        except Exception as e:
            logging.error(f"[SAVE USERS] فشل جدولة حفظ all_users_log في الخلفية: {e}")

    # ✅ تحديث عداد استخدام go في الخلفية (بدون تعطيل رسالة الترحيب والقوائم)
    try:
        asyncio.create_task(update_go_stats_async())
    except Exception as e:
        logging.error(f"[SAVE STATS] فشل جدولة تحديث /go في الخلفية: {e}")

    # ✅ استرجاع بيانات المجموعة المحفوظة للمستخدم
    group_title = context.user_data[user_id].get("group_title", "غير معروف")
    group_id = context.user_data[user_id].get("group_id", user_id)
    previous_user_name = context.user_data[user_id].get("user_name", user_name)

    if chat_id > 0 and user_id in context.bot_data:
        bot_data = context.bot_data[user_id]
        context.user_data[user_id].update(bot_data)
        del context.bot_data[user_id]

        group_title = bot_data.get("group_title", "غير معروف")
        group_id = bot_data.get("group_id", user_id)
        previous_user_name = bot_data.get("user_name", user_name)

    context.user_data[user_id].update({
        "action": None,
        "compose_text": None,
        "compose_media": None,
        "compose_mode": None,
        "group_title": group_title,
        "group_id": group_id,
        "user_name": previous_user_name,
        "final_group_name": group_title,
        "final_group_id": group_id
    })

    await log_event(update, "بدأ المستخدم التفاعل مع /go")

  # ✅ إذا النداء من مجموعة: نرسل بانر الترحيب ونخرج
    if chat_id < 0:
        context.bot_data[user_id] = {
            "group_title": update.effective_chat.title or "غير معروف",
            "group_id": chat_id,
            "user_name": user_name
        }

        video_path = "GO-CHERY.MP4"

    # ✅ اختصار شاشة الترحيب (اسم + سطرين فقط)
        full_caption = (
            f"`👤 {user_name}`\n"
            "✨ مرحبًا بك في نظام الصيانة والدعم الفني GO ومنصة PP لقطع الغيار\n"
        )

        bot_username = context.bot.username
        link = f"https://t.me/{bot_username}?start=go"
        keyboard = [[InlineKeyboardButton("🚀 ابدأ الخدمة الآن", url=link)]]

                try:
            if WELCOME_ANIMATION_FILE_ID:
                try:
                    msg = await context.bot.send_animation(
                        chat_id=chat_id,
                        animation=WELCOME_ANIMATION_FILE_ID,
                        caption=full_caption,
                        reply_markup=InlineKeyboardMarkup(keyboard),
                        parse_mode=constants.ParseMode.MARKDOWN
                    )
                except Exception as e:
                    logging.warning(f"[GO GROUP] تعذر إرسال فيديو الترحيب عبر file_id وسيتم استخدام الملف المحلي: {e}")
                    if os.path.exists(video_path):
                        with open(video_path, "rb") as video:
                            msg = await context.bot.send_animation(
                                chat_id=chat_id,
                                animation=video,
                                caption=full_caption,
                                reply_markup=InlineKeyboardMarkup(keyboard),
                                parse_mode=constants.ParseMode.MARKDOWN
                            )

                        try:
                            attachment = getattr(msg, "effective_attachment", None)
                            file_id = (
                                getattr(attachment, "file_id", None)
                                or getattr(getattr(msg, "animation", None), "file_id", None)
                                or getattr(getattr(msg, "video", None), "file_id", None)
                                or getattr(getattr(msg, "document", None), "file_id", None)
                            )

                            if file_id:
                                logging.info(f"[GO GROUP] WELCOME_ANIMATION_FILE_ID={file_id}")
                            else:
                                logging.warning(f"[GO GROUP] لم يتم العثور على file_id في رسالة الترحيب رقم {msg.message_id}")
                        except Exception as e2:
                            logging.warning(f"[GO GROUP] تعذر استخراج file_id لفيديو الترحيب: {e2}")
                    else:
                        msg = await context.bot.send_message(
                            chat_id=chat_id,
                            text=full_caption,
                            reply_markup=InlineKeyboardMarkup(keyboard),
                            parse_mode=constants.ParseMode.MARKDOWN
                        )

            elif os.path.exists(video_path):
                with open(video_path, "rb") as video:
                    msg = await context.bot.send_animation(
                        chat_id=chat_id,
                        animation=video,
                        caption=full_caption,
                        reply_markup=InlineKeyboardMarkup(keyboard),
                        parse_mode=constants.ParseMode.MARKDOWN
                    )

                try:
                    attachment = getattr(msg, "effective_attachment", None)
                    file_id = (
                        getattr(attachment, "file_id", None)
                        or getattr(getattr(msg, "animation", None), "file_id", None)
                        or getattr(getattr(msg, "video", None), "file_id", None)
                        or getattr(getattr(msg, "document", None), "file_id", None)
                    )

                    if file_id:
                        logging.info(f"[GO GROUP] WELCOME_ANIMATION_FILE_ID={file_id}")
                    else:
                        logging.warning(f"[GO GROUP] لم يتم العثور على file_id في رسالة الترحيب رقم {msg.message_id}")
                except Exception as e:
                    logging.warning(f"[GO GROUP] تعذر استخراج file_id لفيديو الترحيب: {e}")

            else:
                msg = await context.bot.send_message(
                    chat_id=chat_id,
                    text=full_caption,
                    reply_markup=InlineKeyboardMarkup(keyboard),
                    parse_mode=constants.ParseMode.MARKDOWN
                )

            register_message(user_id, msg.message_id, chat_id, context, skip_delete=True)

            if context and hasattr(context, "job_queue") and context.job_queue:
                context.job_queue.run_once(
                    schedule_delete_message,
                    timedelta(seconds=90),
                    data={"user_id": user_id, "message_id": msg.message_id, "chat_id": chat_id}
                )

        except Exception as e:
            logging.error(f"[GO GROUP] فشل إرسال الترحيب بالفيديو: {e}")

    # ✅ بعد إرسال رسالة الترحيب في المجموعة، حدّث group_logs في الخلفية لاستخدامها في بث التوصيات
        try:
            asyncio.create_task(update_group_logs(
                chat.id,
                chat.title or "",
                context
            ))
        except Exception as e:
            logging.warning(f"[GROUP_LOGS] فشل جدولة تحديث group_logs من start للقروب {chat.id}: {e}")

        return  # ← هذا return يُنهي فرع المجموعة فقط

    # ------------------------------------------------------------------------
    # من هنا الخاص
    # ------------------------------------------------------------------------

    context.user_data[user_id].pop("suggestion_used", None)
    context.user_data[user_id].pop("search_attempts", None)

    keyboard = build_main_menu_keyboard(user_id)

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    mmsg1 = await update.message.reply_text(
        f"`🧑‍💼 مرحباً {user_name}`\n\n"
        "🚀 *أهلًا بك في نظام GO للاستعلام الفني والخدمات المساندة.*\n"
        "يوفّر لك النظام معلومات دقيقة عن الصيانة، وحلول الأعطال، ودليل الاستخدام، مع دعم فني مباشر عند الحاجة.\n\n"
        "💡 *أنت الآن في جلسة استعلام تفاعلية، وستظهر لك بالأسفل قائمة خدمات GO لاختيار ما يناسبك.*\n\n"
        f"`⏳ سيتم حذف هذه الرسالة خلال 15 دقائق ({delete_time} / 🇸🇦)`",
        parse_mode=constants.ParseMode.MARKDOWN
    )

    msg2 = await update.message.reply_text(
        "فضلاً اختر الخدمة المطلوبة 🛠️ :",
        reply_markup=keyboard
    )

    # 🧽 تنظيف مفاتيح الجلسة القديمة
    for key in list(context.user_data[user_id].keys()):
        if key.startswith("image_opened_") or key.endswith("_used") or key.endswith("_sent"):
            context.user_data[user_id].pop(key, None)

    register_message(user_id, mmsg1.message_id, chat_id, context)
    register_message(user_id, msg2.message_id, chat_id, context)

    for key in list(context.user_data[user_id].keys()):
        if key.startswith("cat_used_"):
            context.user_data[user_id].pop(key, None)

    context.user_data[user_id]["session_valid"] = False


async def handle_go_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    user = update.effective_user
    user_id = user.id
    user_name = user.full_name
    chat_id = chat.id

    # 🧾 لو جت من مجموعة: جهّز الجلسة ثم أرسل الترحيب أولاً، وبعدها حدّث group_logs في الخلفية
    if chat.type != "private":
        # حفظ بيانات القروب للمستخدم عشان نستخدمها لما ينتقل للخاص
        context.bot_data[user_id] = {
            "group_title": chat.title or "غير معروف",
            "group_id": chat.id,
            "user_name": user_name
        }

        # إنشاء جلسة مؤقتة صالحة لمرة واحدة فقط
        context.user_data[user_id] = context.user_data.get(user_id, {})
        context.user_data[user_id]["session_valid"] = True

        # تنظيف مفاتيح الصور القديمة
        keys_to_remove = [key for key in context.user_data[user_id] if key.startswith("image_opened_")]
        for key in keys_to_remove:
            del context.user_data[user_id][key]

        # ✅ أرسل بانر GO / زر الانطلاق بسرعة
        await start(update, context)

        # ✅ بعد إرسال الرسالة للمجموعة، حدّث group_logs في الخلفية بدون ما تأخر الترحيب
        try:
            asyncio.create_task(update_group_logs(
                chat.id,
                chat.title or "",
                context
            ))
        except Exception as e:
            logging.warning(f"[GROUP_LOGS] فشل جدولة تحديث group_logs للقروب {chat.id}: {e}")

        logging.info(f"[GO من المجموعة] سجلنا بيانات المجموعة {chat.title} / {chat.id} للمستخدم {user.full_name}")
        return

    # ✅ من هنا: التعامل في الخاص
    if chat.type == "private" and (
        not context.user_data.get(user_id, {}).get("session_valid")
    ) and user_id not in AUTHORIZED_USERS:
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

        user_block = f"🧑‍🏫 مرحبا {user_name}"
        alert_message = (
            "📣 يسعدنا اهتمامك بخدمات *نظام الصيانة GO*!\n\n"
            "❌ لا يمكنك بدء الخدمة مباشرة من الخاص.\n"
            "🔐 حفاظًا على الخصوصية، يرجى العودة إلى مجموعتك أو الانضمام إلى المجموعة الرئيسية أدناه وكتابة الأمر (go) هناك.\n\n"
            "[👥 اضغط هنا للانضمام إلى مجموعة CHERY KSA ](https://t.me/CHERYKSA_group)"
        )
        delete_block = f"⏳ سيتم حذف هذا التنبيه تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)"

        msg = await update.message.reply_text(
            f"{user_block}\n\n{alert_message}\n\n{delete_block}",
            parse_mode=constants.ParseMode.MARKDOWN,
            disable_web_page_preview=True
        )
        register_message(user_id, msg.message_id, chat_id, context)
        return

    # ✅ في الخاص مع جلسة صالحة أو مشرف → نترك دالة start تكمل نفس منطق الترحيب والقائمة
    await start(update, context)
    
async def start_suggestion_session(user_id, context):
    
    # ✅ لو عند المستخدم جلسة سابقة غير مرسلة، نعيد استخدام نفس التذكرة
    user_state = context.user_data.get(user_id, {})
    active_id = user_state.get("active_suggestion_id")
    if active_id and user_id in suggestion_records:
        existing_record = suggestion_records[user_id].get(active_id)
        if existing_record and not existing_record.get("submitted"):
            return active_id

    # otherwise نفتح تذكرة جديدة
    suggestion_id = uuid4().hex

    # ✅ توليد رقم تذكرة تسلسلي ثابت من الإكسل (bot_stats)
    global SUGGESTION_TICKET_COUNTER

    # ====== ✅ تعويض الأرقام القديمة + منع الرجوع للصفر ======
    # أول رقم مطلوب يبدأ من 2623، لذلك نخزن/نقرأ آخر رقم كـ 2622
    BASE_COUNTER = 2622

    last_counter = await get_bot_stat_value("suggestion_ticket_counter", BASE_COUNTER)
    try:
        last_counter = int(last_counter)
    except Exception:
        last_counter = BASE_COUNTER

    if last_counter < BASE_COUNTER:
        last_counter = BASE_COUNTER

    if isinstance(last_counter, int) and last_counter >= SUGGESTION_TICKET_COUNTER:
        SUGGESTION_TICKET_COUNTER = last_counter
    # ================================================

    SUGGESTION_TICKET_COUNTER += 1
    ticket_no = SUGGESTION_TICKET_COUNTER

    await set_bot_stat_value("suggestion_ticket_counter", SUGGESTION_TICKET_COUNTER)

    context.user_data.setdefault(user_id, {})

    # ✅ تثبيت بيانات المجموعة بشكل مضمون (من user_data أو bot_data)
    group_name = context.user_data[user_id].get("group_title")
    group_id = context.user_data[user_id].get("group_id")
    user_name = context.user_data[user_id].get("user_name", "—")

    if (not group_name or not group_id) and user_id in context.bot_data:
        fallback = context.bot_data[user_id]
        group_name = fallback.get("group_title", group_name)
        group_id = fallback.get("group_id", group_id)
        user_name = fallback.get("user_name", user_name)
        del context.bot_data[user_id]

    group_name = group_name or "غير معروف"
    group_id = group_id or "غير معروف"

    # ✅ سجل الاقتراح
    suggestion_records.setdefault(user_id, {})
    suggestion_records[user_id][suggestion_id] = {
        "ticket_no": ticket_no,
        "text": None,
        "media": None,
        "submitted": False,
        "admin_messages": {},
        "group_name": group_name,
        "group_id": group_id,
        "user_name": user_name,
        "reply_count": 0,
    }

    context.user_data[user_id]["active_suggestion_id"] = suggestion_id
    return suggestion_id

async def get_bot_stat_value(key: str, default=0):
    try:
        async with EXCEL_LOCK:
            wb = openpyxl.load_workbook("bot_data.xlsx")
            if "bot_stats" not in wb.sheetnames:
                return default

            ws = wb["bot_stats"]

            for row in range(2, ws.max_row + 1):
                k = ws.cell(row=row, column=1).value
                if str(k).strip() == str(key).strip():
                    v = ws.cell(row=row, column=2).value
                    try:
                        return int(v)
                    except Exception:
                        return default

            return default
    except Exception:
        return default

async def set_bot_stat_value(key: str, value):
    async with EXCEL_LOCK:
        wb = openpyxl.load_workbook("bot_data.xlsx")
        if "bot_stats" not in wb.sheetnames:
            ws = wb.create_sheet("bot_stats")
            ws.cell(row=1, column=1).value = "key"
            ws.cell(row=1, column=2).value = "value"
        else:
            ws = wb["bot_stats"]

        found = False
        for row in range(2, ws.max_row + 1):
            k = ws.cell(row=row, column=1).value
            if str(k).strip() == str(key).strip():
                ws.cell(row=row, column=2).value = value
                found = True
                break

        if not found:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1).value = key
            ws.cell(row=next_row, column=2).value = value

        wb.save("bot_data.xlsx")

def _next_team_thread_id() -> int:
    """توليد رقم تسلسلي لكل نقاش داخلي لفريق GO"""
    global TEAM_THREAD_COUNTER
    TEAM_THREAD_COUNTER += 1
    return TEAM_THREAD_COUNTER


async def handle_team_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """استقبال رسالة داخلية من مشرف ضمن نقاش فريق GO"""
    message = update.message
    admin = update.effective_user
    admin_id = admin.id

    text = (message.text or "").strip()
    if not text:
        await message.reply_text("⚠️ الرسالة الداخلية يجب أن تكون نصية.")
        return

    state = context.user_data.get(admin_id, {})
    thread_id = state.get("team_thread_id")
    if not thread_id or thread_id not in team_threads:
        await message.reply_text("⚠️ لا توجد جلسة نقاش داخلي نشطة.")
        state["team_mode"] = False
        state.pop("team_thread_id", None)
        return

    thread = team_threads[thread_id]
    thread.setdefault("messages", [])
    thread["messages"].append(
        {
            "from": admin_id,
            "name": admin.full_name,
            "text": text,
            "at": datetime.now(timezone.utc).isoformat()
        }
    )

    # عداد ردود النقاش
    reply_count = thread.get("reply_count", 0) + 1
    thread["reply_count"] = reply_count

    # سياق النقاش الخاص بالاستفسار (اسم العضو، رقم التذكرة...)
    ctx = thread.get("context", {}) or {}

    header_lines = [
        f"🧵 نقاش فريق GO رقم #{thread_id}",
        f"🔁 رد رقم {reply_count} من: {admin.full_name} ({admin_id})",
    ]

    # لو النقاش مرتبط باستفسار مركز الدعم
    if thread.get("type") == "suggestion":
        member_name = ctx.get("user_name", "غير معروف")
        member_id = ctx.get("user_id", "غير معروف")
        group_name = ctx.get("group_name", "غير معروف")
        group_id = ctx.get("group_id", "غير معروف")

        # رقم التذكرة الجديد
        ticket_no = ctx.get("ticket_no")
        if ticket_no:
            header_lines.append(f"🆔 رقم التذكرة: #{ticket_no}")

        header_lines.append("")
        header_lines.append(f"👤 العضو: {member_name} ({member_id})")
        header_lines.append(f"🏘️ المجموعة: {group_name} ({group_id})")

        # النص الأصلي للاستفسار
        original_text = (ctx.get("text") or "").strip()
        if original_text:
            header_lines.append("")
            header_lines.append("📝 نص استفسار العضو:")
            header_lines.append(f"```{original_text}```")

    header = "\n".join(header_lines)
    body = f"{header}\n\n💬 مداخلة المشرف:\n```{text}```"

    # إيقاف وضع الكتابة لهذا المشرف
    state["team_mode"] = False
    state.pop("team_thread_id", None)

    # إرسال الرسالة لكل المشرفين
    for aid in AUTHORIZED_USERS:
        try:
            buttons = [
                [InlineKeyboardButton("✉️ رد على هذا النقاش", callback_data=f"team_reply_{thread_id}")]
            ]
            reply_markup = InlineKeyboardMarkup(buttons)

            await context.bot.send_message(
                chat_id=aid,
                text=body,
                parse_mode=constants.ParseMode.MARKDOWN,
                reply_markup=reply_markup
            )
        except Exception as e:
            logging.warning(f"[TEAM_THREAD] فشل إرسال رد النقاش للمشرف {aid}: {e}")

# =========================== توصيات فنية عامة للمجموعات ===========================

async def start_recommendation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """بدء وضع كتابة توصية فنية من مشرف"""
    query = update.callback_query
    admin_id = query.from_user.id
    admin_name = query.from_user.full_name

    if admin_id not in AUTHORIZED_USERS:
        await query.answer("هذه الميزة متاحة لمشرفي نظام GO فقط.", show_alert=True)
        return

    # 🔒 حالة المشرف
    ud = context.user_data.setdefault(admin_id, {})

    # ✅ إغلاق أي وضع قد يتعارض مع التوصية
    # مثل: بحث قطع الغيار النصي، مركز الدعم، أو وضع compose قديم
    ud["action"] = None
    ud["compose_mode"] = None
    ud["search_attempts"] = 0  # احتياط، لو كان داخل جلسة بحث
    # نخلي الجلسة صالحة في الخاص
    ud["session_valid"] = True

    # ✅ تفعيل وضع التوصية من جديد
    ud["reco_mode"] = "awaiting_reco"
    ud["reco_text"] = None
    ud["reco_media"] = None
    ud["reco_entities"] = None
    ud["reco_selected"] = []
    ud["reco_pin"] = False  # افتراضياً غير مفعّل

    # تجهيز قائمة المجموعات المتاحة للبث لهذا المشرف
    _prepare_reco_targets_for_admin(admin_id, context)

    await query.answer()

    # 🔹 اسم المشرف أعلى الرسالة (باهت)
    admin_block = f"`👤 المشرف: {admin_name}`"

    # 🔹 نص الرسالة بالتنسيق المطلوب
    text = (
        f"{admin_block}\n\n"
        "📡 إرسال توصية فنية للمجموعات\n"
        "`✏️ أرسل الآن نص التوصية التي ترغب بنشرها، وسيظهر لك بعد ذلك خيار بثها على جميع المجموعات أو اختيار مجموعات معيّنة يدويًا.`\n\n"
        "📎 `يمكنك إرفاق وسيط واحد فقط (صورة أو مستند أو فيديو أو رسالة صوتية) مع التوصية.`\n\n"
        "ℹ️ `بعد الإرسال ستظهر لك معاينة قبل البث النهائي.`"
    )

    # 🔹 زر إلغاء التوصية + الرجوع للقائمة
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ إلغاء التوصية والرجوع للقائمة", callback_data="reco_cancel")]
    ])

    msg = await query.message.reply_text(
        text,
        reply_markup=keyboard,
        parse_mode=ParseMode.MARKDOWN
    )

    # لحذف الرسالة مع الضغط على الزر نحتاج تسجيل رقمها
    context.user_data[admin_id]["reco_message_id"] = msg.message_id
    context.user_data[admin_id]["reco_chat_id"] = msg.chat_id

async def handle_recommendation_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """استقبال نص/وسائط التوصية من المشرف وتجهيز المعاينة"""
    admin_id = update.effective_user.id
    if admin_id not in AUTHORIZED_USERS:
        return

    ud = context.user_data.setdefault(admin_id, {})
    if ud.get("reco_mode") != "awaiting_reco":
        # ليس في وضع التوصية
        return

    message = update.message

    # نص التوصية: إما text أو caption للوسائط
    incoming_text = (message.text or message.caption or "").strip()

    # 👇 التقاط تنسيقات تيليجرام (روابط مخفية، بولد، إلخ)
    entities = None
    if message.text:
        entities = message.entities or []
    elif message.caption:
        entities = message.caption_entities or []

    # التقاط وسيط واحد من هذه الرسالة
    media_item = None
    if message.photo:
        media_item = {"type": "photo", "file_id": message.photo[-1].file_id}
    elif message.document:
        media_item = {"type": "document", "file_id": message.document.file_id}
    elif message.video:
        media_item = {"type": "video", "file_id": message.video.file_id}
    elif message.voice:
        media_item = {"type": "voice", "file_id": message.voice.file_id}

    if not incoming_text and not media_item:
        await message.reply_text("⚠️ لا يمكن حفظ توصية فارغة اكتب نص التوصية أو أرفق وسائط معها.")
        return

    # ✅ دعم أكثر من وسيط: نجمع كل ما يصل من المشرف
    existing_media = ud.get("reco_media") or []
    if isinstance(existing_media, dict):
        existing_media = [existing_media]
    elif not isinstance(existing_media, list):
        existing_media = []

    if media_item:
        existing_media.append(media_item)

    # النص: لو وصل نص جديد نحدّثه، ولو كانت الرسالة وسائط فقط نحافظ على النص السابق إن وجد
    if incoming_text:
        ud["reco_text"] = incoming_text
    else:
        ud.setdefault("reco_text", "")

    ud["reco_media"] = existing_media if existing_media else None
    ud["reco_entities"] = entities  # خزن تنسيقات تيليجرام (تشمل الرابط المخفي)
    # قيمة افتراضية لخيار التثبيت (غير مفعّل)
    ud.setdefault("reco_pin", False)

    # إعادة تجهيز قائمة المجموعات في حال استجدت مجموعات جديدة
    _prepare_reco_targets_for_admin(admin_id, context)

    admin_name = update.effective_user.full_name
    text_for_preview = ud.get("reco_text") or ""
    media_for_preview = ud.get("reco_media") or []

    # 🧾 ملخص الوسائط
    photos_count = sum(1 for m in media_for_preview if m.get("type") == "photo")
    videos_count = sum(1 for m in media_for_preview if m.get("type") == "video")
    docs_count   = sum(1 for m in media_for_preview if m.get("type") == "document")
    voices_count = sum(1 for m in media_for_preview if m.get("type") == "voice")

    media_summary_lines = []

    if photos_count or videos_count or docs_count or voices_count:
        media_summary_lines.append("🧾 *ملخص الوسائط المرفقة:*")
        if photos_count:
            media_summary_lines.append(f"• عدد الصور: `{photos_count}`")
        if videos_count:
            media_summary_lines.append(f"• عدد المقاطع المرئية: `{videos_count}`")
        if docs_count:
            media_summary_lines.append(f"• عدد الملفات المرفقة: `{docs_count}`")
        if voices_count:
            media_summary_lines.append(f"• عدد المقاطع الصوتية: `{voices_count}`")
    else:
        media_summary_lines.append("🧾 *لا توجد وسائط مرفقة حالياً.*")

    text_status = "نص التوصية موجود" if text_for_preview else "التوصية بدون نص (وسائط فقط)"
    media_summary_lines.append(f"✏️ حالة النص: `{text_status}`")

    media_summary_block = "\n".join(media_summary_lines)

    preview_caption = (
        "📡 *معاينة التوصية الفنية قبل الإرسال*\n\n"
        f"👤 *الناشر:* `{admin_name}`\n\n"
        "📄 *نص التوصية:*\n"
        f"```{text_for_preview or 'بدون نص صريح (الوسائط فقط) '}```\n\n"
        f"{media_summary_block}\n\n"
        "*فريق الصيانة والدعم الفني GO*\n\n"
        "✅ الخطوة التالية:\n"
        "• اختر المجموعات المستهدفة.\n"
        "• فعّل/أوقف تثبيت التوصية.\n"
        "• ثم نفّذ البث من نفس شاشة الاختيار.\n"
    )

    # 👇 من هنا ما عاد في بث من المعاينة مباشرة
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🧾 اختيار المجموعات والتثبيت", callback_data="reco_select")],
        [InlineKeyboardButton("❌ إلغاء التوصية", callback_data="reco_cancel")],
    ])

    # 🧹 حذف معاينة سابقة إن وجدت حتى لا تتكرر
    old_preview_id = ud.get("reco_preview_msg_id")
    if old_preview_id:
        try:
            await context.bot.delete_message(chat_id=message.chat_id, message_id=old_preview_id)
        except Exception:
            pass

    sent_preview = await message.reply_text(
        preview_caption,
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=keyboard,
    )

    # حفظ رقم رسالة المعاينة الحالية
    ud["reco_preview_msg_id"] = sent_preview.message_id

def build_reco_groups_keyboard(admin_id: int, context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    """يبني كيبورد اختيار المجموعات مع ترقيم الصفحات + خيار تثبيت التوصية"""

    ud = context.user_data.setdefault(admin_id, {})

    groups = ud.get("reco_targets", []) or []
    selected = set(ud.get("reco_selected") or [])
    page = ud.get("reco_page", 0) or 0
    page_size = 5

    # 🔒 حالة التثبيت
    pin_enabled = bool(ud.get("reco_pin", False))

    rows: list[list[InlineKeyboardButton]] = []

    # ============================================
    # 🛑 حماية: لا توجد مجموعات
    # ============================================
    if not groups:
        rows.append(
            [InlineKeyboardButton("⚠️ لا توجد مجموعات متاحة للبث", callback_data="reco_noop")]
        )
        rows.append([InlineKeyboardButton("⬅️ رجوع", callback_data="reco_cancel")])
        return InlineKeyboardMarkup(rows)

    # ============================================
    # ✂️ تجهيز الصفحة المطلوبة
    # ============================================
    start = page * page_size
    end = start + page_size
    slice_groups = groups[start:end]

    # ============================================
    # 📋 بناء مجموعة الأزرار (المجموعات)
    # ============================================
    for g in slice_groups:
        cid = g.get("id")
        title = g.get("title", "غير معروف")

        # ✂️ تنظيف الاسم (بعض الأسماء طويلة جداً)
        if len(title) > 28:
            title = title[:28] + "…"

        prefix = "✅" if cid in selected else "⬜"

        rows.append([
            InlineKeyboardButton(f"{prefix} {title}", callback_data=f"reco_tgl_{cid}")
        ])

    # ============================================
    # 🔁 أزرار التنقل
    # ============================================
    nav_row = []
    max_page = max((len(groups) - 1) // page_size, 0)

    if page > 0:
        nav_row.append(InlineKeyboardButton("⬅️ السابق", callback_data="reco_page_prev"))

    if page < max_page:
        nav_row.append(InlineKeyboardButton("التالي ➡️", callback_data="reco_page_next"))

    if nav_row:
        rows.append(nav_row)

    # ============================================
    # 📌 زر التثبيت
    # ============================================
    pin_label = "📌 تثبيت التوصية: مفعّل" if pin_enabled else "📌 تثبيت التوصية: غير مفعّل"
    rows.append([InlineKeyboardButton(pin_label, callback_data="reco_pin_toggle")])

    # ============================================
    # 📡 أزرار البث
    # ============================================
    rows.append([InlineKeyboardButton("📡 بث على المجموعات المحددة", callback_data="reco_broadcast")])
    rows.append([InlineKeyboardButton("📡 بث على جميع المجموعات", callback_data="reco_broadcast_all")])
    rows.append([InlineKeyboardButton("❌ إلغاء التوصية", callback_data="reco_cancel")])

    return InlineKeyboardMarkup(rows)

async def show_reco_groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """فتح قائمة اختيار المجموعات يدوياً"""
    query = update.callback_query
    admin_id = query.from_user.id

    # 🔐 التحقق من صلاحية المشرف
    if admin_id not in AUTHORIZED_USERS:
        await query.answer("هذه الميزة متاحة لمشرفي نظام GO فقط.", show_alert=True)
        return

    # ============================================
    # 🔥 حماية مهمة: Reload group_logs → BROADCAST_GROUPS
    # ============================================
    global BROADCAST_GROUPS, df_group_logs

    if not BROADCAST_GROUPS:
        try:
            # لو فاضي → إعادة تحميل من df_group_logs (من الإكسل)
            for _, row in df_group_logs.iterrows():
                gid = int(row.get("chat_id"))
                title = str(row.get("title", "غير معروف"))
                gtype = str(row.get("type", "group"))

                BROADCAST_GROUPS[gid] = {
                    "title": title,
                    "type": gtype,
                }
            logging.info(f"[RECO INIT] تمت إعادة بناء BROADCAST_GROUPS من الإكسل. مجموع: {len(BROADCAST_GROUPS)}")
        except Exception as e:
            logging.error(f"[RECO INIT ERROR] {e}")

    # ============================================
    # 🌐 تجهيز بيانات المجموعات للمشرف
    # ============================================
    _prepare_reco_targets_for_admin(admin_id, context)

    # تأكد من وجود الحقول في user_data
    ud = context.user_data.setdefault(admin_id, {})
    ud["reco_page"] = 0

    # نص واجهة التحكم
    text_lines = [
        "📡 *اختيار المجموعات المستهدفة بالتوصية:*",
        "",
        "• اضغط على اسم المجموعة لتفعيل / إلغاء التحديد.",
        "• يمكنك التنقل بين الصفحات (⬅️ السابق / التالي ➡️).",
        "• عند الانتهاء اضغط «📡 بث على المجموعات المحددة».",
    ]

    # بناء لوحة المجموعات
    keyboard = build_reco_groups_keyboard(admin_id, context)

    # إرسال الرسالة
    await query.message.reply_text(
        "\n".join(text_lines),
        reply_markup=keyboard,
        parse_mode="MARKDOWN"
    )
    await query.answer()

async def toggle_reco_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تفعيل/إلغاء مجموعة من قائمة الاختيار"""
    query = update.callback_query
    admin_id = query.from_user.id

    if admin_id not in AUTHORIZED_USERS:
        await query.answer("غير مصرح.", show_alert=True)
        return

    data = (query.data or "").split("_", 2)
    if len(data) < 3:
        await query.answer()
        return

    try:
        cid = int(data[2])
    except ValueError:
        await query.answer()
        return

    ud = context.user_data.setdefault(admin_id, {})
    selected = set(ud.get("reco_selected") or [])

    if cid in selected:
        selected.remove(cid)
    else:
        selected.add(cid)

    ud["reco_selected"] = list(selected)

    # تحديث الكيبورد فقط
    keyboard = build_reco_groups_keyboard(admin_id, context)
    try:
        await query.edit_message_reply_markup(reply_markup=keyboard)
    except Exception:
        pass

    await query.answer("تم تحديث الاختيار.")


async def change_reco_page(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """التنقل بين صفحات المجموعات"""
    query = update.callback_query
    admin_id = query.from_user.id

    if admin_id not in AUTHORIZED_USERS:
        await query.answer("غير مصرح.", show_alert=True)
        return

    ud = context.user_data.setdefault(admin_id, {})
    groups = ud.get("reco_targets", []) or []
    if not groups:
        await query.answer("لا توجد مجموعات.", show_alert=False)
        return

    page = ud.get("reco_page", 0) or 0
    action = (query.data or "").split("_")[-1]
    page_size = 5
    max_page = max((len(groups) - 1) // page_size, 0)

    if action == "next" and page < max_page:
        ud["reco_page"] = page + 1
    elif action == "prev" and page > 0:
        ud["reco_page"] = page - 1
    else:
        await query.answer("لا توجد صفحات أخرى.", show_alert=False)
        return

    keyboard = build_reco_groups_keyboard(admin_id, context)
    try:
        await query.edit_message_reply_markup(reply_markup=keyboard)
    except Exception:
        pass

    await query.answer()

# ================================================================
# 🧩 STEP 3 — تجميع المجموعات المستهدفة للبث من الذاكرة + الإكسل
# ================================================================
def collect_target_chat_ids(context: ContextTypes.DEFAULT_TYPE) -> list[int]:
    """يعيد قائمة جميع المجموعات المخزنة — سواء من الإكسل أو آخر جلسة"""
    targets = set()

    # 1️⃣ من القوائم المستعادة من Excel (group_logs)
    try:
        if "group_logs" in globals() and not df_group_logs.empty:
            for _, row in df_group_logs.iterrows():
                cid = int(row.get("chat_id", 0))
                if cid < 0:
                    targets.add(cid)
    except Exception as e:
        logging.warning(f"[TARGET GROUPS] خطأ في قراءة شيت المجموعة: {e}")

    # 2️⃣ المجموعات المحفوظة مسبقاً داخل الذاكرة (BROADCAST_GROUPS)
    try:
        global BROADCAST_GROUPS
        for gid in BROADCAST_GROUPS.keys():
            if int(gid) < 0:
                targets.add(int(gid))
    except Exception:
        pass

    # 3️⃣ القروبات النشطة التي اكتشفها البوت خلال الجلسات
    for key, data in context.bot_data.items():
        if isinstance(data, dict) and "group_id" in data:
            gid = data.get("group_id")
            if gid and gid < 0:
                targets.add(gid)

    return list(targets)

async def broadcast_recommendation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """بث التوصية على المجموعات (الكل أو المحدد فقط) + خيار تثبيت الرسالة + إشعار المشرفين"""
    query = update.callback_query
    admin_id = query.from_user.id
    admin_name = query.from_user.full_name

    if admin_id not in AUTHORIZED_USERS:
        await query.answer("هذه الميزة متاحة لمشرفي نظام GO فقط.", show_alert=True)
        return

    ud = context.user_data.setdefault(admin_id, {})
    text = ud.get("reco_text")
    media = ud.get("reco_media")
    pin_enabled = bool(ud.get("reco_pin", False))

    # ✅ تحويل reco_media إلى قائمة وسائط موحدة
    media_list = []
    if isinstance(media, list):
        media_list = media
    elif isinstance(media, dict):
        media_list = [media]

    if not text and not media_list:
        await query.answer("لا توجد توصية جاهزة للبث. يرجى إرسال التوصية أولاً.", show_alert=True)
        return

    data = query.data or ""
    selected_ids = ud.get("reco_selected") or []

    # 🎯 تحديد المجموعات المستهدفة
    if data == "reco_broadcast_all":
        # بث لجميع المجموعات المتاحة
        targets = collect_target_chat_ids(context)
    elif data == "reco_broadcast":
        # بث للمجموعات المحددة فقط – منع لو ما فيه ولا مجموعة
        if not selected_ids:
            await query.answer(
                "فضلاً حدد مجموعة واحدة على الأقل من «اختيار المجموعات والتثبيت» قبل البث.",
                show_alert=True,
            )
            return
        targets = selected_ids
    else:
        # احتياط
        targets = collect_target_chat_ids(context)

    if not targets:
        await query.answer("لا توجد مجموعات متاحة للبث حالياً.", show_alert=True)
        return

    await query.answer("📡 جاري بث التوصية على المجموعات...", show_alert=False)

    # 🆕 تجهيز نص التوصية بصيغة HTML + رابط "اضغط هنا لعرض التفاصيل" إن وجد
    raw_text = text or ""
    url_match = None
    if raw_text:
        url_match = re.search(r"(https?://\S+)", raw_text)

    html_body = ""
    if raw_text:
        if url_match:
            url = url_match.group(1)
            # إزالة الرابط الخام من النص الظاهر
            cleaned = raw_text.replace(url, "").strip()
            html_body = html.escape(cleaned)
        else:
            html_body = html.escape(raw_text)

    # 🆕 إضافة تذييل ثابت في أسفل التوصية
    footer = "فريق الصيانة والدعم الفني GO"
    if html_body:
        html_body += "\n\n" + html.escape(footer)
    else:
        html_body = html.escape(footer)

    html_text = html_body
    if url_match:
        url = url_match.group(1).strip()
        safe_url = html.escape(url, quote=True)
        if html_text:
            html_text += "\n\n"
        html_text += f"🔗 <a href=\"{safe_url}\">اضغط هنا لعرض التفاصيل</a>"

    sent = failed = skipped = 0

    # ترتيب الأنواع للألبوم: فيديو ثم صورة ثم ملف
    type_order = {"video": 0, "photo": 1, "document": 2}

    for chat_id in targets:
        try:
            # تأكد أن البوت مشرف في المجموعة
            member = await context.bot.get_chat_member(chat_id, context.bot.id)
            if member.status not in ("administrator", "creator"):
                skipped += 1
                continue

            sent_msg = None

            if media_list:
                try:
                    # نفصل بين الوسائط الصوتية وغيرها
                    non_voice_media = [m for m in media_list if m.get("type") != "voice"]
                    voice_media = [m for m in media_list if m.get("type") == "voice"]

                    # 🧩 حالة وجود وسائط غير صوتية وصوت معاً
                    if non_voice_media and voice_media:
                        # 🔢 ترتيب غير الصوتية: فيديو → صور → مستندات
                        non_voice_media = sorted(
                            non_voice_media,
                            key=lambda m: type_order.get(m.get("type"), 3)
                        )

                        album = []
                        for m in non_voice_media:
                            mtype = m.get("type")
                            fid = m.get("file_id")
                            if not mtype or not fid:
                                continue

                            if mtype == "photo":
                                album.append(InputMediaPhoto(media=fid))
                            elif mtype == "video":
                                album.append(InputMediaVideo(media=fid))
                            elif mtype == "document":
                                album.append(InputMediaDocument(media=fid))

                        album_msgs = []
                        if album:
                            # 🖼️ كل الوسائط غير الصوتية برسالة واحدة (ألبوم) بدون كابتشن
                            album_msgs = await context.bot.send_media_group(chat_id, album)

                        # 🎧 نرسل أول ملف صوتي مع نص التوصية الكامل + التذييل
                        main_voice = voice_media[0]
                        extra_voices = voice_media[1:]
                        voice_msg = None
                        vf = main_voice.get("file_id")
                        if vf:
                            voice_msg = await context.bot.send_voice(
                                chat_id,
                                vf,
                                caption=html_text or "",
                                parse_mode=constants.ParseMode.HTML,
                            )

                        # أي أصوات إضافية بدون كابتشن
                        for v in extra_voices:
                            try:
                                vf2 = v.get("file_id")
                                if vf2:
                                    await context.bot.send_voice(chat_id, vf2)
                            except Exception as e2:
                                logging.warning(f"[RECO BROADCAST] فشل إرسال voice إضافي إلى {chat_id}: {e2}")

                        # الرسالة التي يمكن تثبيتها: نفضّل رسالة الصوت + النص
                        sent_msg = voice_msg or (album_msgs[0] if album_msgs else None)

                    # 🧩 حالة وجود وسائط غير صوتية فقط (بدون صوت)
                    elif non_voice_media and not voice_media:
                        non_voice_media = sorted(
                            non_voice_media,
                            key=lambda m: type_order.get(m.get("type"), 3)
                        )

                        album = []
                        for idx, m in enumerate(non_voice_media):
                            mtype = m.get("type")
                            fid = m.get("file_id")
                            if not mtype or not fid:
                                continue

                            # أول وسيط فقط نضع معه الكابتشن (النص + التذييل + الرابط)
                            if mtype == "photo":
                                if idx == 0:
                                    album.append(
                                        InputMediaPhoto(
                                            media=fid,
                                            caption=html_text or "",
                                            parse_mode=constants.ParseMode.HTML,
                                        )
                                    )
                                else:
                                    album.append(InputMediaPhoto(media=fid))
                            elif mtype == "video":
                                if idx == 0:
                                    album.append(
                                        InputMediaVideo(
                                            media=fid,
                                            caption=html_text or "",
                                            parse_mode=constants.ParseMode.HTML,
                                        )
                                    )
                                else:
                                    album.append(InputMediaVideo(media=fid))
                            elif mtype == "document":
                                if idx == 0:
                                    album.append(
                                        InputMediaDocument(
                                            media=fid,
                                            caption=html_text or "",
                                            parse_mode=constants.ParseMode.HTML,
                                        )
                                    )
                                else:
                                    album.append(InputMediaDocument(media=fid))

                        album_msgs = []
                        if album:
                            album_msgs = await context.bot.send_media_group(chat_id, album)

                        sent_msg = album_msgs[0] if album_msgs else None

                    # 🧩 حالة الصوت فقط بدون أي وسائط أخرى
                    elif voice_media and not non_voice_media:
                        main_voice = voice_media[0]
                        extra_voices = voice_media[1:]

                        first_voice_msg = None
                        fid = main_voice.get("file_id")
                        if fid:
                            first_voice_msg = await context.bot.send_voice(
                                chat_id,
                                fid,
                                caption=html_text or "",
                                parse_mode=constants.ParseMode.HTML,
                            )

                        for v in extra_voices:
                            try:
                                vf = v.get("file_id")
                                if vf:
                                    await context.bot.send_voice(chat_id, vf)
                            except Exception as e2:
                                logging.warning(f"[RECO BROADCAST] فشل إرسال voice إضافي إلى {chat_id}: {e2}")

                        sent_msg = first_voice_msg

                except Exception as e:
                    logging.warning(f"[RECO BROADCAST] خطأ أثناء إرسال الوسائط المتعددة إلى {chat_id}: {e}")
                    # في حالة أي خطأ نرجع للخطة البسيطة: نص فقط
                    sent_msg = await context.bot.send_message(
                        chat_id,
                        html_text or "",
                        parse_mode=constants.ParseMode.HTML,
                        disable_web_page_preview=True,
                    )
            else:
                # 🆕 لا توجد وسائط → نحاول إرسال صورة GO-NOW.PNG مع نفس النص
                try:
                    with open("GO-NOW.PNG", "rb") as f:
                        sent_msg = await context.bot.send_photo(
                            chat_id,
                            f,
                            caption=html_text or "",
                            parse_mode=constants.ParseMode.HTML,
                        )
                except Exception as e:
                    logging.warning(f"[RECO BROADCAST] تعذر إرسال صورة GO-NOW.PNG إلى {chat_id}: {e}")
                    # في حال فشل تحميل الصورة نرجع لإرسال التوصية كنص HTML فقط
                    sent_msg = await context.bot.send_message(
                        chat_id,
                        html_text or "",
                        parse_mode=constants.ParseMode.HTML,
                        disable_web_page_preview=True,
                    )

            # 📌 تثبيت الرسالة إن كان الخيار مفعّل
            if pin_enabled and sent_msg is not None:
                try:
                    await context.bot.pin_chat_message(
                        chat_id=chat_id,
                        message_id=sent_msg.message_id,
                        disable_notification=True,
                    )
                except BadRequest as e:
                    # غالباً لأن البوت لا يملك صلاحية التثبيت – نتجاهل بدون إيقاف البث
                    logging.warning(f"[RECO PIN] تعذر تثبيت الرسالة في {chat_id}: {e}")
                except Exception as e:
                    logging.warning(f"[RECO PIN] خطأ غير متوقع أثناء التثبيت في {chat_id}: {e}")

            sent += 1
        except Exception as e:
            logging.warning(f"[RECO BROADCAST] فشل إرسال التوصية إلى {chat_id}: {e}")
            failed += 1

    # ملخص للمشرف الناشر
    summary = (
        "📡 تمت عملية بث التوصية الفنية.\n\n"
        f"✅ تم الإرسال إلى: {sent} مجموعة\n"
        f"⏭️ تم التخطي في: {skipped} مجموعة (البوت ليس مشرفاً)\n"
        f"⚠️ فشل الإرسال في: {failed} مجموعة\n\n"
        f"📌 خيار التثبيت كان: {'مفعّل' if pin_enabled else 'غير مفعّل'}"
    )
    try:
        await query.message.reply_text(summary)
    except Exception:
        pass

    # إشعار جميع المشرفين (بدون أرقام تعريفية)
    group_title = ud.get("group_title", "—")

    admin_notification_caption = (
        "📡 تمت عملية بث توصية فنية جديدة.\n\n"
        f"👤 الناشر:\n`{admin_name}`\n\n"
        f"👥 المجموعة التابعة له:\n`{group_title}`\n\n"
        "📊 ملخص البث:\n"
        f"✅ تم الإرسال إلى: `{sent}` مجموعة\n"
        f"⏭️ تم التخطي في: `{skipped}` مجموعة (البوت ليس مشرفاً)\n"
        f"⚠️ فشل الإرسال في: `{failed}` مجموعة\n"
        f"📌 خيار التثبيت: `{'مفعّل' if pin_enabled else 'غير مفعّل'}`\n\n"
        "📄 نص التوصية:\n"
        f"{text or '— التوصية بدون نص (وسائط فقط) —'}"
    )

    # نستخدم وسيط واحد فقط لمعاينة المشرفين
    notify_media = None
    non_voice_media = [m for m in media_list if m.get("type") != "voice"]
    voice_media = [m for m in media_list if m.get("type") == "voice"]
    if non_voice_media:
        non_voice_media = sorted(
            non_voice_media,
            key=lambda m: type_order.get(m.get("type"), 3)
        )
        notify_media = non_voice_media[0]
    elif voice_media:
        notify_media = voice_media[0]

    for aid in AUTHORIZED_USERS:
        try:
            if notify_media:
                mtype = notify_media.get("type")
                fid = notify_media.get("file_id")
                if mtype == "photo":
                    await context.bot.send_photo(aid, fid, caption=admin_notification_caption)
                elif mtype == "video":
                    await context.bot.send_video(aid, fid, caption=admin_notification_caption)
                elif mtype == "document":
                    await context.bot.send_document(aid, fid, caption=admin_notification_caption)
                elif mtype == "voice":
                    await context.bot.send_voice(aid, fid, caption=admin_notification_caption)
            else:
                await context.bot.send_message(aid, admin_notification_caption)
        except Exception as e:
            logging.warning(f"[RECO NOTIFY ADMIN] فشل إشعار المشرف {aid}: {e}")

    # 🧹 تنظيف بيانات التوصية من user_data بعد الانتهاء
    ud.pop("reco_text", None)
    ud.pop("reco_media", None)
    ud.pop("reco_entities", None)
    ud.pop("reco_selected", None)
    ud.pop("reco_pin", None)
    # يمكنك أيضاً إعادة وضع reco_mode لو تحب:
    # ud["reco_mode"] = None

async def cancel_recommendation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """إلغاء وضع التوصية والرجوع للقائمة الرئيسية"""
    query = update.callback_query
    admin_id = query.from_user.id
    ud = context.user_data.setdefault(admin_id, {})

    # تصفير حالة التوصية
    ud["reco_mode"] = None
    ud["reco_text"] = None
    ud["reco_media"] = None

    await query.answer("تم إلغاء التوصية.", show_alert=False)

    # إخفاء رسالة التوصية / المعاينة من الشات
    try:
        await query.message.delete()
    except Exception:
        pass

    # رجوع للقائمة الرئيسية في الخاص
    try:
        keyboard = build_main_menu_keyboard(admin_id)
        msg = await context.bot.send_message(
            chat_id=admin_id,
            text="✅ تم إلغاء التوصية الفنية.\nفضلاً اختر الخدمة المطلوبة 🛠️ :",
            reply_markup=keyboard
        )
        register_message(admin_id, msg.message_id, admin_id, context)
    except Exception as e:
        logging.warning(f"[RECO CANCEL] فشل إرسال القائمة الرئيسية بعد الإلغاء للمشرف {admin_id}: {e}")

async def toggle_reco_pin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تفعيل/إلغاء خيار تثبيت التوصية من شاشة اختيار المجموعات"""
    query = update.callback_query
    admin_id = query.from_user.id

    if admin_id not in AUTHORIZED_USERS:
        await query.answer("غير مصرح.", show_alert=True)
        return

    ud = context.user_data.setdefault(admin_id, {})
    current = bool(ud.get("reco_pin", False))
    ud["reco_pin"] = not current  # عكس الحالة

    # إعادة بناء الكيبورد بنفس الصفحة الحالية
    keyboard = build_reco_groups_keyboard(admin_id, context)
    try:
        await query.edit_message_reply_markup(reply_markup=keyboard)
    except Exception:
        pass

    status = "مفعّل ✅" if ud["reco_pin"] else "غير مفعّل ❌"
    await query.answer(f"خيار تثبيت التوصية الآن: {status}", show_alert=False)

### ✅ الدالة المعدلة: handle_message (فقط جزء الاقتراح)
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global df_admins

    # ✅ تسجيل أي قروب يرسل فيه أي رسالة (بدون أمر go)
    chat = update.effective_chat
    if chat and chat.type in ("group", "supergroup"):
        asyncio.create_task(
            update_group_logs(chat.id, chat.title or "غير معروف", context)
        )

    message = update.message
    user = update.effective_user
    admin_id = user.id
    user_id = user.id
    user_name = user.full_name

    # 🔒 ضمان وجود قواميس للمستخدم/المشرف قبل الكتابة عليها
    context.user_data.setdefault(admin_id, {})
    context.user_data.setdefault(user_id, {})

    # 📨 إذا كان المشرف في وضع نقاش داخلي لفريق GO نوجّه الرسالة هناك
    if context.user_data[admin_id].get("team_mode"):
        await handle_team_message(update, context)
        return

    action = context.user_data.get(user_id, {}).get("action")

    # ✅ حذف مشرف
    if action == "awaiting_admin_removal":
        try:
            target_id = int(message.text.strip())
            if target_id == 1543083749:
                await message.reply_text("🚫 لا يمكن حذف المدير الأعلى.")
                return
            if target_id not in df_admins["manager_id"].astype(int).values:
                await message.reply_text("❌ هذا المعرف غير موجود في قائمة المشرفين.")
                return

            df_admins = df_admins[df_admins["manager_id"].astype(int) != target_id]
            if target_id in AUTHORIZED_USERS:
                AUTHORIZED_USERS.remove(target_id)

            # قفل الكتابة على ملف الإكسل قبل تعديل شيت managers
            async with EXCEL_LOCK:
                with pd.ExcelWriter("bot_data.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_admins.to_excel(writer, sheet_name="managers", index=False)

            await message.reply_text(f"🗑️ تم حذف المشرف بنجاح:\n<code>{target_id}</code>", parse_mode="HTML")
        except Exception as e:
            await message.reply_text(f"❌ حدث خطأ أثناء حذف المشرف:\n<code>{e}</code>", parse_mode="HTML")
        context.user_data[admin_id]["action"] = None
        return

    # ✅ إضافة مشرف
    if action == "awaiting_new_admin_id":
        try:
            text = message.text.strip()
            if not text.isdigit():
                await message.reply_text("❌ يجب إدخال رقم ID رقمي صالح.")
                return
            new_admin_id = int(text)
            if new_admin_id in AUTHORIZED_USERS:
                await message.reply_text("ℹ️ هذا المشرف موجود مسبقًا.")
                return

            AUTHORIZED_USERS.append(new_admin_id)
            df_admins = pd.concat([df_admins, pd.DataFrame([{"manager_id": new_admin_id}])], ignore_index=True)
            # قفل الكتابة على ملف الإكسل قبل تعديل شيت managers
            async with EXCEL_LOCK:
                with pd.ExcelWriter("bot_data.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_admins.to_excel(writer, sheet_name="managers", index=False)

            await message.reply_text(f"✅ تم إضافة المشرف:\n<code>{new_admin_id}</code>", parse_mode="HTML")
        except Exception as e:
            await message.reply_text(f"❌ فشل أثناء حفظ الملف:\n<code>{e}</code>", parse_mode="HTML")
        context.user_data[admin_id]["action"] = None
        return

    # ملاحظة: لا نتدخل في التوصية إذا كان المستخدم في وضع آخر نشط (بحث قطع، دعم فني، ...).
    if admin_id in AUTHORIZED_USERS:
        udata = context.user_data.get(admin_id, {})
        reco_mode = udata.get("reco_mode")
        compose_mode = udata.get("compose_mode")

        # حالة المستخدم الفعلية (لخدمات مثل قطع الغيار، الدعم الفني، ...).
        user_state = context.user_data.get(user_id, {}) or {}
        action = user_state.get("action")

        # ✅ نسمح باستقبال نص التوصية فقط إذا:
        #    - المشرف في وضع التوصية awaiting_reco
        #    - المحادثة خاصة
        #    - وما فيه أي action آخر نشط ولا compose_mode شغال
        if (
            reco_mode == "awaiting_reco"
            and chat.type == "private"
            and not action
            and not compose_mode
        ):
            await handle_recommendation_message(update, context)
            return

    # ✅ حالات الاستفسار والرد المخصص
    admin_state = context.user_data.get(admin_id, {}) or {}

    # ✅ تحديد الحالة: هل هو مستخدم يرسل استفسار؟ أو مشرف يكتب رد مخصص؟
    # - المستخدم: action داخل user_data[user_id]
    # - المشرف: compose_mode داخل user_data[admin_id]
    user_mode = context.user_data.get(user_id, {}).get("action")
    admin_mode = admin_state.get("compose_mode")

    # 1) المستخدم يكتب استفسار (suggestion)
    # =========================
    if user_mode == "suggestion" and chat.type == "private":
        actual_user_id = user_id  # ✅ ثابت: المستخدم نفسه
        context.user_data.setdefault(actual_user_id, {})

        suggestion_id = context.user_data[actual_user_id].get("active_suggestion_id")
        if not suggestion_id:
            suggestion_id = await start_suggestion_session(actual_user_id, context)

        # ✅ منع KeyError
        record = suggestion_records.get(actual_user_id, {}).get(suggestion_id)
        if not record:
            suggestion_id = await start_suggestion_session(actual_user_id, context)
            record = suggestion_records[actual_user_id][suggestion_id]

        if not context.user_data[actual_user_id].get("compose_text") and not context.user_data[actual_user_id].get("compose_media"):
            record["text"] = ""
            record["media"] = None

        group_name = chat.title if chat.type in ["group", "supergroup"] else "خاص"
        group_id = chat.id
        if group_name == "خاص" or group_id == actual_user_id:
            fallback = context.user_data.get(actual_user_id, {}) or context.bot_data.get(actual_user_id, {})
            group_name = fallback.get("group_title", "غير معروف")
            group_id = fallback.get("group_id", actual_user_id)

        record["group_name"] = group_name
        record["group_id"] = group_id
        context.user_data[actual_user_id]["compose_mode"] = "suggestion"

        if message.text:
            context.user_data[actual_user_id]["compose_text"] = message.text.strip()
            record["text"] = message.text.strip()

        elif message.photo or message.video or message.document or message.voice:
            if message.photo:
                file_id = message.photo[-1].file_id
                media_type = "photo"
            elif message.video:
                file_id = message.video.file_id
                media_type = "video"
            elif message.document:
                file_id = message.document.file_id
                media_type = "document"
            elif message.voice:
                file_id = message.voice.file_id
                media_type = "voice"
            context.user_data[actual_user_id]["compose_media"] = {"type": media_type, "file_id": file_id}
            record["media"] = {"type": media_type, "file_id": file_id}

        buttons = [
            [InlineKeyboardButton("📤 إرسال", callback_data="send_suggestion")],
            [InlineKeyboardButton("❌ إلغاء", callback_data="cancel_suggestion")]
        ]

        has_text = context.user_data[actual_user_id].get("compose_text")
        has_media = context.user_data[actual_user_id].get("compose_media")

        if has_text and has_media:
            await message.reply_text("✅ تم حفظ النص والوسائط. يمكنك الإرسال الآن:", reply_markup=InlineKeyboardMarkup(buttons))
        elif has_text:
            await message.reply_text("📎 لقد قمت بادخال النص بنجاج . يمكنك الآن إدخال وسائط أو الإرسال:", reply_markup=InlineKeyboardMarkup(buttons))
        elif has_media:
            await message.reply_text("🖼️ لقد قمت بادخال الوسائط بنجاح . يمكنك الآن إدخال نص أو الإرسال:", reply_markup=InlineKeyboardMarkup(buttons))
        else:
            await message.reply_text("⚠️ لم يتم تسجيل أي محتوى. الرجاء إدخال نص أو وسائط.")
        return

    # =========================
    # 2) المشرف يكتب رد مخصص (custom_reply)
    # =========================
    if admin_mode == "custom_reply" and chat.type == "private":
        actual_user_id = admin_state.get("custom_reply_for", admin_id)

        # ✅ المهم: التذكرة تُقرأ من بيانات المشرف وليس من المستخدم
        suggestion_id = admin_state.get("active_suggestion_id")
        if not suggestion_id:
            await message.reply_text("⚠️ لا توجد تذكرة محددة للرد عليها. افتح التذكرة من زر (الرد على الاستفسار).")
            return

        # ✅ منع KeyError
        record = suggestion_records.get(actual_user_id, {}).get(suggestion_id)
        if not record:
            await message.reply_text("⚠️ التذكرة غير موجودة أو تم تنظيفها. افتح التذكرة مجددًا من القائمة.")
            return

        if not context.user_data[admin_id].get("compose_text") and not context.user_data[admin_id].get("compose_media"):
            record["reply_text"] = ""
            record["reply_media"] = None

        group_name = chat.title if chat.type in ["group", "supergroup"] else "خاص"
        group_id = chat.id
        if group_name == "خاص" or group_id == actual_user_id:
            fallback = context.user_data.get(actual_user_id, {}) or context.bot_data.get(actual_user_id, {})
            group_name = fallback.get("group_title", "غير معروف")
            group_id = fallback.get("group_id", actual_user_id)

        record["group_name"] = group_name
        record["group_id"] = group_id
        context.user_data[admin_id]["compose_mode"] = "custom_reply"

        if message.text:
            context.user_data[admin_id]["compose_text"] = message.text.strip()
            record["reply_text"] = message.text.strip()

        elif message.photo or message.video or message.document or message.voice:
            if message.photo:
                file_id = message.photo[-1].file_id
                media_type = "photo"
            elif message.video:
                file_id = message.video.file_id
                media_type = "video"
            elif message.document:
                file_id = message.document.file_id
                media_type = "document"
            elif message.voice:
                file_id = message.voice.file_id
                media_type = "voice"
            context.user_data[admin_id]["compose_media"] = {"type": media_type, "file_id": file_id}
            record["reply_media"] = {"type": media_type, "file_id": file_id}

        buttons = [
            [InlineKeyboardButton("📤 إرسال الرد", callback_data="submit_admin_reply")],
            [InlineKeyboardButton("❌ إلغاء", callback_data="cancel_custom_reply")]
        ]

        has_text = context.user_data[admin_id].get("compose_text")
        has_media = context.user_data[admin_id].get("compose_media")

        if has_text and has_media:
            await message.reply_text("✅ تم حفظ النص والوسائط. يمكنك الإرسال الآن:", reply_markup=InlineKeyboardMarkup(buttons))
        elif has_text:
            await message.reply_text("📎 لقد قمت بادخال النص بنجاج . يمكنك الآن إدخال وسائط أو الإرسال:", reply_markup=InlineKeyboardMarkup(buttons))
        elif has_media:
            await message.reply_text("🖼️ لقد قمت بادخال الوسائط بنجاح . يمكنك الآن إدخال نص أو الإرسال:", reply_markup=InlineKeyboardMarkup(buttons))
        else:
            await message.reply_text("⚠️ لم يتم تسجيل أي محتوى. الرجاء إدخال نص أو وسائط.")
        return

    # ✅ استعلام قطع الغيار بالنص
    if (
        context.user_data.get(user_id, {}).get("action") == "parts"
        and message.text
        and chat.type == "private"
        and context.user_data.get(user_id, {}).get("session_valid")
    ):
        # ✅ تسجيل رسالة المستخدم نفسها ليتم حذفها بعد 15 دقيقة
        register_message(user_id, message.message_id, chat.id, context)

        part_name = message.text.strip().lower()
        MAX_ATTEMPTS = 8
        current_attempts = context.user_data[user_id].get("search_attempts", 0)

        # ✅ تجاوز الحد الأقصى للمحاولات
        if current_attempts >= MAX_ATTEMPTS:
            msg = await message.reply_text(
                "🚫 لقد استهلكت جميع استعلامات البحث اليدوي (8 استعلامات).\n🔁 ابدأ من جديد باستخدام (go) من المجموعة."
            )
            register_message(user_id, msg.message_id, chat.id, context)
            context.user_data[user_id].clear()
            return

        # ✅ تحديث عداد المحاولات
        context.user_data[user_id]["search_attempts"] = current_attempts + 1
        remaining = MAX_ATTEMPTS - current_attempts - 1

        # ✅ رسالة توضح رقم الاستعلام المتبقي + جدولتها للحذف
        if remaining > 0:
            info_msg = await message.reply_text(
                f"🔁 تم تسجيل الاستعلام رقم {current_attempts + 1}.\nتبقى لك {remaining} من أصل {MAX_ATTEMPTS} استعلامات."
            )
            register_message(user_id, info_msg.message_id, chat.id, context)
        else:
            info_msg = await message.reply_text("⚠️ تبقى آخر استعلام مسموح لك خلال هذي الجلسة.")
            register_message(user_id, info_msg.message_id, chat.id, context)

        # ✅ جدولة تصفير عداد البحث اليدوي بعد 15 دقيقة من آخر استعلام
        if context.job_queue:
            try:
                context.job_queue.run_once(
                    reset_manual_search_state,
                    when=timedelta(minutes=15),
                    data={"user_id": user_id}
                )
            except Exception as e:
                logging.warning(f"[JOB ERROR] فشل في جدولة تصفير عداد البحث اليدوي للمستخدم {user_id}: {e}")

        selected_car = context.user_data[user_id].get("selected_car")
        if not selected_car:
            msg = await message.reply_text("❗ لم يتم اختيار فئة السيارة.")
            register_message(user_id, msg.message_id, chat.id, context)
            return

        filtered_df = df_parts[df_parts["Station No"] == selected_car]
        columns_to_search = ["Station Name", "Part No"]
        matches = filtered_df[
            filtered_df[columns_to_search].apply(
                lambda x: x.str.contains(part_name, case=False, na=False)
            ).any(axis=1)
        ]

        if matches.empty:
            msg = await message.reply_text("❌ لم يتم العثور على نتائج او الادخال خاطي.")
            register_message(user_id, msg.message_id, chat.id, context)
            return

        user_name = message.from_user.full_name
        user_name_safe = html.escape(user_name)
        selected_car_safe = html.escape(selected_car)
        part_name_safe = html.escape(part_name)
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

        header = (
            "🧑‍💼 استعلام خاص بـ: "
            f"<i>{user_name_safe}</i>\n"
            "🚗 فئة السيارة: "
            f"<i>{selected_car_safe}</i>\n\n"
        )

        results_header = (
            f"<b>📌 نتائج البحث عن:</b> <code>{part_name_safe}</code>\n"
        )

        lines = []
        for idx, (_, row) in enumerate(matches.iterrows(), start=1):
            station = html.escape(str(row.get("Station Name", "غير معروف")))
            part_no = html.escape(str(row.get("Part No", "غير معروف")))
            price = get_part_price(row)

            line_parts = [
                f"{idx}️⃣ <b>{station}</b>",
                f"   <code>رقم القطعة: {part_no}</code>",
            ]

            if price:
                price_disp = html.escape(str(price)).strip()
                if "ريال" not in price_disp and "SAR" not in price_disp.upper():
                    price_disp = f"{price_disp} ريال"
                line_parts.append(f"   <code>السعر التقريبي: {price_disp}</code>")

            lines.append("\n".join(line_parts))

        body = "\n\n".join(lines)

        # 💡 ملاحظة بدون span
        note_line = (
            "\n\n<i>💡 يمكن عرض صور قطع الغيار بشكل أوضح من خلال التصنيفات داخل خدمة قطع الغيار.</i>"
        )

        footer = (
            f"\n\n<code>⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة "
            f"({delete_time} / 🇸🇦)</code>"
        )

        text = header + results_header + body + note_line + footer

        keyboard_rows = []

        # ✅ زر "عرض القطع المصنفة" يفتح تصنيفات القطع لنفس الفئة المختارة
        safe_car = str(selected_car).replace(" ", "_")
        keyboard_rows.append(
            [InlineKeyboardButton("🗂 عرض القطع المصنفة", callback_data=f"showparts_{safe_car}_{user_id}")]
        )

        parts_brand = context.user_data[user_id].get("parts_brand")

        if parts_brand:
            safe_brand = parts_brand.replace(" ", "_")
            keyboard_rows.append(
                [InlineKeyboardButton("⬅️ رجوع لاختيار سيارة", callback_data=f"pbrand_{safe_brand}_{user_id}")]
            )

        keyboard_rows.append(
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        )

        msg = await message.reply_text(
            text,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(keyboard_rows),
        )
        register_message(user_id, msg.message_id, chat.id, context)
        return

async def handle_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id

    data = query.data or ""
    mode = context.user_data.get(user_id, {}).get("compose_mode")

    # ✅ إلغاء استفسار المستخدم (cancel_suggestion)
    if data == "cancel_suggestion":
        suggestion_records.pop(user_id, None)
        context.user_data.setdefault(user_id, {})
        context.user_data[user_id].clear()

        # تعديل الرسالة بدل مربع إشعار
        try:
            await query.edit_message_text("❌ تم إلغاء العملية.")
        except Exception:
            pass

    # ✅ إلغاء رد المشرف (cancel_custom_reply)
    elif data == "cancel_custom_reply":
        admin_id = user_id
        admin_state = context.user_data.get(admin_id, {}) or {}
        target_user_id = admin_state.get("custom_reply_for")
        suggestion_id = admin_state.get("active_suggestion_id")

        # ✅ فك القفل إن كانت التذكرة موجودة
        if target_user_id and suggestion_id:
            record = suggestion_records.get(target_user_id, {}).get(suggestion_id)
            if record:
                try:
                    unlock_ticket(record)
                except Exception:
                    pass

        # ✅ تنظيف حالة المشرف فقط (بدون حذف سجلات المستخدم)
        context.user_data.setdefault(admin_id, {})
        for k in ["compose_mode", "custom_reply_for", "active_suggestion_id", "compose_text", "compose_media", "ready_reply_text"]:
            context.user_data[admin_id].pop(k, None)

        # تعديل الرسالة بدل مربع إشعار
        try:
            await query.edit_message_text("❌ تم إلغاء العملية.")
        except Exception:
            pass

    # ✅ إذا ضغط إلغاء عام أثناء suggestion (متوافق مع كودك القديم)
    elif mode == "suggestion":
        suggestion_records.pop(user_id, None)
        context.user_data.setdefault(user_id, {})
        context.user_data[user_id].clear()

        # تعديل الرسالة بدل مربع إشعار
        try:
            await query.edit_message_text("❌ تم إلغاء العملية.")
        except Exception:
            pass

    # إذا ما فيه عملية نشطة → لا نرسل تنبيه، فقط نحذف الرسالة
    else:
        pass  # لا نرسل شيء

    # حذف رسالة الزر دائماً
    try:
        await query.message.delete()
    except Exception:
        pass
        
async def show_manual_car_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data.split("_")
    user_id = int(data[1])

    await log_event(update, "📘 فتح قائمة دليل المالك")

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # نحاول أولاً استخدام البراندات
    try:
        manual_df = df_manual
    except Exception as e:
        await log_event(update, f"❌ فشل في تحميل بيانات دليل المالك من Excel: {e}", level="error")
        msg = await query.message.reply_text("📂 تعذر تحميل بيانات دليل المالك حالياً.")
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        return

    brands = []
    if not manual_df.empty and "brand" in manual_df.columns:
        brands = (
            manual_df["brand"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )
        brands = [b for b in brands if b]

    # ✅ في حال وجود براندات → نعرض قائمة البراندات
    if brands:
        keyboard = []
        for brand in brands:
            safe_brand = brand.replace(" ", "_")
            keyboard.append(
                [InlineKeyboardButton(brand, callback_data=f"mnlbrand_{safe_brand}_{user_id}")]
            )

        keyboard.append(
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
        )

        text = (
            "📘 اختر العلامة التجارية أولاً للاطلاع على دليل المالك:\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

        markup = InlineKeyboardMarkup(keyboard)

        try:
            # لو الرسالة نص نعدلها، لو صورة (غلاف) نرسل رسالة جديدة
            if getattr(query.message, "text", None):
                msg = await query.message.edit_text(
                    text,
                    reply_markup=markup,
                    parse_mode=constants.ParseMode.MARKDOWN
                )
            else:
                msg = await query.message.reply_text(
                    text,
                    reply_markup=markup,
                    parse_mode=constants.ParseMode.MARKDOWN
                )
        except Exception as e:
            await log_event(update, f"❌ فشل في إرسال قائمة براندات دليل المالك: {e}", level="error")
            msg = await query.message.reply_text(
                text,
                reply_markup=markup,
                parse_mode=constants.ParseMode.MARKDOWN
            )

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        context.user_data.setdefault(user_id, {})
        context.user_data[user_id]["manual_msg_id"] = msg.message_id
        context.user_data[user_id]["last_message_id"] = msg.message_id
        return

    # 🔁 في حال عدم وجود عمود brand → نرجع للسلوك القديم (قائمة سيارات مباشرة)
    try:
        car_names = manual_df["car_name"].dropna().drop_duplicates().tolist()
    except Exception as e:
        await log_event(update, f"❌ فشل في تحميل قائمة السيارات من Excel: {e}", level="error")
        msg = await query.message.reply_text("📂 تعذر تحميل قائمة دليل المالك حالياً.")
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        return

    keyboard = [
        [InlineKeyboardButton(car, callback_data=f"manualcar_{car.replace(' ', '_')}_{user_id}")]
        for car in car_names
    ]

    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
    )

    text = (
        "📘 اختر فئة السيارة للاطلاع على دليل المالك:\n\n"
        f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
    )

    markup = InlineKeyboardMarkup(keyboard)

    try:
        if getattr(query.message, "text", None):
            msg = await query.message.edit_text(
                text,
                reply_markup=markup,
                parse_mode=constants.ParseMode.MARKDOWN
            )
        else:
            msg = await query.message.reply_text(
                text,
                reply_markup=markup,
                parse_mode=constants.ParseMode.MARKDOWN
            )
    except Exception as e:
        await log_event(update, f"❌ فشل في إرسال قائمة دليل المالك: {e}", level="error")
        msg = await query.message.reply_text(
            text,
            reply_markup=markup,
            parse_mode=constants.ParseMode.MARKDOWN
        )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["manual_msg_id"] = msg.message_id
    context.user_data[user_id]["last_message_id"] = msg.message_id

async def manual_brand_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    استقبال ضغط زر براند دليل المالك:
    mnlbrand_<BRAND>_<USER_ID>
    """
    query = update.callback_query
    data = (query.data or "").split("_")
    if len(data) < 3:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    try:
        user_id = int(data[-1])
    except ValueError:
        await query.answer("❌ رقم مستخدم غير صحيح.", show_alert=True)
        return

    # البراند قد يحتوي مسافات → نجمع ما بين mnlbrand و user_id
    brand = "_".join(data[1:-1]).replace("_", " ").strip()

    # نحفظ البراند في user_data لاستخدامه لاحقاً (مثلاً مع زر "اختيار سيارة اخرى")
    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["manual_brand"] = brand

    await log_event(update, f"📘 اختيار براند دليل المالك: {brand}")

    try:
        manual_df = df_manual
    except NameError:
        await query.answer("⚠️ بيانات دليل المالك غير متاحة حالياً.", show_alert=True)
        return

    subset = manual_df.copy()
    if "brand" in subset.columns:
        subset = subset[subset["brand"].astype(str).str.strip() == brand]

    car_names = (
        subset.get("car_name", pd.Series(dtype=str))
        .dropna()
        .astype(str)
        .str.strip()
        .drop_duplicates()
        .tolist()
    )

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # 🔁 براند بدون سيارات
    if not car_names:
        text = (
            f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n"
            f"🏷 البراند المختار: {brand}\n\n"
            f"📌 {PLACEHOLDER_TEXT}\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ رجوع لاختيار براند آخر", callback_data=f"manual_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")],
        ])

        try:
            # لو الرسالة الأصلية نص نعدلها، لو غلاف/صورة نرسل رسالة جديدة
            if getattr(query.message, "text", None):
                msg = await query.message.edit_text(
                    text,
                    reply_markup=keyboard,
                    parse_mode=constants.ParseMode.MARKDOWN,
                )
            else:
                msg = await query.message.reply_text(
                    text,
                    reply_markup=keyboard,
                    parse_mode=constants.ParseMode.MARKDOWN,
                )
        except Exception:
            msg = await query.message.reply_text(
                text,
                reply_markup=keyboard,
                parse_mode=constants.ParseMode.MARKDOWN,
            )

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, f"براند دليل المالك بدون سيارات فعلية: {brand}")
        return

    # ✅ لدينا سيارات لهذا البراند → نعرضها
    keyboard = [
        [
            InlineKeyboardButton(
                car,
                callback_data=f"manualcar_{car.replace(' ', '_')}_{user_id}",
            )
        ]
        for car in car_names
    ]

    # أزرار الرجوع (براند آخر + قائمة رئيسية)
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع لاختيار براند آخر", callback_data=f"manual_{user_id}")]
    )
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
    )

    text = (
        f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n"
        f"📘 البراند: {brand}\n\n"
        "🚗 اختر فئة السيارة للاطلاع على دليل المالك:\n\n"
        f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
    )

    try:
        # نفس الفكرة: لو الرسالة نص نعدلها، لو كانت غلاف/صورة نرسل رسالة جديدة
        if getattr(query.message, "text", None):
            msg = await query.message.edit_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode=constants.ParseMode.MARKDOWN,
            )
        else:
            msg = await query.message.reply_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode=constants.ParseMode.MARKDOWN,
            )

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        context.user_data[user_id]["manual_msg_id"] = msg.message_id
        context.user_data[user_id]["last_message_id"] = msg.message_id
    except Exception as e:
        await log_event(
            update,
            f"❌ فشل في إرسال قائمة سيارات دليل المالك للبراند {brand}: {e}",
            level="error",
        )

async def handle_manualcar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    parts = query.data.split("_")
    user_id_from_callback = int(parts[-1])
    car_name = " ".join(parts[1:-1])
    user_name = query.from_user.full_name

    try:
        old_msg_id = context.user_data.get(user_id_from_callback, {}).get("manual_msg_id")
        if old_msg_id:
            await context.bot.delete_message(chat_id=query.message.chat_id, message_id=old_msg_id)
    except Exception:
        pass

    # ✅ نستخدم البراند المخزن إن وجد لتصفية شيت manual
    df = df_manual.copy()
    brand = context.user_data.get(user_id_from_callback, {}).get("manual_brand")
    if brand and "brand" in df.columns:
        df = df[df["brand"].astype(str).str.strip() == str(brand).strip()]

    match = df[df["car_name"].astype(str).str.strip() == car_name.strip()]

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # ✅ زر "اختيار سيارة اخرى" يرجع لقائمة سيارات نفس البراند إن وُجد
    if brand:
        brand_slug = str(brand).strip().replace(" ", "_")
        other_car_cb = f"mnlbrand_{brand_slug}_{user_id_from_callback}"
    else:
        # احتياطاً يرجع لقائمة البراندات
        other_car_cb = f"manual_{user_id_from_callback}"

    # ✅ كيبورد الرجوع الموحد (يُستخدم في كل الحالات)
    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("⬅️ اختيار سيارة اخرى", callback_data=other_car_cb)],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id_from_callback}")],
        ]
    )

    # 🔹 لا توجد أي بيانات لهذا الطراز في شيت manual
    if match.empty:
        caption = get_manual_not_available_message(user_name, car_name, delete_time)
        msg = await query.message.reply_text(
            caption,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=back_keyboard,
        )
        register_message(user_id_from_callback, msg.message_id, query.message.chat_id, context)
        await log_event(update, f"📂 لا توجد بيانات لـ {car_name}", level="error")
        return

    image_url = match["cover_image"].values[0]
    index = match.index[0]

    # 🔹 توجد بيانات لكن لا يوجد غلاف (cover_image فارغ)
    if pd.isna(image_url) or str(image_url).strip() == "":
        caption = get_manual_not_available_message(user_name, car_name, delete_time)
        msg = await query.message.reply_text(
            caption,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=back_keyboard,
        )
        register_message(user_id_from_callback, msg.message_id, query.message.chat_id, context)
        await log_event(update, f"📂 لا يوجد غلاف لـ {car_name}", level="error")
        return

    # 🔹 يوجد غلاف → نعرض الغلاف مع زر استعراض الـ PDF + أزرار الرجوع
    caption = get_manual_caption(user_name, car_name)

    keyboard = [
        [InlineKeyboardButton("📘 استعراض دليل المالك", callback_data=f"openpdf_{index}_{user_id_from_callback}")],
        [InlineKeyboardButton("⬅️ اختيار سيارة اخرى", callback_data=other_car_cb)],
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id_from_callback}")],
    ]

    try:
        msg = await query.message.reply_photo(
            photo=image_url,
            caption=caption,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        register_message(user_id_from_callback, msg.message_id, query.message.chat_id, context)
        context.user_data[user_id_from_callback]["manual_msg_id"] = msg.message_id
        await log_event(update, f"✅ تم عرض غلاف دليل {car_name}")
    except Exception as e:
        await log_event(update, f"❌ خطأ أثناء إرسال الغلاف لـ {car_name}: {e}", level="error")
        msg = await query.message.reply_text("📂 فشل في إرسال الغلاف. يرجى المحاولة لاحقاً.")
        register_message(user_id_from_callback, msg.message_id, query.message.chat_id, context)

    context.user_data[user_id_from_callback].pop("manual_viewed", None)

async def handle_manualdfcar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data or ""
    parts = data.split("_")

    # شكل الكولباك: manualpdf_رقمصف_رقممستخدم
    if len(parts) < 3:
        await query.answer("⚠️ بيانات غير صالحة.", show_alert=True)
        return

    try:
        index = int(parts[1])
        user_id = int(parts[2])
    except ValueError:
        await query.answer("⚠️ بيانات غير صالحة.", show_alert=True)
        return

    try:
        row = df_manual.iloc[index]
        car_name = str(row.get("car_name", "")).strip() or "غير معروف"
        file_id = row.get("pdf_file_id", None)
    except Exception:
        await query.answer("❌ تعذر تحميل الملف – غير متوفر أو بيانات غير صالحة.", show_alert=True)
        return

    user_name = query.from_user.full_name or "الصديق"
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # 🔁 نحدد إلى أين يعيدنا زر "اختيار سيارة اخرى"
    user_data = context.user_data.get(user_id, {})
    brand = user_data.get("manual_brand")
    if brand:
        # يرجع لقائمة سيارات نفس البراند
        brand_slug = str(brand).replace(" ", "_")
        other_car_cb = f"mnlbrand_{brand_slug}_{user_id}"
    else:
        # احتياطاً: يرجع لقائمة البراندات
        other_car_cb = f"manual_{user_id}"

    # نفس الكيبورد في الحالتين (متوفر / غير متوفر)
    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("⬅️ اختيار سيارة اخرى", callback_data=other_car_cb)],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")],
        ]
    )

    # ⛔ لا يوجد ملف PDF متوفر
    if file_id is None or pd.isna(file_id) or str(file_id).strip() == "":
        caption = get_manual_not_available_message(user_name, car_name, delete_time)

        # نحاول حذف الرسالة السابقة (الغلاف مثلاً) قبل إرسال الرسالة الجديدة
        try:
            await context.bot.delete_message(
                chat_id=query.message.chat_id,
                message_id=query.message.message_id,
            )
        except Exception:
        # لو ما قدر يحذفها نكمل عادي
            pass

        msg = await query.message.reply_text(
            caption,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=back_keyboard,
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, f"📂 لا يوجد ملف PDF لـ {car_name}", level="error")
        return

    # ✅ يوجد ملف PDF
    caption = get_manual_caption(user_name, car_name)

    # نحاول حذف الرسالة السابقة (الغلاف مثلاً) قبل إرسال الملف
    try:
        await context.bot.delete_message(
            chat_id=query.message.chat_id,
            message_id=query.message.message_id,
        )
    except Exception:
        pass

    try:
        msg = await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=file_id,
            caption=caption,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=back_keyboard,
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)

        context.user_data.setdefault(user_id, {})
        context.user_data[user_id]["manual_sent"] = True

        await log_event(update, f"📘 تم إرسال ملف دليل {car_name}")
    except Exception as e:
        await log_event(update, f"❌ فشل في إرسال دليل PDF لـ {car_name}: {e}", level="error")
        await query.message.reply_text("📂 تعذر إرسال الملف. حاول لاحقاً.")

def get_manual_not_available_message(user_name: str, car_name: str, delete_time: str) -> str:
    return (
        f"`🧑‍💻 استعلام خاص بـ {user_name}`\n\n"
        f"📘 نعتذر، دليل المالك للسيارة ({car_name}) غير متوفر حالياً.\n"
        f"📂 سيتم رفع الملف قريباً بالتحديث القادم.\n\n"
        f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
    )


def get_manual_caption(user_name: str, car_name: str) -> str:
    return (
        f"`🧑‍💼 استعلام خاص بـ {user_name}`\n\n"
        f"📜 دليل المالك للسيارة ({car_name})\n\n"
    )

async def select_car_for_parts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data.split("_")
    user_id = int(data[-1])
    car = " ".join(data[1:-1])

    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["selected_car"] = car
    context.user_data[user_id]["action"] = "parts"
    context.user_data[user_id]["session_valid"] = True  # ✅ تفعيل الجلسة اليدوية

    if "search_attempts" not in context.user_data[user_id]:
        context.user_data[user_id]["search_attempts"] = 0

    # التصنيفات الرئيسية للقطع الاستهلاكية
    part_categories = {
        "🧴 الزيوت": "زيت",
        "🌀 الفلاتر": "فلتر",
        "🔌 البواجي": "بواجي",
        "⚙️ السيور": "سير",
        "🛞 الاقمشة فحمات": "فحمات",
        "💧 السوائل ": "سائل ",
        "🔋 البطاريات": "بطارية",
        "🧼 منتجات مساعدة": "منتج",
    }

    keyboard = [
        [InlineKeyboardButton(name, callback_data=f"catpart_{keyword}_{user_id}")]
        for name, keyword in part_categories.items()
    ]

    # 🔙 زر رجوع لاختيار سيارة أخرى من نفس البراند (إن وجد براند)
    parts_brand = context.user_data[user_id].get("parts_brand")
    if parts_brand:
        safe_brand = parts_brand.replace(" ", "_")
        keyboard.append(
            [InlineKeyboardButton("⬅️ رجوع لاختيار سيارة اخرى", callback_data=f"pbrand_{safe_brand}_{user_id}")]
        )

    # 🔙 زر رجوع للقائمة الرئيسية
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
    )

    # ✅ تنسيق الرد النهائي بصيغة احترافية
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")
    user_name = query.from_user.full_name

    text = (
        f"`🧑‍💼 استعلام خاص بـ {user_name}`\n\n"
        f"🚗 الفئة المختارة: {car}\n\n"
        "اختر تصنيف القطعة التي تريد استعلامها:\n"
        "مثال: فلاتر – زيوت – بواجي – سيور – فحمات – سوائل – بطاريات – منتجات مساعدة.\n\n"
        f"`⏳ سيتم حذف هذه الجلسة تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
    )

    # ⬅️ مهم: لو الرسالة الأصلية صورة، edit_message_text سيفشل → نستخدم reply_text
    try:
        msg = await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=constants.ParseMode.MARKDOWN,
        )
    except Exception:
        msg = await query.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=constants.ParseMode.MARKDOWN,
        )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    await log_event(update, f"عرض تصنيفات القطع للفئة: {car}")
    
async def send_part_image(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    عرض صورة القطعة مع أزرار:
    - 🗂 رجوع لقائمة تصنيفات القطع لنفس الفئة
    - 🚗 اختيار سيارة أخرى
    - ⬅️ رجوع للقائمة الرئيسية
    """
    query = update.callback_query
    data = (query.data or "").split("_")

    # نتوقع: partimg_<index>_<user_id>
    if len(data) < 4:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    try:
        index = int(data[2])
        user_id = int(data[3])
    except ValueError:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    # علامة أن هذه الصورة انفتحت (لو حاب تستخدمها لاحقاً)
    context.user_data.setdefault(user_id, {})[f"image_opened_{index}"] = True
    user_data = context.user_data.setdefault(user_id, {})

    # 🔹 الصف من شيت parts
    try:
        row = df_parts.iloc[index]
    except Exception:
        await query.answer("⚠️ لم أتمكن من قراءة بيانات هذه القطعة.", show_alert=True)
        return

    user_name = query.from_user.full_name or "المستخدم"
    selected_car = user_data.get("selected_car", "غير معروف")

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    station = html.escape(str(row.get("Station Name", "غير معروف"))) if pd.notna(row.get("Station Name")) else "غير معروف"
    part_no = html.escape(str(row.get("Part No", "غير متوفر"))) if pd.notna(row.get("Part No")) else "غير متوفر"

    caption = (
        f"`🧑‍💻 استعلام خاص بـ: {user_name}`\n"
        f"`🚗 الفئة: {selected_car}`\n\n"
        f"القطعة: {station}\n"
        f"رقم القطعة: {part_no}\n\n"
        f"`⏳ سيتم حذف هذه الرسالة تلقائياً بعد 15 دقيقة ({delete_time} / 🇸🇦)`"
    )

    # ============================
    # 🎛 أزرار الرجوع أسفل الصورة
    # ============================
    buttons = []

    # 1) رجوع لقائمة التصنيفات لنفس السيارة (لو محددة)
    safe_car = None
    if selected_car not in (None, "", "غير معروف"):
        safe_car = str(selected_car).replace(" ", "_")
        buttons.append([
            InlineKeyboardButton(
                "🗂 رجوع لقائمة تصنيفات القطع",
                callback_data=f"showparts_{safe_car}_{user_id}"
            )
        ])

    # 2) زر اختيار سيارة أخرى
    # نحاول أولاً نرجع لنفس البراند لو محفوظ، وإلا نفتح قائمة قطع الغيار من جديد
    parts_brand = user_data.get("parts_brand")
    if parts_brand:
        safe_brand = str(parts_brand).replace(" ", "_")
        buttons.append([
            InlineKeyboardButton(
                "🚗 اختيار سيارة أخرى",
                callback_data=f"pbrand_{safe_brand}_{user_id}"
            )
        ])
    else:
        # حالة احتياطية نرجع لقائمة استعلامات قطع الغيار
        buttons.append([
            InlineKeyboardButton(
                "🚗 اختيار سيارة أخرى",
                callback_data=f"parts_{user_id}"
            )
        ])

    # 3) رجوع للقائمة الرئيسية
    buttons.append([
        InlineKeyboardButton(
            "⬅️ رجوع للقائمة الرئيسية",
            callback_data=f"back_main_{user_id}"
        )
    ])

    reply_markup = InlineKeyboardMarkup(buttons)

    msg = await context.bot.send_photo(
        chat_id=query.message.chat_id,
        photo=row.get("Image"),
        caption=caption,
        parse_mode=constants.ParseMode.MARKDOWN,
        reply_markup=reply_markup,
    )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    
async def car_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data.split("_")
    user_id = int(data[-1])

    # اسم السيارة من الكول باك
    car = "_".join(data[1:-1]).replace("_", " ")

    # حفظ نوع السيارة في جلسة المستخدم
    user_data = context.user_data.setdefault(user_id, {})
    user_data["car_type"] = car

    # جلب مسافات الصيانة لهذه السيارة
    kms = (
        df_maintenance[df_maintenance["car_type"] == car]["km_service"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    keyboard = [
        [InlineKeyboardButton(f"{km}", callback_data=f"km_{km}_{user_id}")]
        for km in kms
    ]

    # (اختياري) رجوع لقائمة سيارات نفس البراند إن كان محفوظاً
    brand = user_data.get("brand")
    if brand:
        safe_brand = str(brand).replace(" ", "_")
        keyboard.append(
            [InlineKeyboardButton("⬅️ رجوع لقائمة السيارات", callback_data=f"mbrand_{safe_brand}_{user_id}")]
        )

    # زر رجوع للقائمة الرئيسية
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
    )

    # النص مع اسم السيارة في الأعلى
    text = f"🚗 {car}\nاختر مسافة km الصيانة 🧾 :"

    # 🔁 لو الرسالة الأصلية نص → نعدلها، لو كانت ملف/صورة → نرسل رسالة جديدة
    try:
        if getattr(query.message, "text", None):
            msg = await query.edit_message_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard),
            )
        else:
            raise Exception("message has no text")
    except Exception:
        msg = await query.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
        )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    await log_event(update, f"اختار {car} من قائمة السيارات")


async def km_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data.split("_")

    # شكل الكول باك: km_<km>_<user_id>
    if len(data) < 3:
        await query.answer("❌ استعلام غير صالح.", show_alert=True)
        return

    km_value = data[1]
    try:
        user_id = int(data[2])
    except ValueError:
        await query.answer("❌ استعلام غير صالح.", show_alert=True)
        return

    # 🔐 حماية الاستعلام ليبقى خاص بصاحبه
    if query.from_user.id != user_id:
        requester = await context.bot.get_chat(user_id)
        await query.answer(
            f"❌ هذا الاستعلام خاص ب‏ {requester.first_name} {requester.last_name} - استخدم الأمر go",
            show_alert=True
        )
        return

    user_data = context.user_data.setdefault(user_id, {})
    car = user_data.get("car_type")
    if not car:
        await query.answer("⚠️ لا توجد سيارة محددة لهذه الجلسة.", show_alert=True)
        return

    # 🔎 اختيار الصفوف المطابقة لنوع السيارة والمسافة
    results = df_maintenance[
        (df_maintenance["car_type"] == car) &
        (df_maintenance["km_service"].astype(str) == str(km_value))
    ]

    if results.empty:
        await query.answer("⚠️ لا توجد بيانات صيانة لهذا الطراز عند هذه المسافة.", show_alert=True)
        return

    user_name = query.from_user.full_name
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")
    header = f"`🧑‍💻 استعلام خاص بـ {user_name}`\n\n"

    for i, row in results.iterrows():
        maintenance_action = str(row.get("maintenance_action", "")).strip()

        # 🧩 حالة الطراز قيد التجهيز
        if PLACEHOLDER_TEXT in maintenance_action:
            text = (
                f"{header}"
                f"🚗 *نوع السيارة:* {car}\n"
                f"📏 *المسافة:* {km_value} كم\n\n"
                f"📌 {PLACEHOLDER_TEXT}\n\n"
                f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
            )
        else:
            # ✳️ الحالة العادية: عرض الإجراءات الفعلية من الإكسل
            text = (
                f"{header}"
                f"🚗 *نوع السيارة:* {car}\n"
                f"📏 *المسافة:* {km_value}\n"
                f"🛠️ *الإجراءات:* _{maintenance_action}_\n\n"
                f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
            )

        safe_car = str(car).replace(" ", "_")

        keyboard = [
            [InlineKeyboardButton("عرض تكلفة الصيانة 💰", callback_data=f"cost_{i}_{user_id}")],
            [InlineKeyboardButton("عرض ملف الصيانة 📂", callback_data=f"brochure_{i}_{user_id}")],
            # رجوع لقائمة مسافات الصيانة لنفس السيارة
            [InlineKeyboardButton("⬅️ رجوع لقائمة مسافات الصيانة", callback_data=f"car_{safe_car}_{user_id}")],
            # رجوع للقائمة الرئيسية
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        ]


        msg = await query.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=constants.ParseMode.MARKDOWN
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)

    await log_event(update, f"اختار {car} على مسافة {km_value} كم")

    # محاولة حذف رسالة اختيار الـ KM بعد الإرسال
    try:
        await asyncio.sleep(1)
        await context.bot.delete_message(
            chat_id=query.message.chat_id,
            message_id=query.message.message_id
        )
    except Exception:
        pass

    # ✅ تفريغ الجلسة بعد انتهاء الاستخدام
    # context.user_data[user_id] = {}

async def send_cost(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    index, user_id = int(query.data.split("_")[1]), int(query.data.split("_")[2])

    # 🔐 حماية الاستعلام
    if query.from_user.id != user_id:
        requester = await context.bot.get_chat(user_id)
        await query.answer(
            f"❌ هذا الاستعلام خاص ب‏ {requester.first_name} {requester.last_name} - استخدم الأمر go",
            show_alert=True
        )
        return

    result = df_maintenance.iloc[index]
    car_type = result["car_type"]
    km_service = result["km_service"]
    cost = result["cost_in_riyals"]
    maintenance_action = str(result.get("maintenance_action", "")).strip()

    # 🏷 قراءة البراند من شيت الصيانة كما هو
    brand_raw = str(result.get("brand", "")).strip()

    # 🧩 ربط البراند بوكيله:
    if brand_raw:
        br_low = brand_raw.strip().lower()  # إزالة الفراغات وتحويل للحروف الصغيرة

        # ✅ تطبيع كل صيغ إكسيد → EXEED
        if ("exeed" in br_low) or ("exceed" in br_low):
            norm_brand = "EXEED"
        # ✅ شيري
        elif "chery" in br_low:
            norm_brand = "CHERY"
        # ✅ جيتور
        elif "jetour" in br_low:
            norm_brand = "JETOUR"
        else:
            # أي براند آخر نستخدمه كما هو بعد إزالة الفراغات
            norm_brand = brand_raw.strip()

        dealer_key = DEALER_FOR_BRAND.get(norm_brand, norm_brand)
    else:
        dealer_key = "CHERY"

    # جلب بيانات الشركة والرقم من القاموس
    contact_info = BRAND_CONTACTS.get(dealer_key, {})

    if dealer_key == "EXEED":
        # لو البراند EXEED، نجهز النص لكل شركة بشكل منفصل
        exeeds = contact_info["companies"]
        company_name = ""
        company_phone = ""
        for c in exeeds:
            company_name += f"{c['name']}\n"
            company_phone += f"{c['phone']}\n"
    else:
        company_name = contact_info.get("company", "")
        company_phone = contact_info.get("phone", "")

    user_name = query.from_user.full_name
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # ✳️ حالة الطراز قيد التجهيز، لكن نعرض الشركة ورقم الهاتف دائماً
    if PLACEHOLDER_TEXT in maintenance_action or (not cost and company_name):
        caption = (
            f"`🧑‍💻 استعلام خاص بـ {user_name}`\n"
            f"🚗 نوع السيارة: {car_type}\n"
            f"📏 المسافة: {km_service} كم\n\n"
            f"🏢 الشركة:\n{company_name}"
            f"📞 للحجز اتصل:\n{company_phone}"
            + (f"📌 {PLACEHOLDER_TEXT}\n\n" if PLACEHOLDER_TEXT in maintenance_action else "")
            + f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )
    else:
        # ✳️ الحالة العادية: عرض تكلفة الصيانة
        caption = (
            f"`🧑‍💻 استعلام خاص بـ {user_name}`\n"
            f"`📅 آخر تحديث للأسعار: شهر يونيو/2026`\n"
            f"🚗 نوع السيارة: {car_type}\n"
            f"📏 المسافة: {km_service} كم\n"
            f"💰 تكلفة الصيانة: {cost} ريال\n"
            f"🏢 الشركة:\n{company_name}"
            f"📞 للحجز اتصل:\n{company_phone}\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

    # حذف زرّي "عرض التكلفة" و "عرض ملف الصيانة" من الرسالة الأصلية
    try:
        keyboard = query.message.reply_markup.inline_keyboard
        updated_keyboard = [
            row for row in keyboard
            if not any(
                (btn.callback_data and ("cost_" in btn.callback_data or "brochure_" in btn.callback_data))
                for btn in row
            )
        ]
        await query.message.edit_reply_markup(
            reply_markup=InlineKeyboardMarkup(updated_keyboard) if updated_keyboard else None
        )
    except Exception:
        pass

    safe_car = str(car_type).replace(" ", "_")

    # 🔙 أزرار الرسالة الجديدة لتكلفة الصيانة:
    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("📄 عرض ملف الصيانة", callback_data=f"brochure_{index}_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع لقائمة مسافات الصيانة", callback_data=f"car_{safe_car}_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")],
        ]
    )

    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=caption,
        parse_mode=constants.ParseMode.MARKDOWN,
        reply_markup=back_keyboard
    )
    register_message(user_id, msg.message_id, query.message.chat_id, context)

    await log_event(update, f"عرض تكلفة الصيانة للسيارة {car_type} عند {km_service} كم")

    # ✅ لا نمسح الجلسة بالكامل حتى يبقى زر "اختيار سيارة" يعمل بعد الرجوع
    user_data = context.user_data.get(user_id, {})
    if isinstance(user_data, dict):
        # فقط نمسح القيم المؤقتة لو حاب في المستقبل
        for k in ["km_value", "maintenance_results"]:
            user_data.pop(k, None)
            
async def maintenance_brand_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    استقبال ضغط زر براند الصيانة:
    mbrand_<BRAND>_<USER_ID>
    """
    query = update.callback_query
    data = query.data.split("_")
    user_id = int(data[-1])

    # قد يكون البراند فيه مسافات، نجمع ما بين mbrand و user_id
    brand = "_".join(data[1:-1]).replace("_", " ").strip()

    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["brand"] = brand

    if "brand" not in df_maintenance.columns:
        await query.answer("⚠️ بيانات البراند غير متوفرة حالياً.", show_alert=True)
        return

    # استخراج السيارات لهذا البراند من شيت الصيانة
    cars = (
        df_maintenance[
            df_maintenance["brand"].astype(str).str.strip() == brand
        ]["car_type"]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )

    # لو ما في أي سيارة (يعني البراند كله مجرد صفوف تحضيرية)
    if not cars:
        text = (
            f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n"
            f"🚗 البراند المختار: {brand}\n\n"
            f"📌 {PLACEHOLDER_TEXT}\n\n"
            "`⏳ سيتم إضافة تفاصيل الصيانة لهذا البراند في التحديثات القادمة من قبل فريق GO.`"
        )

        keyboard = [
            [InlineKeyboardButton("⬅️ رجوع لاختيار براند آخر", callback_data=f"maintenance_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")],
        ]

        msg = await query.edit_message_text(
            text,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, f"براند بدون سيارات فعلياً في الصيانة: {brand}")
        return

    # ✅ لدينا سيارات لهذا البراند → نعرض القائمة
    keyboard = [
        [
            InlineKeyboardButton(
                car,
                callback_data=f"car_{car.replace(' ', '_')}_{user_id}"
            )
        ]
        for car in cars
    ]
    # زر رجوع لاختيار براند آخر
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع لاختيار براند آخر", callback_data=f"maintenance_{user_id}")]
    )
    # زر رجوع للقائمة الرئيسية
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
    )

    msg = await query.edit_message_text(
        f"🚗 اختر فئة السيارة ضمن {brand}:",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    register_message(user_id, msg.message_id, query.message.chat_id, context)
    await log_event(update, f"عرض سيارات الصيانة للبراند: {brand}")


async def parts_brand_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    استقبال ضغط زر براند قطع الغيار:
    pbrand_<BRAND>_<USER_ID>
    """
    query = update.callback_query
    data = (query.data or "").split("_")
    if len(data) < 3:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    try:
        user_id = int(data[-1])
    except ValueError:
        await query.answer("❌ رقم مستخدم غير صحيح.", show_alert=True)
        return

    # البراند قد يحتوي مسافات → نجمع ما بين pbrand و user_id
    brand = "_".join(data[1:-1]).replace("_", " ").strip()

    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["parts_brand"] = brand

    await log_event(update, f"🔧 فتح سيارات قطع الغيار للبراند: {brand}")

    # ✅ جلب شيت قطع الغيار
    try:
        parts_df = df_parts
    except NameError:
        await query.answer("⚠️ بيانات قطع الغيار غير متاحة حالياً.", show_alert=True)
        return

    subset = parts_df.copy()

    # نحاول نفلتر حسب عمود البراند لو موجود
    brand_cols = ["Brand", "brand", "BRAND", "البراند"]
    brand_col = next((c for c in brand_cols if c in subset.columns), None)
    if brand_col:
        subset = subset[brand_col].astype(str).str.strip() == brand
        subset = parts_df[subset]

    # تحديد عمود فئة السيارة
    car_col_candidates = ["Station No", "car_name", "Car", "الفئة"]
    car_col = next((c for c in car_col_candidates if c in parts_df.columns), None)

    if not car_col:
        await query.answer("⚠️ لا توجد أعمدة فئات سيارات معرفة لهذا البراند.", show_alert=True)
        return

    car_names = (
        parts_df.loc[parts_df[brand_col].astype(str).str.strip() == brand, car_col]
        .dropna()
        .astype(str)
        .str.strip()
        .drop_duplicates()
        .tolist()
        if brand_col
        else []
    )

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # 🔁 براند بدون أي سيارات
    if not car_names:
        text = (
            f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n"
            f"🔧 البراند المختار: {brand}\n\n"
            "⚠️ لا توجد فئات سيارات متاحة حالياً لهذا البراند في خدمة قطع الغيار.\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ رجوع لاختيار براند آخر", callback_data=f"parts_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")],
        ])

        try:
            if getattr(query.message, "text", None):
                msg = await query.edit_message_text(
                    text,
                    reply_markup=keyboard,
                    parse_mode=constants.ParseMode.MARKDOWN,
                )
            else:
                msg = await query.message.reply_text(
                    text,
                    reply_markup=keyboard,
                    parse_mode=constants.ParseMode.MARKDOWN,
                )
        except Exception:
            msg = await query.message.reply_text(
                text,
                reply_markup=keyboard,
                parse_mode=constants.ParseMode.MARKDOWN,
            )

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        return

    # ✅ لدينا سيارات لهذا البراند → نعرضها
    keyboard = []
    for car in car_names:
        safe_car = str(car).replace(" ", "_")
        # مهم جداً: نستخدم showparts_ عشان يروح لـ select_car_for_parts
        callback_data = f"showparts_{safe_car}_{user_id}"
        keyboard.append([InlineKeyboardButton(car, callback_data=callback_data)])

    # أزرار رجوع
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع لاختيار براند آخر", callback_data=f"consumable_{user_id}")]
    )
    keyboard.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
    )

    text = (
        f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n"
        f"🔧 البراند المختار: {brand}\n\n"
        "🚗 اختر فئة السيارة المناسبة من القائمة:\n\n"
        f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
    )

    markup = InlineKeyboardMarkup(keyboard)

    # 🔐 هنا نعالج مشكلة: There is no text in the message to edit
    try:
        if getattr(query.message, "text", None):
            msg = await query.edit_message_text(
                text,
                reply_markup=markup,
                parse_mode=constants.ParseMode.MARKDOWN,
            )
        else:
            msg = await query.message.reply_text(
                text,
                reply_markup=markup,
                parse_mode=constants.ParseMode.MARKDOWN,
            )
    except Exception:
        msg = await query.message.reply_text(
            text,
            reply_markup=markup,
            parse_mode=constants.ParseMode.MARKDOWN,
        )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    await log_event(update, f"عرض سيارات قطع الغيار للبراند: {brand}")

async def save_parts(df: pd.DataFrame):
    """حفظ شيت قطع الغيار parts"""
    global df_parts
    df_parts = df.copy()

    async with EXCEL_LOCK:
        await asyncio.to_thread(
            write_excel_background,
            "bot_data.xlsx",
            df_parts,
            "parts"
        )

async def send_brochure(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    index, user_id = int(query.data.split("_")[1]), int(query.data.split("_")[2])

    # 🔐 حماية الاستعلام ليبقى خاص بصاحبه
    if query.from_user.id != user_id:
        requester = await context.bot.get_chat(user_id)
        await query.answer(
            f"❌ هذا الاستعلام خاص ب‏ {requester.first_name} {requester.last_name} - استخدم الأمر /go",
            show_alert=True
        )
        return

    result = df_maintenance.iloc[index]
    user_name = query.from_user.full_name
    car_type = result["car_type"]
    km_service = result["km_service"]
    maintenance_action = str(result.get("maintenance_action", "")).strip()

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    header = f"`🧑‍💻 استعلام خاص بـ {user_name}`\n"

    safe_car = str(car_type).replace(" ", "_")

    # 🔙 أزرار الرسالة الجديدة لملف الصيانة:
    # 1) عرض تكلفة الصيانة
    # 2) رجوع لقائمة المسافات
    # 3) رجوع للقائمة الرئيسية
    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("💰 عرض تكلفة الصيانة", callback_data=f"cost_{index}_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع لقائمة مسافات الصيانة", callback_data=f"car_{safe_car}_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")],
        ]
    )

    # 🧩 إذا كان الطراز قيد التجهيز → لا نحاول إرسال صورة
    if PLACEHOLDER_TEXT in maintenance_action:
        caption = (
            f"{header}"
            f"*نوع السيارة 🚗:* {car_type}\n"
            f"*المسافة 📏:* {km_service}\n\n"
            f"📌 {PLACEHOLDER_TEXT}\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=caption,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=back_keyboard,
        )
    else:
        # ✳️ الحالة العادية: إرسال صورة البروشور من العمود brochure_display
        caption = (
            f"{header}"
            f"*نوع السيارة 🚗:* {car_type}\n"
            f"*المسافة 📏:* {km_service}\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

        try:
            msg = await context.bot.send_photo(
                chat_id=query.message.chat_id,
                photo=result["brochure_display"],
                caption=caption,
                parse_mode=constants.ParseMode.MARKDOWN,
                reply_markup=back_keyboard,
            )
        except Exception:
            # لو ما فيه صورة أو في خطأ
            msg = await query.message.reply_text(
                "📂 الملف قيد التحديث حاليا سيكون متاح لاحقا.",
                reply_markup=back_keyboard,
            )

    register_message(user_id, msg.message_id, query.message.chat_id, context)

    # حذف زرّي "عرض ملف الصيانة" و "عرض التكلفة" من الرسالة الأصلية (حتى لا يتكرروا فوق)
    try:
        keyboard = query.message.reply_markup.inline_keyboard
        updated_keyboard = [
            row for row in keyboard
            if not any(
                (btn.callback_data and ("brochure_" in btn.callback_data or "cost_" in btn.callback_data))
                for btn in row
            )
        ]
        await query.message.edit_reply_markup(
            reply_markup=InlineKeyboardMarkup(updated_keyboard) if updated_keyboard else None
        )
    except Exception:
        pass

async def handle_service_centers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id

    context.user_data.setdefault(user_id, {})["service_used"] = True

    try:
        await context.bot.delete_message(chat_id=query.message.chat_id, message_id=query.message.message_id)
    except Exception:
        pass

    # ✅ إرسال الفيديو وتسجيله
    video_path = "مراكز خدمة شيري.MP4"
    if os.path.exists(video_path):
        with open(video_path, "rb") as video_file:
            user_name = query.from_user.full_name
            now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
            delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")
            caption = (
                f"`🧑‍💻 استعلام خاص بـ {user_name}`\n\n"
                f"🗺️  مراكز الخدمة CHERY\n\n"
                f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
            )
            msg1 = await context.bot.send_video(
                chat_id=query.message.chat_id,
                video=video_file,
                caption=caption,
                parse_mode=constants.ParseMode.MARKDOWN
            )
            context.user_data[user_id]["map_msg_id"] = msg1.message_id
            register_message(user_id, msg1.message_id, query.message.chat_id, context)

    # ✅ زرّين + زر رجوع في رسالة واحدة
    keyboard = [
        [InlineKeyboardButton("📍 مواقع فروع شركة شيري", callback_data=f"branches_{user_id}")],
        [InlineKeyboardButton("🔧 المتاجر ومراكز الصيانة المستقلة", callback_data=f"independent_{user_id}")],
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
    ]

    msg2 = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="🛠️ الرجاء اختيار أحد الخيارات التالية:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    register_message(user_id, msg2.message_id, query.message.chat_id, context)

    await log_event(update, "عرض مراكز الخدمة الرسمية للمستخدم")

async def handle_branch_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data.split("_")
    user_id = int(data[1])

    # 🧹 حذف فيديو المواقع السابق إن وجد
    map_msg_id = context.user_data.get(user_id, {}).get("map_msg_id")
    if map_msg_id:
        try:
            await context.bot.delete_message(
                chat_id=query.message.chat_id,
                message_id=map_msg_id
            )
        except Exception:
            pass
        context.user_data[user_id]["map_msg_id"] = None

    # 🧹 حذف زرّي "📍 مواقع الفروع" و"🔧 المتاجر المستقلة" من الرسالة السابقة
    try:
        old_keyboard = query.message.reply_markup.inline_keyboard
        new_keyboard = [
            row for row in old_keyboard
            if not any(
                btn.callback_data
                and ("branches_" in btn.callback_data or "independent_" in btn.callback_data)
                for btn in row
            )
        ]
        await query.message.edit_reply_markup(
            reply_markup=InlineKeyboardMarkup(new_keyboard) if new_keyboard else None
        )
    except Exception:
        pass

    user_name = query.from_user.full_name
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    header = f"`🧑‍💼 استعلام خاص بـ {user_name}`"
    middle = "🚨 مواقع مراكز الصيانة شيري CHERY"
    footer = f"\n\n`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"

    # ==========================================================
    # 🛑 حماية مهمة: branches قد تكون dict وليس list → تسبب خطأ
    # ==========================================================

    raw_branches = context.bot_data.get("branches", [])

    branches: list = []

    if isinstance(raw_branches, list):
        branches = raw_branches

    elif isinstance(raw_branches, dict):
        # إذا رفعنا البيانات على شكل dict من الإكسل
        # نجمع كل العناصر داخلها
        for v in raw_branches.values():
            if isinstance(v, list):
                branches.extend(v)
            elif isinstance(v, dict):
                branches.append(v)

    # الآن branches مضمونة أنها قائمة من dicts

    keyboard_rows: list[list[InlineKeyboardButton]] = []

    for branch in branches:
        if not isinstance(branch, dict):
            continue  # حماية إضافية

        city = str(branch.get("city", "")).strip()
        name = str(branch.get("branch_name", "")).strip()
        url = str(branch.get("url", "")).strip()

        if not city:
            continue

        label = f"📍 {city} / {name}" if name else f"📍 {city}"

        if url and url.startswith("http"):
            keyboard_rows.append([InlineKeyboardButton(label, url=url)])
        else:
            keyboard_rows.append([InlineKeyboardButton(label, callback_data=f"not_ready_{user_id}")])

    if not keyboard_rows:
        await query.answer("❌ لا يوجد فروع صالحة للعرض حالياً.", show_alert=True)
        return

    # زر المراكز المستقلة
    keyboard_rows.append(
        [InlineKeyboardButton("🔧 المتاجر ومراكز الصيانة المستقلة", callback_data=f"independent_{user_id}")]
    )

    # زر الرجوع
    keyboard_rows.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
    )

    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=f"{header}\n{middle}:{footer}",
        parse_mode=constants.ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(keyboard_rows),
    )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    await log_event(update, "عرض قائمة فروع مراكز شيري الرسمية")

async def handle_independent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = int(query.data.split("_")[1])

    # 🧹 حذف فيديو المواقع السابق إن وجد
    map_msg_id = context.user_data.get(user_id, {}).get("map_msg_id")
    if map_msg_id:
        try:
            await context.bot.delete_message(chat_id=query.message.chat_id, message_id=map_msg_id)
        except Exception:
            pass
        context.user_data[user_id]["map_msg_id"] = None

    # 🧹 حذف زرّي "🔧 المتاجر والمراكز المستقلة" و "📍 مواقع فروع شركة شيري" من الرسالة القديمة
    try:
        keyboard = query.message.reply_markup.inline_keyboard
        updated_keyboard = [
            row for row in keyboard
            if not any(
                btn.callback_data
                and ("independent_" in btn.callback_data or "branches_" in btn.callback_data)
                for btn in row
            )
        ]
        await query.message.edit_reply_markup(
            reply_markup=InlineKeyboardMarkup(updated_keyboard) if updated_keyboard else None
        )
    except Exception:
        pass

    context.user_data.setdefault(user_id, {})["independent_used"] = True

    image_path = "شروط-الصيانة.jpg"
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # 🖼 إرسال صورة شروط الصيانة إن وجدت
    if os.path.exists(image_path):
        with open(image_path, "rb") as image_file:
            caption = (
                f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n"
                f"📋 شروط الصيانة للمراكز المستقلة:\n\n"
                f"يمكنك إجراء الصيانة الدورية لدى المراكز المستقلة مع الحفاظ على الضمان متى ما التزمت "
                f"بقطع الغيار والزيوت المطابقة لتعليمات الشركة الصانعة، وتم تدوين بيانات السيارة والفاتورة "
                f"بشكل صحيح وواضح.\n\n"
                f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`"
            )
            msg1 = await context.bot.send_photo(
                chat_id=query.message.chat_id,
                photo=image_file,
                caption=caption,
                parse_mode=constants.ParseMode.MARKDOWN
            )
            register_message(user_id, msg1.message_id, query.message.chat_id, context)

    # 🌍 قائمة المدن من شيت المراكز المستقلة
    cities = df_independent["city"].dropna().unique().tolist()
    city_buttons = [
        [InlineKeyboardButton(city, callback_data=f"setcity_{city}_{user_id}")]
        for city in cities
    ]

    # ✅ إضافة زر "مواقع فروع شركة شيري" أسفل المدن
    city_buttons.append(
        [InlineKeyboardButton("📍 مواقع فروع شركة شيري", callback_data=f"branches_{user_id}")]
    )

    # ✅ زر رجوع للقائمة الرئيسية أسفل المدن
    city_buttons.append(
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
    )

    msg2 = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="🌍 اختر المدينة لعرض المراكز والمتاجر مباشرة:",
        reply_markup=InlineKeyboardMarkup(city_buttons),
        parse_mode=constants.ParseMode.MARKDOWN,
    )
    register_message(user_id, msg2.message_id, query.message.chat_id, context)
    await log_event(update, "عرض قائمة المدن للمراكز والمتاجر المستقلة")


async def set_city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    parts = query.data.split("_")
    city = parts[1]
    user_id = int(parts[2])

    # 🔴 إزالة قفل تكرار المدينة (معطل)
    # if context.user_data.get(user_id, {}).get("city_selected"):

    context.user_data.setdefault(user_id, {})["city"] = city

    try:
        await context.bot.delete_message(chat_id=query.message.chat_id, message_id=query.message.message_id)
    except Exception:
        pass

    keyboard = [
        [InlineKeyboardButton("✅ قائمة المراكز المعتمدة", callback_data=f"show_centers_{user_id}")],
        [InlineKeyboardButton("🛒 قائمة متاجر قطع الغيار", callback_data=f"show_stores_{user_id}")],
        [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")]
    ]

    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=f"`🧑‍💻 استعلام خاص بـ {query.from_user.full_name}`\n\n🔍 اختر نوع الخدمة بعد اختيار المدينة ({city}):",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode=constants.ParseMode.MARKDOWN
    )

    register_message(user_id, msg.message_id, query.message.chat_id, context)
    await log_event(update, f"اختار مدينة: {city}")

async def _send_independent_results(update: Update, context: ContextTypes.DEFAULT_TYPE, filter_type: str):
    """
    عرض نتائج المراكز / المتاجر المستقلة مع صورة المتجر (إن وجدت) + رابط الموقع من ملف Excel.
    يعتمد على شيت independent بالأعمدة:
    name, phone, type, image_url, location_url, city
    """
    query = update.callback_query
    user_id = query.from_user.id
    city = context.user_data.get(user_id, {}).get("city")

    if not city:
        await query.answer("❌ لم يتم تحديد المدينة. استخدم /go لإعادة التحديد.", show_alert=True)
        return

    # فلترة حسب المدينة ونوع السجل (مثلاً: 'مركز' أو 'متجر')
    try:
        results = df_independent[
            (df_independent["city"] == city) &
            (df_independent["type"].astype(str).str.contains(filter_type))
        ]
    except Exception as e:
        logging.error(f"[INDEPENDENT] خطأ أثناء فلترة البيانات: {e}")
        await query.answer("❌ حدث خطأ أثناء قراءة بيانات المراكز المستقلة.", show_alert=True)
        return

    if results.empty:
        msg = await query.message.reply_text(f"🚫 لا توجد بيانات {filter_type} حالياً في {city}.")
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, f"🚫 لا توجد نتائج {filter_type} في {city}", level="error")
        return

    user_name = query.from_user.full_name or "العميل"
    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # 🆕 تجهيز نصوص آمنة لـ HTML
    user_name_safe = html.escape(user_name)
    city_safe = html.escape(str(city))

    for _, row in results.iterrows():
        name = row.get("name", "بدون اسم")
        phone = row.get("phone", "غير متوفر")
        result_type = row.get("type", "")
        image_url = row.get("image_url", "")
        location_url = row.get("location_url", "")

        name_safe = html.escape(str(name))
        phone_safe = html.escape(str(phone))
        result_type_safe = html.escape(str(result_type)) if result_type else "غير محدد"

        # 📝 نص الوصف (HTML بدل ماركداون)
        text = (
            f"<code>🧑‍💻 استعلام خاص بـ {user_name_safe}</code>\n"
            f"<code>🏙️ المدينة: {city_safe}</code>\n\n"
            f"🏪 الاسم: {name_safe}\n"
            f"🔧 التصنيف: {result_type_safe}\n"   # 👈 النوع (متجر / مركز)
            f"📞 الهاتف: {phone_safe}\n"
        )

        # 🌐 رابط الموقع إن وجد (رابط مخفي داخل نص عربي قابل للنقر فقط)
        if isinstance(location_url, str) and location_url.strip():
            safe_url = location_url.strip()
            safe_url_escaped = html.escape(safe_url)
            text += (
                f"🌐 <a href=\"{safe_url_escaped}\">اضغط هنا لعرض الموقع والتفاصيل </a>\n"
            )

        text += (
            f"\n<code>⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)</code>"
        )

        # 🖼 إذا عندنا رابط صورة صالح نرسلها كصورة + كابشن، غير كذا نرسل نص فقط
        try:
            if isinstance(image_url, str) and image_url.strip().lower().startswith("http"):
                msg = await context.bot.send_photo(
                    chat_id=query.message.chat_id,
                    photo=image_url.strip(),
                    caption=text,
                    parse_mode=constants.ParseMode.HTML,
                )
            else:
                msg = await query.message.reply_text(
                    text,
                    parse_mode=constants.ParseMode.HTML
                )
            register_message(user_id, msg.message_id, query.message.chat_id, context)
        except Exception as e:
            logging.warning(f"[INDEPENDENT] فشل إرسال نتيجة مع الصورة لـ {name}: {e}")
            try:
                # fallback: إرسال نص فقط لو الصورة فشلت
                msg = await query.message.reply_text(
                    text,
                    parse_mode=constants.ParseMode.HTML
                )
                register_message(user_id, msg.message_id, query.message.chat_id, context)
            except Exception as e2:
                logging.error(f"[INDEPENDENT] فشل إرسال نتيجة نصية لـ {name}: {e2}")

    await log_event(update, f"✅ عرض نتائج {filter_type} في {city}")

async def show_center_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = int(query.data.split("_")[2])

    # 🧹 إزالة أزرار اختيار نوع الخدمة من الرسالة القديمة (المراكز + المتاجر)
    try:
        keyboard = query.message.reply_markup.inline_keyboard
        updated_keyboard = [
            row for row in keyboard
            if not any(
                btn.callback_data
                and ("show_centers_" in btn.callback_data or "show_stores_" in btn.callback_data)
                for btn in row
            )
        ]
        await query.message.edit_reply_markup(
            reply_markup=InlineKeyboardMarkup(updated_keyboard) if updated_keyboard else None
        )
    except Exception:
        pass

    # 📋 عرض قائمة المراكز المعتمدة
    await _send_independent_results(update, context, filter_type="مركز")

    # 🔁 بعد عرض النتائج: زر "متاجر" + "رجوع"
    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("🛒 قائمة متاجر قطع الغيار", callback_data=f"show_stores_{user_id}")],
            [InlineKeyboardButton("🏙️ اختيار مدينة أخرى", callback_data=f"independent_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")],
        ]
    )

    back_msg = await query.message.reply_text(
        "يمكنك الآن استعراض متاجر قطع الغيار أو العودة للقائمة الرئيسية:",
        reply_markup=back_keyboard,
    )
    register_message(user_id, back_msg.message_id, query.message.chat_id, context)

    await log_event(
        update,
        f"📜 عرض قائمة المراكز المعتمدة في {context.user_data[user_id].get('city', 'غير معروفة')}"
    )

async def show_store_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = int(query.data.split("_")[2])

    # 🧹 إزالة أزرار اختيار نوع الخدمة من الرسالة القديمة (المراكز + المتاجر)
    try:
        keyboard = query.message.reply_markup.inline_keyboard
        updated_keyboard = [
            row for row in keyboard
            if not any(
                btn.callback_data
                and ("show_centers_" in btn.callback_data or "show_stores_" in btn.callback_data)
                for btn in row
            )
        ]
        await query.message.edit_reply_markup(
            reply_markup=InlineKeyboardMarkup(updated_keyboard) if updated_keyboard else None
        )
    except Exception:
        pass

    # 📋 عرض قائمة المتاجر
    await _send_independent_results(update, context, filter_type="متجر")

    # 🔁 بعد عرض النتائج: زر "مراكز" + "رجوع"
    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ قائمة المراكز المعتمدة", callback_data=f"show_centers_{user_id}")],
            [InlineKeyboardButton("🏙️ اختيار مدينة أخرى", callback_data=f"independent_{user_id}")],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back:main:{user_id}")],
        ]
    )

    back_msg = await query.message.reply_text(
        "يمكنك الآن استعراض المراكز المعتمدة أو العودة للقائمة الرئيسية:",
        reply_markup=back_keyboard,
    )
    register_message(user_id, back_msg.message_id, query.message.chat_id, context)

    await log_event(
        update,
        f"📜 عرض قائمة المتاجر في {context.user_data[user_id].get('city', 'غير معروفة')}"
    )

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    raw_data = query.data or ""

    # ✅ تعريف الاسم بشكل آمن لـ HTML (يحل مشكلة NameError نهائيا)
    _user = query.from_user
    name = html.escape((_user.full_name or "الصديق").strip())

    # 🔙 زر رجوع للقائمة الرئيسية back_main_USERID
    if raw_data.startswith("back_main_"):
        try:
            user_id = int(raw_data.split("_")[2])
        except Exception:
            await query.answer("❌ خطأ في زر الرجوع", show_alert=True)
            return

        keyboard = build_main_menu_keyboard(user_id)

        msg = None
        try:
            if getattr(query.message, "text", None):
                msg = await query.edit_message_text(
                    "اختر الخدمة المطلوبة:",
                    reply_markup=keyboard
                )
            else:
                raise Exception("message has no text")
        except Exception:
            msg = await query.message.reply_text(
                "اختر الخدمة المطلوبة:",
                reply_markup=keyboard
            )

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "⬅️ رجوع الى القائمة الرئيسية")
        return

    # ❌ زر إلغاء النقاش والعودة للقائمة الرئيسية
    if raw_data == "cancelteam":
        admin_id = query.from_user.id
        state = context.user_data.get(admin_id, {}) or {}

        # حذف ثريد النقاش من الذاكرة
        thread_id = state.get("team_thread_id")
        if thread_id is not None:
            team_threads.pop(thread_id, None)

        # حذف رسالة تعليمات النقاش إن وُجدت
        chat_id = state.get("team_msg_chat_id")
        msg_id = state.get("team_msg_id")
        if chat_id and msg_id:
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=msg_id)
            except Exception:
                pass

        # تصفير حالة النقاش للمشرف
        state["team_mode"] = False
        state["team_thread_id"] = None
        state["team_msg_chat_id"] = None
        state["team_msg_id"] = None
        context.user_data[admin_id] = state

        # الرجوع للقائمة الرئيسية في الخاص
        keyboard = build_main_menu_keyboard(admin_id)
        await context.bot.send_message(
            chat_id=admin_id,
            text="🔙 تم إلغاء النقاش.\nاختر من القائمة الرئيسية:",
            reply_markup=keyboard,
        )

        await query.answer()
        return

    # ✅ ازرار الرجوع الموحدة back:target:user_id
    if raw_data.startswith("back:"):
        await handle_back(update, context)
        return
 
     # ✅ معالجة خاصة لزر showparts_ (لأن الاسم فيه مسافات تتحول إلى _)
    if raw_data.startswith("showparts_"):
        try:
            data = raw_data[len("showparts_"):]
            last_underscore = data.rfind("_")
            selected_car = data[:last_underscore].replace("_", " ").strip()
            user_id = int(data[last_underscore + 1:])

            context.user_data.setdefault(user_id, {})
            context.user_data[user_id]["selected_car"] = selected_car

            await select_car_for_parts(update, context)
        except Exception as e:
            logging.error(f"🔴 Error in showparts callback: {e}")
            await query.answer("❌ حدث خطأ أثناء معالجة التصنيف.", show_alert=True)
        return

    # من هنا يكمل كود الأزرار العام
    data = raw_data.split("_")

    # ✅ تحضير action و user_id مع حالات خاصة catpart_ و faultcat_
    action = None
    user_id: Optional[int] = None

    if raw_data.startswith("catpart_"):
        # شكل الداتا: catpart_keyword_userid
        if len(data) < 3:
            await query.answer("⚠️ بيانات غير صالحة، يرجى المحاولة مجددًا.", show_alert=True)
            return
        _, keyword, user_id_str = data
        action = "catpart"
        try:
            user_id = int(user_id_str)
        except ValueError:
            logging.error(f"🔴 فشل في تحليل user_id في catpart: {user_id_str}")
            await query.answer("⚠️ خطأ في البيانات، يرجى المحاولة مجددًا.", show_alert=True)
            return

    elif raw_data.startswith("faultcat_"):
    # شكل الداتا: faultcat_idx_userid
        if len(data) < 3:
            await query.answer("❌ بيانات غير صالحة لهذا الاختيار.", show_alert=True)
            return

        action = "faultcat"
        try:
            user_id = int(data[2])
        except ValueError:
            await query.answer("❌ خطأ في بيانات المستخدم.", show_alert=True)
            return

    else:
    # باقي الأنواع الأخرى مثل parts_123 أو suggestion_123 أو faults_123 أو maintenance_123 أو coming_soon_123 ...
        valid = True

        if len(data) < 2:
            await query.answer("⚠️ زر غير مفهوم، يرجى المحاولة مجددًا.", show_alert=True)
            valid = False

        if valid:
        # ✅ آخر جزء دائما هو user_id حتى لو action فيه _
            user_id_str = data[-1]
            try:
                user_id = int(user_id_str)
            except ValueError:
                logging.error(f"🔴 فشل في تحليل user_id: {user_id_str}")
                await query.answer("⚠️ خطأ في البيانات، يرجى المحاولة مجددًا.", show_alert=True)
                valid = False

        if valid:
        # ✅ action قد يكون كلمة واحدة أو كلمتين مثل coming_soon
            action = "_".join(data[:-1])

    chat = query.message.chat
    context.user_data.setdefault(user_id, {})
    context.user_data[user_id]["group_title"] = chat.title or "خاص"
    context.user_data[user_id]["group_id"] = chat.id

    # ================== 🔧 خدمة الأعطال الشائعة ==================
    if action == "faults":
        try:
            faults_df = df_faults
        except NameError:
            faults_df = pd.DataFrame()

        # لا يوجد شيت او فارغ
        if faults_df is None or faults_df.empty or "category" not in faults_df.columns:
            text = (
                "🔧 الأعطال الشائعة وحلولها\n\n"
                "هذه الخدمة تحت التحديث حالياً أو لم يتم إضافة بيانات في ملف Excel بعد.\n\n"
                "عند تجهيز قاعدة بيانات الأعطال سوف تظهر لك قائمة بالأنظمة والأعراض والحلول بإذن الله."
            )
            keyboard = [
                [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
            ]
            msg = await query.edit_message_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)
            await log_event(update, "محاولة فتح خدمة الاعطال الشائعة بدون بيانات")
            return

        # تجهيز قائمة الانظمة / التصنيفات
        categories = (
            faults_df["category"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not categories:
            text = (
                "🔧 الأعطال الشائعة وحلولها\n\n"
                "لم يتم العثور على أي تصنيفات للأعطال في ملف Excel.\n"
                "فضلاً قم بإضافة بيانات في شيت faults."
            )
            keyboard = [
                [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
            ]
            msg = await query.edit_message_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)
            return

        # حفظ التصنيفات في user_data مع الفهرس
        context.user_data[user_id]["fault_categories"] = categories

        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

        keyboard = []
        for idx, cat in enumerate(categories):
            keyboard.append(
                [InlineKeyboardButton(cat, callback_data=f"faultcat_{idx}_{user_id}")]
            )

        # زر رجوع
        keyboard.append(
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        )

        text = (
            "🔧 الأعطال الشائعة وحلولها\n\n"
            "اختر النظام أو التصنيف الذي ترغب عرض الأعطال الشائعة الخاصة به:\n\n"
            "`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة "
            f"({delete_time} / 🇸🇦)`"
        )

        msg = await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=constants.ParseMode.MARKDOWN
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "فتح قائمة الاعطال الشائعة الرئيسية")
        return

    elif action == "faultcat":
        # عرض اعطال تصنيف معين
        if len(data) < 3:
            await query.answer("❌ بيانات غير صالحة لهذا الاختيار.", show_alert=True)
            return

        idx = int(data[1])

        user_store = context.user_data.get(user_id, {})
        categories = user_store.get("fault_categories", [])

        if not categories or idx < 0 or idx >= len(categories):
            await query.answer("❌ لم يتم العثور على هذا التصنيف. حاول من جديد عبر القائمة الرئيسية.", show_alert=True)
            return

        selected_category = categories[idx]

        try:
            faults_df = df_faults
        except NameError:
            faults_df = pd.DataFrame()

        if faults_df is None or faults_df.empty:
            await query.answer("❌ لا توجد بيانات أعطال حالياً.", show_alert=True)
            return

        # تصفية الاعطال حسب التصنيف
        subset = faults_df[
            faults_df["category"].astype(str).str.strip() == str(selected_category).strip()
        ]

        if subset.empty:
            msg = await query.message.reply_text(
                f"🚫 لا توجد أعطال مسجلة حالياً تحت التصنيف:\n• {selected_category}"
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)
            await log_event(update, f"لا توجد اعطال لتصنيف {selected_category}")
            return

        user_name = query.from_user.full_name
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

        for _, row in subset.iterrows():
            car_type = row.get("car_type", "")
            symptom = row.get("symptom", "")
            cause = row.get("cause", "")
            solution = row.get("solution", "")

            text = (
                f"`🧑‍💻 استعلام خاص بـ {user_name}`\n"
                f"`🔧 النظام / التصنيف: {selected_category}`\n"
            )

            if str(car_type).strip():
                text += f"`🚗 نوع السيارة (إن وجد): {car_type}`\n"

            text += "\n"

            if str(symptom).strip():
                text += f"🔹 العَرَض:\n{symptom}\n\n"
            if str(cause).strip():
                text += f"🔹 السبب المحتمل:\n{cause}\n\n"
            if str(solution).strip():
                text += f"🔹 الحل المقترح:\n{solution}\n\n"

            text += (
                f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة "
                f"({delete_time} / 🇸🇦)`"
            )

            msg = await query.message.reply_text(
                text,
                parse_mode=constants.ParseMode.MARKDOWN
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)

                # 2) رجوع للقائمة الرئيسية
        back_keyboard = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("⬅️ العودة لقائمة الأعطال", callback_data=f"faults_{user_id}")],
                [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")],
            ]
        )

        # 🔥 إرسال الأزرار مع نص بسيط حتى تقبل تيليجرام الرسالة
        back_msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="يمكنك المتابعة من خلال الخيارات التالية:",
            reply_markup=back_keyboard
        )

        register_message(user_id, back_msg.message_id, query.message.chat_id, context)

        await log_event(update, f"عرض اعطال التصنيف: {selected_category}")
        return

    # ================== الصيانة الدورية بنظام البراندات ==================
    if action == "maintenance":
        # نحدد أن المستخدم داخل مسار الصيانة
        context.user_data.setdefault(user_id, {})
        context.user_data[user_id]["action"] = "maintenance"

        # نحاول نقرأ البراندات من شيت الصيانة
        if "brand" in df_maintenance.columns:
            brands = (
                df_maintenance["brand"]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )
            brands = [b for b in brands if b]  # حذف الفراغات إن وجدت
        else:
            brands = []

        # لو مافي عمود brand لأي سبب نرجع للسلوك القديم (قائمة سيارات واحدة)
        if not brands:
            cars = (
                df_maintenance["car_type"]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )

            keyboard = [
                [
                    InlineKeyboardButton(
                        car,
                        callback_data=f"car_{car.replace(' ', '_')}_{user_id}"
                    )
                ]
                for car in cars
            ]
            keyboard.append(
                [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
            )

            msg = await query.edit_message_text(
                "🚗 اختر فئة السيارة للصيانة الدورية:",
                reply_markup=InlineKeyboardMarkup(keyboard),
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)
            await log_event(update, "فتح قائمة الصيانة الدورية (بدون براندات)")
            return

        # ✅ هنا السلوك الجديد: عرض براندات أولاً
        keyboard = []
        for brand in brands:
            safe_brand = brand.replace(" ", "_")
            keyboard.append(
                [
                    InlineKeyboardButton(
                        brand,
                        callback_data=f"mbrand_{safe_brand}_{user_id}"
                    )
                ]
            )

        # زر رجوع للقائمة الرئيسية
        keyboard.append(
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        )

        msg = await query.edit_message_text(
            "🏷 اختر العلامة التجارية أولاً ثم سيتم عرض فئات السيارات:",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "فتح قائمة الصيانة الدورية حسب البراند")
        return


    if action == "parts":
        keyboard = [
            # استعلام القطع الاستهلاكية (يبقى كما هو)
            [InlineKeyboardButton(
                "🧩 استعلام قطع الغيار الاستهلاكية",
                callback_data=f"consumable_{user_id}"
            )],
            # استعلام قطع غيار عام → يفتح موقع شيري مباشرة كرابط
            [InlineKeyboardButton(
                "🧩 استعلام قطع غيار عام (موقع شيري الرسمي)",
                url="https://www.cheryksa.com/ar/spareparts"
            )],
            # زر الرجوع للقائمة الرئيسية
            [InlineKeyboardButton(
                "⬅️ رجوع للقائمة الرئيسية",
                callback_data=f"back_main_{user_id}"
            )],
        ]

        msg = await query.edit_message_text(
            "اختر نوع استعلام قطع الغيار ⚙️ :",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "اختار استعلام قطع الغيار")
        return

    elif action in ("external", "extparts"):
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")
        keyboard = [[InlineKeyboardButton("🔗 فتح موقع الاستعلام", url="https://www.cheryksa.com/ar/spareparts")]]
        msg = await query.edit_message_text(
            "🌐 تم تجهيز الرابط، اضغط الزر بالأسفل للانتقال إلى موقع استعلام قطع غيار شيري الرسمي:\n\n"
            f"`⏳ سيتم حذف هذا الاستعلام تلقائياً خلال 15 دقيقة ({delete_time} / 🇸🇦)`",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=constants.ParseMode.MARKDOWN
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "تم فتح رابط قطع الغيار الخارجي (extparts)")
        return

    elif action == "consumable":
        # أولاً نحاول عرض البراندات من شيت parts
        try:
            parts_df = df_parts
        except NameError:
            parts_df = pd.DataFrame()

        brands = []
        if not parts_df.empty and "brand" in parts_df.columns:
            brands = (
                parts_df["brand"]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )
            brands = [b for b in brands if b]

        # في حال توفر البراندات → نعرض قائمة البراندات أولاً
        if brands:
            keyboard = []
            for brand in brands:
                safe_brand = brand.replace(" ", "_")
                keyboard.append(
                    [InlineKeyboardButton(brand, callback_data=f"pbrand_{safe_brand}_{user_id}")]
                )

            keyboard.append(
                [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
            )

            msg = await query.edit_message_text(
                "🏷 اختر العلامة التجارية أولاً لعرض فئات السيارات للقطع الاستهلاكية:",
                reply_markup=InlineKeyboardMarkup(keyboard),
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)
            await log_event(update, "فتح قائمة البراندات للقطع الاستهلاكية (parts)")
            return

        # في حال عدم توفر عمود brand نعود للسلوك القديم (قائمة سيارات واحدة)
        keyboard = []

        for car in unique_cars:
            callback_data = f"showparts_{car.replace(' ', '_')}_{user_id}"
            keyboard.append([InlineKeyboardButton(car, callback_data=callback_data)])

        # زر رجوع في اسفل القائمة
        keyboard.append([InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")])

        if not unique_cars:
            await query.edit_message_text("❌ لا توجد سيارات متاحة في قاعدة البيانات.")
            await log_event(update, "❌ لا توجد سيارات متاحة في قاعدة البيانات (consumable)")
            return

        msg = await query.edit_message_text("🚗 اختر فئة السيارة المطلوبة:", reply_markup=InlineKeyboardMarkup(keyboard))
        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "عرض قائمة السيارات للقطع الاستهلاكية (بدون براندات)")
        return

    elif action == "catpart":
        keyword = data[1]
        user_id = int(data[2])
        selected_car = context.user_data[user_id].get("selected_car")

        if not selected_car:
            await query.answer("❌ يرجى اختيار فئة السيارة أولاً.", show_alert=True)
            return

        filtered_df = df_parts[df_parts["Station No"] == selected_car]
        matches = filtered_df[
            filtered_df["Station Name"]
            .astype(str)
            .str.strip()
            .str.contains(f"^{keyword}|\\s{keyword}", case=False, na=False)
        ]

        if matches.empty:
            await query.answer("❌ لم يتم توفير بيانات لهذا التصنيف بعد.\nهذا الطراز قيد الإعداد من فريق GO.", show_alert=True)
            return

    # 📌 ➤ إضافة بسيطة فقط: حفظ آخر صورة في هذا التصنيف
        last_image_index = None
        for idx, row in matches.iterrows():
            if pd.notna(row.get("Image")):
                last_image_index = idx

        context.user_data.setdefault(user_id, {})
        context.user_data[user_id]["last_image_index_for_cat"] = last_image_index
    # 📌 انتهى التعديل الوحيد هنا

        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")
        footer = f"\n<code>⏳ سيتم حذف هذا الاستعلام تلقائيًا خلال 15 دقيقة ({delete_time} / 🇸🇦)</code>"

        user_name = query.from_user.full_name

    # 🔹 رسائل القطع داخل التصنيف
        for i, row in matches.iterrows():
            part_name_value = row.get("Station Name", "غير معروف")
            part_number_value = row.get("Part No", "غير معروف")
            price = get_part_price(row)  # 💰 استخراج السعر إن وجد

            text = (
                f"<code>🧑‍💼 استعلام خاص بـ {user_name}</code>\n"
                f"<code>🚗 الفئة: {selected_car}</code>\n\n"
                f"🔹 <b>اسم القطعة:</b> {part_name_value}\n"
                f"🔹 <b>رقم القطعة:</b> {part_number_value}\n"
            )

            if price:
                price_display = price
                if "ريال" not in price and "SAR" not in price.upper():
                    price_display = f"{price} ريال"
                text += f"🔹 <b>السعر التقريبي:</b> {price_display}\n"

            text += f"\n<code>📌 تم العثور على نتائج بناءً على التصنيف</code>{footer}"

            keyboard = []
            if pd.notna(row.get("Image")):
                keyboard.append(
                    [InlineKeyboardButton("عرض الصورة 📸", callback_data=f"part_image_{i}_{user_id}")]
                )

            msg = await query.message.reply_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard) if keyboard else None,
                parse_mode=ParseMode.HTML
            )
            register_message(user_id, msg.message_id, query.message.chat_id, context)

        # 🔹 رسالة ختامية فيها أزرار رجوع
        safe_car = selected_car.replace(" ", "_")

        # نحاول نجيب البراند من user_data لو محفوظ
        parts_brand = context.user_data.get(user_id, {}).get("parts_brand")
        back_buttons = [
            [InlineKeyboardButton("🗂 رجوع لقائمة تصنيفات القطع", callback_data=f"showparts_{safe_car}_{user_id}")],
        ]

        if parts_brand:
            # يرجع لقائمة سيارات نفس البراند
            safe_brand = str(parts_brand).replace(" ", "_")
            back_buttons.append(
                [InlineKeyboardButton("🚘 اختيار سيارة أخرى", callback_data=f"pbrand_{safe_brand}_{user_id}")]
            )
        else:
            # احتياط: يرجعه لقائمة خدمة قطع الغيار العامة
            back_buttons.append(
                [InlineKeyboardButton("🚘 اختيار سيارة أخرى", callback_data=f"parts_{user_id}")]
            )

        back_buttons.append(
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        )

        back_keyboard = InlineKeyboardMarkup(back_buttons)

        # 🔥 نرسل رسالة جديدة أسفل النتائج تحتوي أزرار الرجوع
        back_msg = await query.message.reply_text(
            "يمكنك المتابعة من خلال الأزرار التالية:",
            reply_markup=back_keyboard,
        )
        register_message(user_id, back_msg.message_id, query.message.chat_id, context)

        await log_event(update, f"✅ استعلام تصنيفي: {keyword} ضمن {selected_car}")
        return
    
    elif action == "coming":
        teaser = (
            f"<b>🚀 اهلا {name}</b>\n\n"
            "<b>PP | منصة عروض قطع الغيار</b>\n\n"
            "<i>"
            "قدّم طلبك مرة واحدة\n"
            "وسيصل تلقائيًا لمجموعة من التجار\n"
            "لتحصل على عروض متعددة وتختار الأنسب لك"
            "</i>\n\n"
            "<b>⬇️ اضغط الزر وابدأ طلبك الآن</b>"
        )

        # ✅ زر ديناميكي: رابط عند التفعيل / تنبيه عند عدم التفعيل
        if PP_DIRECT_ENABLED and PP_BOT_USERNAME:
            start_btn = InlineKeyboardButton(
                "↗️ ابدأ بطلب القطع الآن",
                url=f"https://t.me/{PP_BOT_USERNAME}?start=pp"
            )
        else:
            start_btn = InlineKeyboardButton(
                "↗️ ابدأ بطلب القطع الآن",
                callback_data=f"coming_soon_{user_id}"
            )

        kb = InlineKeyboardMarkup([
            [start_btn],
            [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
                    ])

        try:
            await query.edit_message_text(
                teaser,
                reply_markup=kb,
                parse_mode=constants.ParseMode.HTML,
                disable_web_page_preview=True
        )
        except Exception:
            await query.message.reply_text(
                teaser,
                reply_markup=kb,
                parse_mode=constants.ParseMode.HTML,
                disable_web_page_preview=True
            )

        await log_event(update, "فتح زر (قريبا شراء قطع غيار مباشر)")
        return


    elif action == "coming_soon":
        await query.answer("⏳ الخدمة قيد التجهيز وسيتم تفعيلها قريبا", show_alert=True)
        await log_event(update, "ضغط زر بدء خدمة PP قبل التفعيل")
        return

    elif action == "suggestion":
        context.user_data[user_id]["action"] = "suggestion"

        query = update.callback_query
        user_obj = query.from_user
        chat = query.message.chat

        user_name = user_obj.full_name
        now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
        delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

        # ✅ اسم المستخدم: باهت صغير (code)
        user_block = f"🧑‍💼 استفسار دعم فني خاص بـ `{query.from_user.full_name}`"

        prompt_block = (
            "💬 أهلاً بك في مركز الدعم الفني لبرنامج GO.\n\n"
            "✉️ يرجى كتابة استفسارك أو ملاحظتك.\n\n"
            "⚠️ لخدمتك بشكل أدق "
            "`نرجو إضافة فئة السيارة والموديل والسنة داخل الاستفسار.`\n\n"
            "ℹ️ إذا احتجت إرسال أكثر من ملف يمكنك إرسالها في استفسارات منفصلة.\n\n"
            "`يتم الاحتفاظ بهذه الجلسة مؤقتاً لمتابعة رد فريق GO`\n"
            f"`⏳ سيتم حذف هذه الجلسة بعد 15 دقيقة ({delete_time} / 🇸🇦)`"
        )

        text = f"{user_block}\n\n{prompt_block}"

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📤 يمكنك اضافة وسائط مع الاستفسار ", callback_data="send_suggestion")],
          # [InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]
        ])

        # 👇 هنا الذكاء: لو الرسالة الحالية هي رسالة الشكر (فيها رقم تذكرة) نخليها كما هي ونرسل رسالة جديدة تحتها
        current_text = (query.message.text or "") if query.message else ""

        if "🎫 رقم تذكرتك" in current_text:
            msg = await query.message.reply_text(
                text,
                reply_markup=keyboard,
                parse_mode=constants.ParseMode.MARKDOWN
            )
        else:
            # السلوك القديم للقائمة الرئيسية أو أماكن أخرى
            try:
                msg = await query.edit_message_text(
                    text,
                    reply_markup=keyboard,
                    parse_mode=constants.ParseMode.MARKDOWN
                )
            except Exception:
                # لو فشل التعديل لأي سبب، نرجع نرسلها كرسالة جديدة
                msg = await query.message.reply_text(
                    text,
                    reply_markup=keyboard,
                    parse_mode=constants.ParseMode.MARKDOWN
                )

        register_message(user_id, msg.message_id, query.message.chat_id, context)
        await log_event(update, "بدأ المستخدم إرسال استفسار أو ملاحظة عبر مركز الدعم الفني")

        if "active_suggestion_id" not in context.user_data[user_id]:
            suggestion_id = await start_suggestion_session(user_id, context)
        else:
            suggestion_id = context.user_data[user_id]["active_suggestion_id"]

        record = suggestion_records.get(user_id, {}).get(suggestion_id)
        if record:
            record["group_name"] = chat.title if getattr(chat, "title", None) else "خاص"
            record["group_id"] = chat.id
            record["user_name"] = user_name

        return
        
async def start_team_general_thread(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """زر: team_main_USERID من القائمة الرئيسية"""
    query = update.callback_query
    data = (query.data or "").split("_")

    if len(data) != 3:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    try:
        admin_id_from_cb = int(data[2])
    except ValueError:
        await query.answer("❌ خطأ في رقم المستخدم.", show_alert=True)
        return

    admin = query.from_user
    admin_id = admin.id

    if admin_id != admin_id_from_cb or admin_id not in AUTHORIZED_USERS:
        await query.answer("❌ غير مصرح لك باستخدام هذا الزر.", show_alert=True)
        return

    thread_id = _next_team_thread_id()
    team_threads[thread_id] = {
        "type": "general",
        "created_by": admin_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "context": {
            "source": "main_menu",
            "chat_id": query.message.chat.id,
            "chat_title": getattr(query.message.chat, "title", "خاص"),
        },
        "reply_count": 0,
    }

    state = context.user_data.setdefault(admin_id, {})
    state["team_mode"] = True
    state["team_thread_id"] = thread_id

    await query.answer()

    # 👤 اسم المشرف أعلى الرسالة باهت
    admin_block = f"`👤 المشرف: {admin.full_name}`"

    # العنوان عادي، والتفاصيل باهت
    text = (
        f"{admin_block}\n\n"
        "🧵 فتح نقاش داخلي جديد لفريق GO\n"
        f"`تم فتح نقاش داخلي جديد برقم #{thread_id}.`\n\n"
        "`✍️ اكتب رسالتك الأولى الآن، وسيتم إرسالها لبقية المشرفين في قنواتهم الخاصة.`"
    )

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ إلغاء النقاش والعودة للقائمة الرئيسية", callback_data="cancelteam")]
    ])

    msg = await context.bot.send_message(
        chat_id=admin_id,
        text=text,
        reply_markup=keyboard,
        parse_mode=ParseMode.MARKDOWN,
    )

    # حفظ رسالة النقاش لحذفها عند الإلغاء
    state["team_msg_chat_id"] = msg.chat_id
    state["team_msg_id"] = msg.message_id

async def start_team_opinion_thread(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """زر: team_opinion_userId_suggestionId من إشعارات الرد"""
    query = update.callback_query
    data = (query.data or "").split("_")

    if len(data) < 3:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    admin = query.from_user
    admin_id = admin.id
    if admin_id not in AUTHORIZED_USERS:
        await query.answer("❌ غير مصرح لك باستخدام هذا الزر.", show_alert=True)
        return

    try:
        user_id = int(data[2])
    except ValueError:
        await query.answer("❌ رقم مستخدم غير صحيح.", show_alert=True)
        return

    # suggestion_id هو بقية السلسلة
    suggestion_id = "_".join(data[3:]) if len(data) > 3 else ""
    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if not record:
        await query.answer("⚠️ لا يوجد سجل لهذا الاستفسار.", show_alert=True)
        return

    # ✅ تصحيح اسم/ID المجموعة لو كانت ناقصة/غير معروف (من user_data ثم bot_data)
    def _bad(v):
        return v in (None, "", "خاص", "غير معروف")

    if _bad(record.get("group_name")) or _bad(record.get("group_id")) or record.get("group_id") in (None, user_id, "غير معروف"):
        uctx = context.user_data.get(user_id, {}) or {}

        fixed_name = uctx.get("final_group_name") or uctx.get("group_title")
        fixed_id = uctx.get("final_group_id") or uctx.get("group_id")

        if (_bad(fixed_name) or _bad(fixed_id) or fixed_id == user_id) and user_id in context.bot_data:
            bctx = context.bot_data.get(user_id, {}) or {}
            fixed_name = bctx.get("group_title") or fixed_name
            fixed_id = bctx.get("group_id") or fixed_id

        if not _bad(fixed_name):
            record["group_name"] = fixed_name
        if not _bad(fixed_id) and fixed_id != user_id:
            record["group_id"] = fixed_id

    thread_id = _next_team_thread_id()
    team_threads[thread_id] = {
        "type": "suggestion",
        "created_by": admin_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "context": {
            "user_id": user_id,
            "user_name": record.get("user_name"),
            "group_name": record.get("group_name"),
            "group_id": record.get("group_id"),
            "suggestion_id": suggestion_id,
            "ticket_no": record.get("ticket_no"),  # ✅ مهم جداً
            "text": record.get("text"),
        },
        "reply_count": 0,
    }

    state = context.user_data.setdefault(admin_id, {})
    state["team_mode"] = True
    state["team_thread_id"] = thread_id

    await query.answer()
    await context.bot.send_message(
        chat_id=admin_id,
        text=(
            f"🧵 تم فتح نقاش داخلي حول استفسار العضو {record.get('user_name','')} "
            f"(نقاش #{thread_id}).\n\n"
            "✍️ اكتب رأيك أو ملاحظتك الآن، وسيتم إرسالها لبقية المشرفين."
        ),
    )

async def team_reply_existing_thread(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """زر: team_reply_threadId من رسالة نقاش سابقة"""
    query = update.callback_query
    data = (query.data or "").split("_")

    if len(data) != 3:
        await query.answer("❌ بيانات غير صالحة.", show_alert=True)
        return

    try:
        thread_id = int(data[2])
    except ValueError:
        await query.answer("❌ رقم نقاش غير صحيح.", show_alert=True)
        return

    admin = query.from_user
    admin_id = admin.id
    if admin_id not in AUTHORIZED_USERS:
        await query.answer("❌ غير مصرح لك باستخدام هذا الزر.", show_alert=True)
        return

    if thread_id not in team_threads:
        await query.answer("⚠️ هذا النقاش لم يعد موجوداً.", show_alert=True)
        return

    state = context.user_data.setdefault(admin_id, {})
    state["team_mode"] = True
    state["team_thread_id"] = thread_id

    await query.answer()
    await context.bot.send_message(
        chat_id=admin_id,
        text=(
            f"🧵 نقاش فريق GO #{thread_id}\n\n"
            "✍️ اكتب ردك الآن ليتم إرساله لبقية المشرفين ضمن هذا النقاش."
        ),
    )

    ### ✅ الدالة المعدلة: handle_suggestion
async def handle_suggestion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    chat = update.effective_chat
    user_id = user.id

    # ... نفس الكود السابق أعلاه بدون تغيير ...

    now_saudi = datetime.now(timezone.utc) + timedelta(hours=3)
    delete_time = (now_saudi + timedelta(minutes=15)).strftime("%I:%M %p")

    # ✅ اسم المستخدم: باهت صغير (code)
    user_block = f"🧑‍💼 استفسار دعم فني خاص بـ `{user.full_name}`"

    prompt_block = (
        "💬 أهلاً بك في مركز الدعم الفني لبرنامج GO.\n\n"
        "✉️ يرجى كتابة استفسارك أو ملاحظتك.\n\n"
        "⚠️ لخدمتك بشكل أدق "
        "`نرجو إضافة فئة السيارة والموديل والسنة داخل الاستفسار.`\n\n"
        "ℹ️ إذا احتجت إرسال أكثر من ملف يمكنك إرسالها في استفسارات منفصلة.\n\n"
        "`يتم الاحتفاظ بهذه الجلسة مؤقتاً لمتابعة رد فريق GO`\n"
        f"`⏳ سيتم حذف هذه الجلسة بعد 15 دقيقة ({delete_time} / 🇸🇦)`"
    )

    text = f"{user_block}\n\n{prompt_block}"

    msg = await update.message.reply_text(
        text,
        parse_mode=constants.ParseMode.MARKDOWN
    )
    register_message(user_id, msg.message_id, chat.id, context)
    await log_event(update, "بدأ المستخدم إرسال استفسار أو ملاحظة عبر مركز الدعم الفني")

    suggestion_id = await start_suggestion_session(user_id, context)

    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if record:
        record["group_name"] = chat.title if chat.type != "private" else "خاص"
        record["group_id"] = chat.id if chat.type != "private" else "غير معروف"
        record["user_name"] = user.full_name

async def handle_suggestion_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data_parts = query.data.split("_")

    # prefix = reply أو replyready
    prefix = data_parts[0] if data_parts else ""

    if len(data_parts) < 3 or not data_parts[1].isdigit():
        await query.answer("❌ لا يمكن معالجة الطلب، البيانات غير مكتملة.", show_alert=True)
        return

    user_id = int(data_parts[1])
    suggestion_id = data_parts[2]
    admin_id = query.from_user.id
    admin_name = query.from_user.full_name

    if admin_id not in AUTHORIZED_USERS:
        await query.answer("❌ غير مصرح لك بالرد.", show_alert=True)
        return

    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if not record:
        await query.answer("❌ لا يوجد سجل لهذا الاستفسار.", show_alert=True)
        return

    # ✅ بعد أول رد: زر (الرد على الاستفسار الوارد) يصبح للاستخدام مرة واحدة فقط للجميع
    # الرد الإضافي يكون فقط عبر زر (✉️ إرسال رد آخر) الذي يصل لصاحب الرد
    if record.get("replied_by") and record.get("caption"):
        ticket_no = record.get("ticket_no")
        ticket_part = f" 🎫 #{ticket_no}" if ticket_no else ""
        await query.answer(
            f"🟥 تم الرد على التذكرة{ticket_part} مسبقًا من قبل: {record['replied_by']}",
            show_alert=True
        )
        return

    # ✅ قفل مؤقت فقط قبل أول رد (منع مشرفين يفتحون التذكرة بنفس الوقت)
    if not record.get("replied_by"):
        locked_by_id = record.get("locked_by_id")
        if locked_by_id and not _lock_expired(record) and int(locked_by_id) != int(admin_id):
            locker = record.get("locked_by_name") or "مشرف آخر"
            await query.answer(f"🔒 التذكرة قيد المعالجة بواسطة: {locker}", show_alert=True)
            return

        ok, reason = lock_ticket(record, admin_id, admin_name)
        if not ok:
            await query.answer(reason, show_alert=True)
            return

    record["reply_opened_by"] = admin_name
    record["user_name"] = record.get("user_name", query.from_user.full_name)

    # ✅ تصحيح بيانات المجموعة إذا كانت ناقصة أو غير صحيحة
    if record.get("group_name") in ["خاص", None] or record.get("group_id") == user_id:
        user_ctx = context.user_data.get(user_id, {})
        record["group_name"] = user_ctx.get("group_title") or user_ctx.get("final_group_name", "غير معروف")
        record["group_id"] = user_ctx.get("group_id") or user_ctx.get("final_group_id", "غير معروف")

    # 👇 هنا التفريع على مرحلتين حسب البادئة
    if prefix == "reply":
        # المرحلة الأولى: اختيار نوع الرد
        keyboard = [
            [InlineKeyboardButton("📋 رد جاهز", callback_data=f"replyready_{user_id}_{suggestion_id}")],
            [InlineKeyboardButton("✍️ رد مخصص", callback_data=f"customreply_{user_id}_{suggestion_id}")],
        ]

        text = (
            "كيف تفضل الرد على هذا الاستفسار؟\n\n"
            f"👤 <b>اسم المستخدم:</b> {record.get('user_name')}\n"
            f"🆔 <b>رقم المستخدم:</b> <code>{user_id}</code>\n"
            f"🏘️ <b>المجموعة:</b> {record.get('group_name')}\n"
            f"🔢 <b>رقم المجموعة:</b> <code>{record.get('group_id')}</code>"
        )
    else:
        # المرحلة الثانية: قائمة الردود الجاهزة
        keyboard = [
            [InlineKeyboardButton(text, callback_data=f"sendreply_{key}_{user_id}_{suggestion_id}")]
            for key, text in SUGGESTION_REPLIES.items()
        ]
        keyboard.append([
            InlineKeyboardButton("⬅️ رجوع لاختيار نوع الرد", callback_data=f"reply_{user_id}_{suggestion_id}")
        ])

        text = (
            "✉️ اختر الرد الجاهز الذي تريد إرساله للمستخدم:\n\n"
            f"👤 <b>اسم المستخدم:</b> {record.get('user_name')}\n"
            f"🆔 <b>رقم المستخدم:</b> <code>{user_id}</code>\n"
            f"🏘️ <b>المجموعة:</b> {record.get('group_name')}\n"
            f"🔢 <b>رقم المجموعة:</b> <code>{record.get('group_id')}</code>"
        )

    msg = await context.bot.send_message(
        chat_id=admin_id,
        text=text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode=ParseMode.HTML
    )

    # حذف القائمة القديمة إن وجدت
    if "reply_menu_chat" in record and "reply_menu_id" in record:
        try:
            await context.bot.delete_message(record["reply_menu_chat"], record["reply_menu_id"])
        except Exception:
            pass

    record["reply_menu_id"] = msg.message_id
    record["reply_menu_chat"] = msg.chat_id

async def send_suggestion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id

    user_data = context.user_data.setdefault(user_id, {})
    attempts = user_data.get("support_attempts", 0)

    if attempts >= 3:
        await query.answer(
            "انتهت الاستفسارات في هذه الجلسة. أغلق الاتصال من زر الالغاء واستخدم GO بالمجموعة للعودة لمركز الدعم مره اخرى .",
            show_alert=True
        )
        return

    user_data["support_attempts"] = attempts + 1

    suggestion_id = user_data.get("active_suggestion_id")
    if not suggestion_id:
        await query.answer("⚠️ لا توجد جلسة دعم نشطة.", show_alert=True)
        return

    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if not record:
        await query.answer("⚠️ لا يوجد استفسار أو ملاحظة محفوظ.", show_alert=True)
        return

    text = record.get("text", "")
    media = record.get("media")

    if not text and not media:
        await query.answer("⚠️ لا يمكن إرسال الاستفسار فارغ.", show_alert=True)
        return

    # تنظيف بيانات الرد السابقة
    record.pop("replied_by", None)
    record.pop("caption", None)

    user_name = query.from_user.full_name
    record["user_name"] = user_name

    # ✅ تثبيت بيانات المجموعة بدون كسر القيم الصحيحة داخل التذكرة
    group_name = record.get("group_name")
    group_id = record.get("group_id")

    if group_name in [None, "غير معروف", "خاص"] or group_id in [None, "غير معروف", user_id]:
        group_name = user_data.get("final_group_name")
        group_id = user_data.get("final_group_id")

    if (not group_name or group_name in ["غير معروف", "خاص"]) and user_id in context.bot_data:
        fallback = context.bot_data[user_id]
        group_name = fallback.get("group_title", group_name)
        group_id = fallback.get("group_id", group_id)

    record["group_name"] = group_name or "غير معروف"
    record["group_id"] = group_id or "غير معروف"

    logging.info(f"[تأكيد المجموعة] المستخدم: {user_id} | المجموعة: {group_name} | ID: {group_id}")

    # 👉 استخراج رقم التذكرة
    ticket_no = record.get("ticket_no", "—")

    # هيدر يرسل للمشرفين
    header = (
        f"👤 الاسم: {user_name}\n"
        f"🆔 رقم المستخدم: <code>{user_id}</code>\n"
        f"🏘️ المجموعة: {group_name}\n"
        f"🔢 رقم المجموعة: <code>{group_id}</code>\n"
        f"🎫 رقم التذكرة: <code>#{ticket_no}</code>\n"
        "╰─────────╯"
    )

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 الرد على الاستفسار الوارد", callback_data=f"reply_{user_id}_{suggestion_id}")]
    ])

    record["admin_messages"] = {}

    # إرسال الاستفسار لكل مشرف
    for admin_id in AUTHORIZED_USERS:
        try:
            sent = None
            full_caption = header

            if media:
                mtype = media["type"]
                fid = media["file_id"]
                if text:
                    full_caption += f"\n\n📝 <b>الاستفسار الوارد :</b>\n<code>{text}</code>"

                if mtype == "photo":
                    sent = await context.bot.send_photo(
                        admin_id, fid,
                        caption=full_caption,
                        parse_mode=ParseMode.HTML,
                        reply_markup=keyboard
                    )
                elif mtype == "video":
                    sent = await context.bot.send_video(
                        admin_id, fid,
                        caption=full_caption,
                        parse_mode=ParseMode.HTML,
                        reply_markup=keyboard
                    )
                elif mtype == "document":
                    sent = await context.bot.send_document(
                        admin_id, fid,
                        caption=full_caption,
                        parse_mode=ParseMode.HTML,
                        reply_markup=keyboard
                    )
                elif mtype == "voice":
                    sent = await context.bot.send_voice(
                        admin_id, fid,
                        caption=full_caption,
                        parse_mode=ParseMode.HTML,
                        reply_markup=keyboard
                    )
            else:
                suggestion_block = f"\n\n📝 <b>الاستفسار الوارد:</b>\n<code>{text}</code>" if text else ""
                full_caption += suggestion_block
                sent = await context.bot.send_message(
                    admin_id,
                    text=full_caption,
                    parse_mode=ParseMode.HTML,
                    reply_markup=keyboard
                )

            if sent:
                record["admin_messages"][admin_id] = sent.message_id

        except Exception as e:
            logging.error(f"[استفسار] فشل في إرسال الاستفسار للمشرف {admin_id}: {e}")

    record["submitted"] = True
    record["timestamp"] = datetime.now()

    # حذف رسالة المعاينة (Preview) إن وُجدت
    try:
        await query.message.delete()
    except Exception:
        pass

    # ✅ رسالة شكر + رقم التذكرة + اقتباس للاستفسار (بنفس تنسيقك)
    thank_you_message = (
        f"🧑‍💼 استفسار دعم فني خاص بـ {user_name}\n\n"
        "🎉 شكرًا لثقتك بفريق الصيانة والدعم الفني GO!\n"
        f"🎫 رقم تذكرتك: #{ticket_no}\n\n"
    )

    if text:
        thank_you_message += (
            "ملخص استفسارك:\n"
            f"{text}\n\n"
        )

    thank_you_message += (
        "✅ تم إرسال الاستفسار بنجاح إلى فريق الدعم الفني .\n"
        "📌 سيتم مراجعة طلبك والرد عليك في هذه المحادثة.\n\n"
    )

    back_keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✉️ إرسال استفسار آخر", callback_data=f"suggestion_{user_id}")],
        ]
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=thank_you_message,
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=back_keyboard
    )

    # ✅ مهم: فصل التذكرة بعد الإرسال حتى ينشئ تذكرة جديدة لاحقًا
    user_data.pop("active_suggestion_id", None)

    # تفريغ سياق المستخدم بعد الإرسال
    keys_to_clear = ["action", "compose_mode", "compose_text", "compose_media"]
    for key in keys_to_clear:
        user_data.pop(key, None)


async def handle_send_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data_parts = (query.data or "").split("_")

    if len(data_parts) < 4:
        await query.answer("❌ تنسيق البيانات غير صحيح.", show_alert=True)
        return

    reply_key = data_parts[1]
    user_id = int(data_parts[2])
    suggestion_id = data_parts[3]
    admin_id = query.from_user.id
    admin_name = query.from_user.full_name

    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if not record:
        await query.answer("❌ لا يوجد سجل لهذا الاستفسار.", show_alert=True)
        return

    # ✅ قفل مؤقت فقط قبل أول رد
    locked_now = False
    if not record.get("replied_by"):
        locked_by_id = record.get("locked_by_id")
        if locked_by_id and not _lock_expired(record) and int(locked_by_id) != int(admin_id):
            locker = record.get("locked_by_name") or "مشرف آخر"
            await query.answer(f"🔒 التذكرة قيد المعالجة بواسطة: {locker}", show_alert=True)
            return

        ok, reason = lock_ticket(record, admin_id, admin_name)
        if not ok:
            await query.answer(reason, show_alert=True)
            return
        locked_now = True

    existing_admin = record.get("replied_by")
    if existing_admin and existing_admin != admin_name:
        await query.answer(
            f"🟥 تم الرد مسبقًا على هذا الاستفسار من قبل: {existing_admin}",
            show_alert=True
        )
        return

    # ✅ حماية: لو ما فيه replied_by (تذكرة جديدة) نضمن أن العداد صفر
    if not record.get("replied_by"):
        record["reply_count"] = int(record.get("reply_count", 0) or 0) if str(record.get("reply_count", "0")).isdigit() else 0
        if record["reply_count"] != 0:
            record["reply_count"] = 0

    # 🔁 نحسب رقم الرد بدون ما نثبته إلا بعد نجاح الإرسال (حتى لا يصير رد إضافي وهمي بعد فشل)
    current_count = int(record.get("reply_count", 0) or 0)
    new_count = current_count + 1
    is_additional = current_count >= 1

    # =========================
    # ✅ تصحيح بيانات المجموعة (اعتبار "غير معروف" نقص + استخدام final_* ثم group_* ثم bot_data)
    # =========================
    def _bad(v):
        return v in (None, "", "خاص", "غير معروف")

    if _bad(record.get("group_name")) or _bad(record.get("group_id")) or record.get("group_id") in (None, user_id, "غير معروف"):
        uctx = context.user_data.get(user_id, {}) or {}

        fixed_name = uctx.get("final_group_name") or uctx.get("group_title")
        fixed_id = uctx.get("final_group_id") or uctx.get("group_id")

        # احتياط: bot_data (خصوصًا إذا التذكرة بدأت من مجموعة ثم انتقل للخاص)
        if (_bad(fixed_name) or _bad(fixed_id) or fixed_id == user_id) and user_id in context.bot_data:
            bctx = context.bot_data.get(user_id, {}) or {}
            fixed_name = bctx.get("group_title") or fixed_name
            fixed_id = bctx.get("group_id") or fixed_id

        if not _bad(fixed_name):
            record["group_name"] = fixed_name
        if not _bad(fixed_id) and fixed_id != user_id:
            record["group_id"] = fixed_id
    # =========================

    group_name = record.get("group_name", "غير معروف")
    group_id = record.get("group_id", "غير معروف")
    user_name = record.get("user_name", "—")
    original_text = record.get("text") or "❓ لا يوجد استفسار محفوظ."
    reply_text = SUGGESTION_REPLIES.get(reply_key, "📌 تم الرد على استفسارك.")
    has_media = record.get("media")

    # 🔗 استخراج أي رابط وتحويله إلى نص قابل للنقر
    url_match = re.search(r"(https?://\S+)", reply_text)
    hidden_link = ""
    if url_match:
        raw_url = url_match.group(1)
        reply_text = reply_text.replace(raw_url, "").strip()
        hidden_link = f"\n\n[🔗 اضغط هنا لعرض التفاصيل]({raw_url})"

    ticket_no = record.get("ticket_no")
    if ticket_no:
        ticket_info_user = (
            f"\u200F🎫 *رقم التذكرة:* `#{ticket_no}`\n"
            f"\u200F🔁 *رقم الرد داخل التذكرة:* `{new_count}`\n\n"
        )
        ticket_info_admin = ticket_info_user
    else:
        ticket_info_user = ""
        ticket_info_admin = ""

    if is_additional:
        user_caption = (
            f"\u200F🔁 *رد إضافي رقم {new_count} من فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_user}"
            f"\u200F📝 *استفسارك أو ملاحظتك:*\n"
            f"```{original_text.strip()}```\n\n"
            f"\u200F💬 *رد المشرف:*\n"
            f"```{reply_text.strip()}```{hidden_link}\n\n"
            f"\u200F🤖 *شكرًا لمتابعتك معنا.*"
        )
    else:
        user_caption = (
            f"\u200F📣 *رد من قبل فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_user}"
            f"\u200F📝 *استفسارك أو ملاحظتك:*\n"
            f"```{original_text.strip()}```\n\n"
            f"\u200F💬 *رد المشرف:*\n"
            f"```{reply_text.strip()}```{hidden_link}\n\n"
            f"\u200F🤖 *شكرًا لثقتك بفريق الصيانة والدعم الفني GO.*"
        )

    if is_additional:
        admin_caption = (
            f"\u200F🔁 *رد إضافي رقم {new_count} من فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_admin}"
            f"\u200F👤 `{user_name}`\n"
            f"\u200F🆔 {user_id}\n"
            f"\u200F🏘️ \u202B{group_name}\u202C\n"
            f"\u200F🔢 `{group_id}`\n"
            + ("\u200F📎 يحتوي على وسائط\n" if has_media else "") + "\n"
            f"\u200F📝 *المداخلة:*\n```{original_text.strip()}```\n\n"
            f"\u200F💬 *رد المشرف:*\n```{reply_text.strip()}```{hidden_link}\n\n"
            f"\u200F✅ تم الرد من قبل: `{admin_name}`"
        )
    else:
        admin_caption = (
            f"\u200F📣 *رد من قبل فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_admin}"
            f"\u200F👤 `{user_name}`\n"
            f"\u200F🆔 {user_id}\n"
            f"\u200F🏘️ \u202B{group_name}\u202C\n"
            f"\u200F🔢 `{group_id}`\n"
            + ("\u200F📎 يحتوي على وسائط\n" if has_media else "") + "\n"
            f"\u200F📝 *المداخلة:*\n```{original_text.strip()}```\n\n"
            f"\u200F💬 *رد المشرف:*\n```{reply_text.strip()}```{hidden_link}\n\n"
            f"\u200F✅ تم الرد من قبل: `{admin_name}`"
        )

    try:
        media = record.get("media")

        # ✅ إرسال الرد للمستخدم
        if media:
            mtype = media["type"]
            fid = media["file_id"]
            if mtype == "photo":
                await context.bot.send_photo(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            elif mtype == "video":
                await context.bot.send_video(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            elif mtype == "document":
                await context.bot.send_document(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            elif mtype == "voice":
                await context.bot.send_voice(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
        else:
            try:
                with open("GO-NOW.PNG", "rb") as image:
                    await context.bot.send_photo(user_id, image, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            except Exception:
                await context.bot.send_message(user_id, text=user_caption, parse_mode=ParseMode.MARKDOWN, disable_web_page_preview=True)

        # ✅ نجاح الإرسال: نثبت العداد واسم المشرف
        record["reply_count"] = new_count
        if not existing_admin:
            record["replied_by"] = admin_name
        record["caption"] = user_caption

        # ✅ فك القفل بعد نجاح الإرسال
        unlock_ticket(record)

        try:
            await query.message.delete()
        except Exception:
            pass

        # إشعار جميع المشرفين
        for aid in AUTHORIZED_USERS:
            try:
                buttons = [
                    [InlineKeyboardButton("🟦 دعوة فريق GO للنقاش", callback_data=f"team_main_{aid}")],
                    [InlineKeyboardButton("🗣️ دعوة إبداء رأي", callback_data=f"team_opinion_{user_id}_{suggestion_id}")],
                ]

                if aid == admin_id:
                    buttons.insert(
                        0,
                        [InlineKeyboardButton("✉️ إرسال رد آخر", callback_data=f"customreply_{user_id}_{suggestion_id}")]
                    )

                reply_markup = InlineKeyboardMarkup(buttons)

                if media:
                    mtype = media["type"]
                    fid = media["file_id"]
                    if mtype == "photo":
                        await context.bot.send_photo(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    elif mtype == "video":
                        await context.bot.send_video(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    elif mtype == "document":
                        await context.bot.send_document(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    elif mtype == "voice":
                        await context.bot.send_voice(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                else:
                    try:
                        with open("GO-NOW.PNG", "rb") as image:
                            await context.bot.send_photo(aid, image, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    except Exception:
                        await context.bot.send_message(aid, text=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup, disable_web_page_preview=True)

            except Exception as e:
                logging.warning(f"[HANDLE_SEND_REPLY][admin_notify {aid}] فشل إرسال الإشعار: {e}")

    except Exception as e:
        # ✅ لو فشل الإرسال: نفك القفل لو كان هو اللي قفّل (حتى لا تعلق التذكرة)
        if locked_now:
            try:
                unlock_ticket(record)
            except Exception:
                pass

        logging.error(f"[HANDLE_SEND_REPLY] فشل في إرسال الرد للمستخدم {user_id}: {e}")
        try:
            await query.answer("❌ حصل خطأ أثناء إرسال الرد. جرّب مرة أخرى.", show_alert=True)
        except Exception:
            pass

async def handle_custom_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    admin_id = query.from_user.id
    admin_name = query.from_user.full_name

    if not data.startswith("customreply_"):
        await query.answer("🚫 بيانات غير صالحة.", show_alert=True)
        return

    try:
        parts = data.split("_")
        user_id = int(parts[1])
        suggestion_id = parts[2]
    except Exception:
        await query.answer("🚫 فشل في استخراج بيانات الاستفسار.", show_alert=True)
        return

    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if not record:
        await query.answer("❌ لا يوجد سجل لهذه الاستفسار.", show_alert=True)
        return

    existing_admin = record.get("replied_by")
    existing_admin_id = record.get("replied_by_id")

    # ✅ بعد أول رد: ممنوع على غير نفس المشرف فتح رد مخصص
    if existing_admin_id and int(existing_admin_id) != int(admin_id):
        await query.answer(
            f"🟥 تم الرد مسبقًا على هذا الاستفسار من قبل: {existing_admin}",
            show_alert=True
        )
        return

    if existing_admin and existing_admin != admin_name and not existing_admin_id:
        await query.answer(
            f"🟥 تم الرد مسبقًا على هذا الاستفسار من قبل: {existing_admin}",
            show_alert=True
        )
        return

    # ✅ تصحيح معلومات المجموعة إن كانت ناقصة
    if record.get("group_name") in ["خاص", None] or record.get("group_id") in [None, user_id]:
        record["group_name"] = context.user_data.get(user_id, {}).get("group_title", "غير معروف")
        record["group_id"] = context.user_data.get(user_id, {}).get("group_id", "غير معروف")

    # 📌 تفعيل وضع الإدخال اليدوي
    context.user_data.setdefault(admin_id, {})
    context.user_data[admin_id]["compose_mode"] = "custom_reply"
    context.user_data[admin_id]["custom_reply_for"] = user_id
    context.user_data[admin_id]["active_suggestion_id"] = suggestion_id

    msg = await query.message.reply_text(
        f"✍️ أرسل الآن الرد المخصص ليتم إرساله للمستخدم `{user_id}`:",
        parse_mode=ParseMode.MARKDOWN
    )

    # ✅ تسجيل الرسالة للحذف التلقائي إن أردت
    register_message(admin_id, msg.message_id, query.message.chat_id, context)

### ✅ الدالة المعدلة: submit_admin_reply
async def submit_admin_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    user_id = context.user_data.get(admin_id, {}).get("custom_reply_for")
    suggestion_id = context.user_data.get(admin_id, {}).get("active_suggestion_id")

    if not user_id or not suggestion_id:
        await query.answer("❌ لا توجد جلسة رد نشطة.", show_alert=True)
        return

    record = suggestion_records.get(user_id, {}).get(suggestion_id)
    if not record:
        await query.answer("❌ لا يوجد سجل لهذه الاستفسار.", show_alert=True)
        return

    admin_name = update.effective_user.full_name

    # ✅ قفل مؤقت فقط قبل أول رد
    locked_now = False
    if not record.get("replied_by"):
        locked_by_id = record.get("locked_by_id")
        if locked_by_id and not _lock_expired(record) and int(locked_by_id) != int(admin_id):
            locker = record.get("locked_by_name") or "مشرف آخر"
            await query.answer(f"🔒 التذكرة قيد المعالجة بواسطة: {locker}", show_alert=True)
            return

        ok, reason = lock_ticket(record, admin_id, admin_name)
        if not ok:
            await query.answer(reason, show_alert=True)
            return
        locked_now = True

    existing_admin = record.get("replied_by")
    if existing_admin and existing_admin != admin_name:
        await query.answer(
            f"🟥 تم الرد مسبقًا على هذا الاستفسار من قبل: {existing_admin}",
            show_alert=True
        )
        return

    # ✅ حماية: لو ما فيه replied_by (تذكرة جديدة) نضمن أن العداد صفر
    if not record.get("replied_by"):
        record["reply_count"] = int(record.get("reply_count", 0) or 0) if str(record.get("reply_count", "0")).isdigit() else 0
        if record["reply_count"] != 0:
            record["reply_count"] = 0

    # 🔁 نحسب رقم الرد بدون ما نثبته إلا بعد نجاح الإرسال
    current_count = int(record.get("reply_count", 0) or 0)
    new_count = current_count + 1
    is_additional = current_count >= 1

    # ✅ نص الرد (قد يكون None أو فاضي)
    raw_text = context.user_data.get(admin_id, {}).get("compose_text")
    reply_text = (raw_text or "").strip()

    # ⭐⭐ دعم الرد الجاهز في الرد المخصص
    if not reply_text:
        ready_reply = context.user_data.get(admin_id, {}).get("ready_reply_text")
        if ready_reply:
            reply_text = ready_reply.strip()

    # ✅ الوسائط: لو المشرف أرسل وسائط نستخدمها، وإلا نعيد إرسال وسائط الاستفسار الأصلية (السلوك القديم)
    admin_media = context.user_data.get(admin_id, {}).get("compose_media")
    orig_media = record.get("media")
    media = admin_media or orig_media

    # ✅ لا نسمح بالرد الفارغ تماماً (لا نص ولا وسائط)
    if not reply_text and not media:
        await query.answer("⚠️ لا يمكن إرسال رد فارغ.", show_alert=True)
        return

    # 🔗 استخراج أي رابط وتحويله إلى نص قابل للنقر
    url_match = re.search(r"(https?://\S+)", reply_text)
    hidden_link = ""
    if url_match:
        raw_url = url_match.group(1)
        reply_text = reply_text.replace(raw_url, "").strip()
        hidden_link = f"\n\n[🔗 اضغط هنا لعرض التفاصيل]({raw_url})"

    user_name = record.get("user_name", "—")
    original_text = (record.get("text") or "❓ لا يوجد استفسار محفوظ.").strip()
    has_media = bool(media)

    # =========================
    # ✅ تصحيح بيانات المجموعة (اعتبار "غير معروف" نقص + استخدام final_* ثم group_* ثم bot_data)
    # =========================
    def _bad(v):
        return v in (None, "", "خاص", "غير معروف")

    if _bad(record.get("group_name")) or _bad(record.get("group_id")) or record.get("group_id") in (None, user_id, "غير معروف"):
        uctx = context.user_data.get(user_id, {}) or {}

        fixed_name = uctx.get("final_group_name") or uctx.get("group_title")
        fixed_id = uctx.get("final_group_id") or uctx.get("group_id")

        if (_bad(fixed_name) or _bad(fixed_id) or fixed_id == user_id) and user_id in context.bot_data:
            bctx = context.bot_data.get(user_id, {}) or {}
            fixed_name = bctx.get("group_title") or fixed_name
            fixed_id = bctx.get("group_id") or fixed_id

        if not _bad(fixed_name):
            record["group_name"] = fixed_name
        if not _bad(fixed_id) and fixed_id != user_id:
            record["group_id"] = fixed_id
    # =========================

    group_name = record.get("group_name", "غير معروف")
    group_id = record.get("group_id", "غير معروف")

    ticket_no = record.get("ticket_no")
    if ticket_no:
        ticket_info_user = (
            f"\u200F🎫 *رقم التذكرة:* `#{ticket_no}`\n"
            f"\u200F🔁 *رقم الرد داخل التذكرة:* `{new_count}`\n\n"
        )
        ticket_info_admin = ticket_info_user
    else:
        ticket_info_user = ""
        ticket_info_admin = ""

    # ===================== رسالة المستخدم =====================
    if is_additional:
        user_caption = (
            f"\u200F🔁 *رد إضافي رقم {new_count} من فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_user}"
            f"\u200F📝 *استفسارك أو ملاحظتك:*\n```{original_text}```\n\n"
        )
    else:
        user_caption = (
            f"\u200F📣 *رد من قبل فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_user}"
            f"\u200F📝 *استفسارك أو ملاحظتك:*\n```{original_text}```\n\n"
        )

    if reply_text:
        user_caption += f"\u200F💬 *رد المشرف:*\n```{reply_text}```{hidden_link}\n\n"

    user_caption += "\u200F🤖 *شكرًا لثقتك بفريق الصيانة والدعم الفني GO.*"

    # ===================== رسالة المشرفين =====================
    if is_additional:
        admin_caption = (
            f"\u200F🔁 *رد إضافي رقم {new_count} من فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_admin}"
            f"\u200F👤 `{user_name}`\n"
            f"\u200F🆔 {user_id}\n"
            f"\u200F🏘️ \u202B{group_name}\u202C\n"
            f"\u200F🔢 `{group_id}`\n"
            + ("\u200F📎 يحتوي على وسائط\n" if has_media else "") + "\n"
            f"\u200F📝 *الاستفسار:*\n```{original_text}```\n\n"
        )
    else:
        admin_caption = (
            f"\u200F📣 *رد من قبل فريق الدعم الفني GO:*\n\n"
            f"{ticket_info_admin}"
            f"\u200F👤 `{user_name}`\n"
            f"\u200F🆔 {user_id}\n"
            f"\u200F🏘️ \u202B{group_name}\u202C\n"
            f"\u200F🔢 `{group_id}`\n"
            + ("\u200F📎 يحتوي على وسائط\n" if has_media else "") + "\n"
            f"\u200F📝 *الاستفسار:*\n```{original_text}```\n\n"
        )

    if reply_text:
        admin_caption += f"\u200F💬 *رد المشرف:*\n```{reply_text}```{hidden_link}\n\n"

    admin_caption += f"\u200F✅ تم الرد من قبل: `{admin_name}`"

    try:
        # ✅ إرسال الرد للمستخدم
        if media:
            mtype = media["type"]
            fid = media["file_id"]
            if mtype == "photo":
                await context.bot.send_photo(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            elif mtype == "video":
                await context.bot.send_video(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            elif mtype == "document":
                await context.bot.send_document(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            elif mtype == "voice":
                await context.bot.send_voice(user_id, fid, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
        else:
            try:
                with open("GO-NOW.PNG", "rb") as image:
                    await context.bot.send_photo(user_id, image, caption=user_caption, parse_mode=ParseMode.MARKDOWN)
            except Exception:
                await context.bot.send_message(
                    user_id,
                    text=user_caption,
                    parse_mode=ParseMode.MARKDOWN,
                    disable_web_page_preview=True
                )

        # ✅ نجاح الإرسال: نثبت العداد واسم المشرف
        record["reply_count"] = new_count
        if not existing_admin:
            record["replied_by"] = admin_name
        record["caption"] = user_caption

        # ✅ فك القفل بعد نجاح الإرسال
        unlock_ticket(record)

        try:
            await query.message.delete()
        except Exception:
            pass

        # حذف أي منيو قديم للرد من ملفات السجل
        if "reply_menu_chat" in record and "reply_menu_id" in record:
            for aid in AUTHORIZED_USERS:
                try:
                    await context.bot.delete_message(record["reply_menu_chat"], record["reply_menu_id"])
                except Exception:
                    pass
            record.pop("reply_menu_chat", None)
            record.pop("reply_menu_id", None)

        # إشعار جميع المشرفين بالرد
        for aid in AUTHORIZED_USERS:
            try:
                buttons = [
                    [InlineKeyboardButton("🟦 دعوة فريق GO للنقاش", callback_data=f"team_main_{aid}")],
                    [InlineKeyboardButton("🗳 دعوة إبداء رأي", callback_data=f"team_opinion_{user_id}_{suggestion_id}")],
                ]

                if aid == admin_id:
                    buttons.insert(
                        0,
                        [InlineKeyboardButton("✉️ إرسال رد آخر", callback_data=f"customreply_{user_id}_{suggestion_id}")]
                    )

                reply_markup = InlineKeyboardMarkup(buttons)

                if media:
                    mtype = media["type"]
                    fid = media["file_id"]
                    if mtype == "photo":
                        await context.bot.send_photo(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    elif mtype == "video":
                        await context.bot.send_video(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    elif mtype == "document":
                        await context.bot.send_document(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    elif mtype == "voice":
                        await context.bot.send_voice(aid, fid, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                else:
                    try:
                        with open("GO-NOW.PNG", "rb") as image:
                            await context.bot.send_photo(aid, image, caption=admin_caption, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                    except Exception:
                        await context.bot.send_message(
                            aid,
                            text=admin_caption,
                            parse_mode=ParseMode.MARKDOWN,
                            reply_markup=reply_markup,
                            disable_web_page_preview=True
                        )

            except Exception as e:
                logging.warning(f"[رد مخصص - إشعار مشرف {aid}] فشل: {e}")

        # تنظيف حالة المشرف بعد الإرسال
        context.user_data.pop(admin_id, None)

    except Exception as e:
        if locked_now:
            try:
                unlock_ticket(record)
            except Exception:
                pass
        logging.error(f"[رد مخصص] فشل في إرسال الرد للمستخدم {user_id}: {e}")
        try:
            await query.answer("❌ حصل خطأ أثناء إرسال الرد. جرّب مرة أخرى.", show_alert=True)
        except Exception:
            pass

async def handle_control_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # 🧠 سجل محاولة الدخول
    await log_event(update, "🛠️ المستخدم طلب الدخول إلى لوحة التحكم")

    if user_id not in AUTHORIZED_USERS:
        await update.message.reply_text("🚫 غير مصرح لك بالدخول إلى لوحة التحكم.")
        return

    keyboard = [
        [InlineKeyboardButton("👤 المشرفون", callback_data="admins_menu")],
        [InlineKeyboardButton("🧹 تنظيف الجلسات", callback_data="clear_sessions")],
        [InlineKeyboardButton("♻️ إعادة تحميل الإعدادات", callback_data="reload_settings")],
        [InlineKeyboardButton("🚧 تفعيل وضع الصيانة", callback_data="ctrl_maintenance_on")],
        [InlineKeyboardButton("✅ إنهاء وضع الصيانة", callback_data="ctrl_maintenance_off")],
        [InlineKeyboardButton("🧨 تدمير البيانات", callback_data="self_destruct")],
        [InlineKeyboardButton("🔁 إعادة تشغيل الجلسة", callback_data="restart_session")],
        [InlineKeyboardButton("💾 النسخ الاحتياطي الآن", callback_data="ctrl_backup")],
        [InlineKeyboardButton("🚪 خروج", callback_data="exit_control")],
    ]

    await update.message.reply_text(
        "🎛️ *لوحة التحكم الخاصة بالمشرفين*\n\nيرجى اختيار الإجراء المطلوب:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode=ParseMode.MARKDOWN
    )

# ✅ معالجة الضغط على أزرار الصيانة
async def handle_control_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    action = query.data
    user_id = query.from_user.id

    if user_id not in AUTHORIZED_USERS:
        await query.answer("🚫 لا تملك صلاحية الوصول.", show_alert=True)
        return

    # ✅ تفعيل وضع الصيانة
    if action == "ctrl_maintenance_on":
        context.bot_data["maintenance_mode"] = True
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ تم تفعيل وضع الصيانة.\nلن يستطيع المستخدمون استخدام الخدمات مؤقتًا.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]])
        )
        return

    # ✅ إنهاء وضع الصيانة
    if action == "ctrl_maintenance_off":
        context.bot_data["maintenance_mode"] = False
        await context.bot.send_message(
            chat_id=user_id,
            text="✅ تم إنهاء وضع الصيانة.\nيمكن للمستخدمين استخدام الخدمات الآن.",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]]
            )
        )
        return

    # ✅ نسخ احتياطي يدوي من لوحة التحكم
    if action == "ctrl_backup":
        await query.answer("⏳ يتم الآن إنشاء نسخة احتياطية للبيانات...", show_alert=True)
        await create_excel_backup(reason="manual", context=context, notify_chat_id=user_id)
        return

    # باقي الإجراءات كما هي
    if action == "control_back":
        await query.message.edit_text(
            "🛠️ *لوحة التحكم:*",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("👤 المشرفون", callback_data="admins_menu")],
                [InlineKeyboardButton("🧹 تنظيف الجلسات", callback_data="clear_sessions")],
                [InlineKeyboardButton("♻️ إعادة تحميل الإعدادات", callback_data="reload_settings")],
                [InlineKeyboardButton("🚧 تفعيل وضع الصيانة", callback_data="ctrl_maintenance_on")],
                [InlineKeyboardButton("✅ إنهاء وضع الصيانة", callback_data="ctrl_maintenance_off")],
                [InlineKeyboardButton("🧨 تدمير البيانات", callback_data="self_destruct")],
                [InlineKeyboardButton("🔁 إعادة تشغيل الجلسة", callback_data="restart_session")],
                [InlineKeyboardButton("💾 النسخ الاحتياطي الآن", callback_data="ctrl_backup")],
                [InlineKeyboardButton("🚪 خروج", callback_data="exit_control")]
            ]),
            parse_mode=constants.ParseMode.MARKDOWN
        )
        return

    if query.data == "exit_control":
        try:
            await query.message.delete()
        except Exception:
            pass

        try:
            await context.bot.send_message(
                chat_id=user_id,
                text="🚪 تم الخروج من لوحة التحكم.",
            )
        except Exception:
            pass

        return

    if query.data == "self_destruct":
        if user_id == 1543083749:
            await query.answer("💣 لاتملك هذي الصلاحية  (تدمير البيانات).", show_alert=True)
        else:
            await query.answer("🚫 أنت لا تملك الصلاحية لتنفيذ هذا الإجراء.", show_alert=True)
        return

    if query.data == "admins_menu":
        await query.message.edit_text(
            "👤 *إدارة المشرفين: اختر الإجراء المطلوب*",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📑 عرض المشرفين", callback_data="list_admins")],
                [InlineKeyboardButton("➕ إضافة مشرف", callback_data="add_admin")],
                [InlineKeyboardButton("🗑️ حذف مشرف", callback_data="delete_admin")],
                [InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]
            ]),
            parse_mode=constants.ParseMode.MARKDOWN
        )
        return

    if query.data == "list_admins":
        try:
            # تحميل آخر نسخة حديثة من شيت managers فورياً
            try:
                df_admins_local = pd.read_excel("bot_data.xlsx", sheet_name="managers")
            except Exception:
                df_admins_local = globals().get("df_admins", pd.DataFrame(columns=["manager_id"]))  # نسخة fallback

            if df_admins_local is None or df_admins_local.empty:
                await query.message.edit_text(
                    "⚠️ لا يوجد مشرفون مسجلون حالياً.",
                    reply_markup=InlineKeyboardMarkup(
                        [[InlineKeyboardButton("⬅️ عودة", callback_data="admins_menu")]]
                    )
                )
                return

            rows = []
            for i, row in df_admins_local.iterrows():
                try:
                    id_ = int(row.get("manager_id"))
                except Exception:
                    continue

                # جلب اسم المشرف من تيليجرام
                try:
                    user = await context.bot.get_chat(id_)
                    name = user.full_name
                except Exception:
                    name = "❓ غير معروف"

                rows.append(f"{i+1}. {name}\n🆔 `{id_}`")

            if not rows:
                await query.message.edit_text(
                    "⚠️ لم يتم العثور على مشرفين.",
                    reply_markup=InlineKeyboardMarkup(
                        [[InlineKeyboardButton("⬅️ عودة", callback_data="admins_menu")]]
                    )
                )
                return

            await query.message.edit_text(
                "📑 *قائمة المشرفين:*\n\n" + "\n\n".join(rows),
                parse_mode=constants.ParseMode.MARKDOWN,
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("⬅️ عودة", callback_data="admins_menu")]]
                )
            )

        except Exception as e:
            await query.message.reply_text(f"❌ فشل في تحميل القائمة: {e}")
        return

    if query.data == "add_admin":
        context.user_data[user_id] = {"action": "awaiting_new_admin_id"}
        await query.message.reply_text("✏️ أرسل الآن رقم ID الخاص بالمشرف الجديد.")
        return

    if query.data == "delete_admin":
        context.user_data[user_id] = {"action": "awaiting_admin_removal"}
        await query.message.reply_text("🗑️ أرسل رقم ID للمشرف الذي ترغب بحذفه نهائيًا.")
        return

    if query.data == "clear_sessions":
        removed_count = await cleanup_old_sessions(context)
        await query.answer("🧼 تم تنفيذ التنظيف", show_alert=False)
        await query.message.edit_text(
            f"🧹 تم تنظيف الجلسات المؤقتة.\n📌 عدد الرسائل المحذوفة: {removed_count}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]])
        )
        return

    if query.data == "reload_settings":
        try:
            df_admins = pd.read_excel("bot_data.xlsx", sheet_name="managers")
            AUTHORIZED_USERS.clear()
            for _, row in df_admins.iterrows():
                AUTHORIZED_USERS.append(int(row["manager_id"]))
            await query.message.edit_text("✅ تم إعادة تحميل ملف الإعدادات وتحديث البيانات.",
                                          reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]]))
        except Exception as e:
            await query.message.edit_text(f"❌ حدث خطأ أثناء تحميل الإعدادات:\n{e}",
                                          reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]]))
        return

    if query.data == "restart_session":
        context.user_data.clear()
        context.bot_data.clear()
        await query.answer("🔁 تم إعادة تشغيل الجلسة بنجاح.", show_alert=True)
        await query.message.edit_text("♻️ تم تفريغ جميع بيانات الجلسة.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ عودة", callback_data="control_back")]]))
        return

async def handle_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id

    if query.data != f"rate_{user_id}":
        await query.answer("⚠️ حدث خطأ في البيانات.", show_alert=True)
        return

    context.user_data.setdefault(user_id, {})["rating_mode"] = True

    await query.answer()

    # فقط باراميترين
    await show_statistics(update, context)

async def save_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global RATED_USERS

    query = update.callback_query
    data = query.data or ""
    parts = data.split("_")

    # شكل الكول باك: ratingval_رقم_رقم
    if len(parts) != 3:
        await query.answer("⚠️ تنسيق غير صالح.", show_alert=True)
        return

    try:
        rating_value = int(parts[1])
        user_id = int(parts[2])
    except ValueError:
        await query.answer("⚠️ بيانات تقييم غير صالحة.", show_alert=True)
        return

    # منع أي أحد غير صاحب الجلسة من التقييم
    if query.from_user.id != user_id:
        try:
            requester = await context.bot.get_chat(user_id)
            requester_name = requester.full_name
        except Exception:
            requester_name = "المستخدم"

        await query.answer(
            f"❌ هذا التقييم خاص بـ {requester_name} - استخدم الأمر /go",
            show_alert=True,
        )
        return

    now = datetime.now(timezone.utc) + timedelta(hours=3)
    user_name = query.from_user.full_name

    # محاولة جلب اسم ورقم المجموعة من user_data
    group_name = context.user_data.get(user_id, {}).get("group_title", "غير معروف")
    group_id = context.user_data.get(user_id, {}).get("group_id", "غير معروف")

    # لو مافي بيانات كافية نحاول من bot_data
    if group_name == "غير معروف" and user_id in context.bot_data:
        group_name = context.bot_data[user_id].get("group_title", "غير معروف")
        group_id = context.bot_data[user_id].get("group_id", "غير معروف")

    # لو ما زالت البيانات غير واضحة، ونفس رسالة التقييم جايه من مجموعة، نستخدم عنوان المجموعة
    chat = query.message.chat if query.message else None
    if (
        (group_name in ["غير معروف", None, "خاص"] or group_id in ["غير معروف", None, user_id])
        and chat is not None
        and chat.type != "private"
    ):
        group_name = chat.title or group_name
        group_id = chat.id

    rating_entry = {
        "user_id": user_id,
        "name": user_name,
        "rating": rating_value,
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "group_name": group_name,
        "group_id": group_id,
    }

    try:
        ratings_file = "bot_data.xlsx"

        # قراءة شيت ratings إن وجد
        try:
            df_ratings = pd.read_excel(ratings_file, sheet_name="ratings")
        except Exception:
            df_ratings = pd.DataFrame(
                columns=["user_id", "name", "rating", "timestamp", "group_name", "group_id"]
            )

        # ✅ توحيد نوع user_id داخل df_ratings
        if not df_ratings.empty and "user_id" in df_ratings.columns:
            try:
                df_ratings["user_id"] = pd.to_numeric(df_ratings["user_id"], errors="coerce")
            except Exception:
                pass

        # ✅ هل هذا المستخدم قيّم من قبل؟
        already_rated = False
        if not df_ratings.empty and "user_id" in df_ratings.columns:
            try:
                already_rated = int(user_id) in df_ratings["user_id"].dropna().astype(int).tolist()
            except Exception:
                already_rated = False

        # طبقة حماية إضافية من الكاش
        if user_id in RATED_USERS:
            already_rated = True

        if already_rated:
            # إزالة أزرار التقييم من الرسالة الأصلية (إن أمكن)
            try:
                if query.message:
                    await context.bot.edit_message_reply_markup(
                        chat_id=query.message.chat_id,
                        message_id=query.message.message_id,
                        reply_markup=None,
                    )
            except Exception:
                pass

            # تنظيف مود التقييم
            user_dict = context.user_data.get(user_id)
            if isinstance(user_dict, dict):
                user_dict.pop("rating_mode", None)

            # إظهار قائمة الخدمات
            try:
                main_keyboard = build_main_menu_keyboard(user_id)
                msg = await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text="🔙 تم تسجيل تقييمك سابقًا، وهذه قائمة خدمات GO:",
                    reply_markup=main_keyboard,
                )
                register_message(user_id, msg.message_id, query.message.chat_id, context)
            except Exception as e:
                logging.warning(f"[RATING] فشل إرسال قائمة الخدمات بعد اكتشاف تقييم سابق: {e}")

            # تنبيه منبثق بالنص الكامل
            alert_text = (
                "🌟 شكرًا لك من جديد على ثقتك ودعمك لنظام GO.\n\n"
                f"{user_name}\n"
                "تم تسجيل تقييمك في وقت سابق، ووجودك معنا هو أهم تقييم ❤️"
            )
            await query.answer(alert_text, show_alert=True)
            return

        # ✅ مستخدم جديد في التقييم → نضيفه إلى الإكسل
        df_ratings = pd.concat([df_ratings, pd.DataFrame([rating_entry])], ignore_index=True)

        # تحديث قائمة المقيمين في الذاكرة
        RATED_USERS.add(user_id)

        # حفظ التقييم في الخلفية بدون تجميد البوت
        async with EXCEL_LOCK:
            await asyncio.to_thread(
                write_excel_background,
                ratings_file,
                df_ratings,
                "ratings"
            )

        # محاولة حذف رسالة أزرار التقييم القديمة
        try:
            if query.message:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=query.message.message_id,
                )
        except Exception:
            pass

        # تنظيف مود التقييم من user_data
        user_dict = context.user_data.get(user_id)
        if isinstance(user_dict, dict):
            user_dict.pop("rating_mode", None)

        # قاموس الايموجيات
        rating_emojis = {
            1: "😞 غير راضٍ",
            2: "😐 مقبول",
            3: "😊 جيد",
            4: "😍 ممتاز",
        }

        thank_you_message = (
            f"🟦 شكراً لتقييمك،\n"
            f"`{user_name}`\n\n"
            f"`تقييمك: {rating_emojis.get(rating_value, '⭐')}`\n\n"
            "🎉 رأيك يهمنا ويساعدنا في تحسين البرنامج!"
        )

        back_keyboard = InlineKeyboardMarkup(
            [[InlineKeyboardButton("⬅️ رجوع للقائمة الرئيسية", callback_data=f"back_main_{user_id}")]]
        )

        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=thank_you_message,
            parse_mode=constants.ParseMode.MARKDOWN,
            reply_markup=back_keyboard,
        )
        register_message(user_id, msg.message_id, query.message.chat_id, context)

        # اسم عرض للمجموعة في إشعار المشرفين
        display_group_name = group_name or "استعلام خاص (بدون مجموعة مرتبطة)"

        # إشعار المشرفين
        for admin_id in AUTHORIZED_USERS:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=(
                        "🌟 *تقييم جديد من مستخدم*\n\n"
                        f"👤 الاسم:\n`{user_name}`\n\n"
                        f"👥 المجموعة:\n`{display_group_name}`\n\n"
                        f"🆔 رقم المجموعة:\n`{group_id}`\n\n"
                        f"📝 التقييم:\n`{rating_emojis.get(rating_value, '⭐')}`\n\n"
                        f"🕓 الوقت:\n`{rating_entry['timestamp']}`"
                    ),
                    parse_mode=constants.ParseMode.MARKDOWN,
                )
            except Exception as e:
                logging.warning(f"❌ فشل إرسال إشعار التقييم للمشرف {admin_id}: {e}")

    except Exception as e:
        logging.error(f"[RATING] ❌ فشل في حفظ التقييم: {e}", exc_info=True)
        await query.answer("⚠️ حدث خطأ أثناء حفظ التقييم، حاول لاحقًا.", show_alert=True)

async def handle_add_admin_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    message = update.message

    if context.user_data.get(user_id, {}).get("action") != "awaiting_new_admin_id":
        return  # تجاهل الرسائل خارج السياق

    new_admin_id_text = message.text.strip()
    if not new_admin_id_text.isdigit():
        await message.reply_text("❌ يجب إدخال رقم ID رقمي صالح.")
        return

    new_admin_id = int(new_admin_id_text)

    global df_admins  # ✅ استخدم النسخة المحملة في الذاكرة

    if new_admin_id in AUTHORIZED_USERS:
        await message.reply_text("ℹ️ هذا المشرف موجود مسبقًا.")
        return

    # ✅ إضافة إلى القائمة الحالية
    AUTHORIZED_USERS.append(new_admin_id)
    df_admins = pd.concat([df_admins, pd.DataFrame([{"manager_id": new_admin_id}])], ignore_index=True)

    # ✅ حفظ التغييرات في الملف Excel
    try:
        # قفل الكتابة على ملف الإكسل قبل تعديل شيت managers
        async with EXCEL_LOCK:
            with pd.ExcelWriter("bot_data.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_admins.to_excel(writer, sheet_name="managers", index=False)

        await message.reply_text(f"✅ تم إضافة المشرف بنجاح: `{new_admin_id}`", parse_mode=ParseMode.MARKDOWN)

    except Exception as e:
        await message.reply_text(f"❌ حدث خطأ أثناء حفظ التغييرات:\n{e}")

    # 🧼 مسح الحالة
    context.user_data[user_id]["action"] = None


def _prepare_reco_targets_for_admin(admin_id: int, context: ContextTypes.DEFAULT_TYPE):
    """
    يبني قائمة المجموعات المتاحة للبث مع أسماء من df_group_logs إن أمكن.
    يخزنها في user_data[admin_id]["reco_targets"]
    """
    ud = context.user_data.setdefault(admin_id, {})
    targets = sorted(list(collect_target_chat_ids(context)))

    groups = []
    global df_group_logs
    for cid in targets:
        title = None
        try:
            if df_group_logs is not None and not df_group_logs.empty and "chat_id" in df_group_logs.columns:
                mask = df_group_logs["chat_id"].astype(str) == str(cid)
                row = df_group_logs[mask]
                if not row.empty:
                    title = str(row.iloc[0].get("title") or "").strip()
        except Exception as e:
            logging.warning(f"[RECO TARGETS] فشل قراءة اسم المجموعة {cid} من df_group_logs: {e}")

        if not title:
            title = f"مجموعة {cid}"

        groups.append({"id": cid, "title": title})

    ud["reco_targets"] = groups
    ud["reco_selected"] = ud.get("reco_selected") or []
    ud["reco_page"] = 0

    logging.info(f"[RECO GROUPS] للمشرف {admin_id}: عدد المجموعات المتاحة للبث = {len(groups)}")

application.add_handler(CommandHandler("start", start))
application.add_handler(CommandHandler("go", start))
application.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"(?i)^go$"), handle_go_text))
application.add_handler(CommandHandler("go25s", handle_control_panel))

# ✅ أوامر لوحة التحكم العامة + إشعار التحديث + وضع الصيانة
application.add_handler(
    CallbackQueryHandler(
        handle_control_buttons,
        pattern="^(ctrl_maintenance_on|ctrl_maintenance_off|reload_settings|add_admin|list_admins|clear_sessions|self_destruct|control_back|admins_menu|restart_session|delete_admin|broadcast_update|ctrl_backup|exit_control)$"
    )
)

# ✅ استقبال رسائل المستخدمين والمشرفين (اقتراحات وردود مخصصة)
application.add_handler(MessageHandler(filters.ALL & ~filters.COMMAND, handle_message))

# ✅ نظام الاقتراحات (إرسال + ردود سريعة + رد مخصص)
application.add_handler(CallbackQueryHandler(send_suggestion, pattern=r"^send_suggestion$"))
# ✅ نقاشات فريق GO الداخلية
application.add_handler(CallbackQueryHandler(start_team_general_thread, pattern=r"^team_main_\d+$"))
# ✅ إرسال توصية فنية عامة للمجموعات
# ✅ إرسال توصية فنية عامة للمجموعات
# ✅ إرسال توصية فنية عامة للمجموعات
application.add_handler(CallbackQueryHandler(start_recommendation, pattern=r"^send_reco$"))
application.add_handler(CallbackQueryHandler(broadcast_recommendation, pattern=r"^reco_broadcast(_all)?$"))
application.add_handler(CallbackQueryHandler(cancel_recommendation, pattern=r"^reco_cancel$"))

# ✅ اختيار المجموعات يدوياً للتوصية + التثبيت
application.add_handler(CallbackQueryHandler(show_reco_groups, pattern=r"^reco_select$"))
application.add_handler(CallbackQueryHandler(toggle_reco_group, pattern=r"^reco_tgl_-?\d+$"))
application.add_handler(CallbackQueryHandler(change_reco_page, pattern=r"^reco_page_(prev|next)$"))
application.add_handler(CallbackQueryHandler(toggle_reco_pin, pattern=r"^reco_pin_toggle$"))

application.add_handler(CallbackQueryHandler(start_team_opinion_thread, pattern=r"^team_opinion_\d+_.+$"))
application.add_handler(CallbackQueryHandler(team_reply_existing_thread, pattern=r"^team_reply_\d+$"))

application.add_handler(
    CallbackQueryHandler(handle_suggestion_reply, pattern=r"^reply(?:ready)?_\d+_.+$")
)
application.add_handler(CallbackQueryHandler(handle_send_reply, pattern=r"^sendreply_[a-zA-Z0-9]+_\d+_.+$"))
application.add_handler(CallbackQueryHandler(handle_custom_reply, pattern=r"^customreply_\d+_.+$"))
application.add_handler(CallbackQueryHandler(submit_admin_reply, pattern=r"^submit_admin_reply$"))

# ✅ التقييم
application.add_handler(CallbackQueryHandler(show_statistics, pattern=r"^rate_\d+$"))
application.add_handler(CallbackQueryHandler(save_rating, pattern=r"^ratingval_\d+_\d+$"))

# ✅ الصيانة وقطع الغيار
application.add_handler(CallbackQueryHandler(car_choice, pattern=r"^car_.*_\d+$"))
application.add_handler(CallbackQueryHandler(maintenance_brand_choice, pattern=r"^mbrand_.*_\d+$"))
application.add_handler(CallbackQueryHandler(parts_brand_choice, pattern=r"^pbrand_.*_\d+$"))
application.add_handler(CallbackQueryHandler(km_choice, pattern=r"^km_.*_\d+$"))
application.add_handler(CallbackQueryHandler(send_cost, pattern=r"^cost_\d+_\d+$"))
application.add_handler(CallbackQueryHandler(send_part_image, pattern=r"^part_image_\d+_\d+$"))

# ✅ أزرار القوائم الخاصة بالصيانة وقطع الغيار والاقتراحات والأعطال + الرجوع
# ✅ أزرار التصنيف داخل نفس القائمة (تحت كل فئة)
application.add_handler(CallbackQueryHandler(button, pattern=r"^catpart_.*_\d+$"))

# ✅ أزرار الخدمات الرئيسية (قطع غيار / صيانة / ... من القائمة الرئيسية)
application.add_handler(
    CallbackQueryHandler(
        button,
        pattern=r"^(parts|maintenance|consumable|external|suggestion|coming|coming_soon)_\d+$"
    )
)

# ✅ اختيار سيارة للقطع الاستهلاكية + الرجوع لقائمة التصنيفات لنفس السيارة
application.add_handler(CallbackQueryHandler(select_car_for_parts, pattern=r"^(carpart_|showparts_).*"))

# الأعطال الشائعة من القائمة الرئيسية
application.add_handler(CallbackQueryHandler(button, pattern=r"^faults_\d+$"))
# ✅ تصنيفات الأعطال الفرعية
application.add_handler(CallbackQueryHandler(button, pattern=r"^faultcat_\d+_\d+$"))
# أزرار الرجوع القديمة من نوع back_main_USERID
application.add_handler(CallbackQueryHandler(button, pattern=r"^back_main_\d+$"))
# أزرار الرجوع الموحدة من نوع back:target:user_id
application.add_handler(CallbackQueryHandler(button, pattern=r"^back:"))
application.add_handler(CallbackQueryHandler(button, pattern=r"^cancelteam$"))

application.add_handler(CallbackQueryHandler(send_brochure, pattern=r"^brochure_\d+_\d+$"))

# ✅ دليل المالك
application.add_handler(CallbackQueryHandler(show_manual_car_list, pattern=r"^manual_"))
application.add_handler(CallbackQueryHandler(manual_brand_choice, pattern=r"^mnlbrand_.*_\d+$"))
application.add_handler(CallbackQueryHandler(handle_manualcar, pattern=r"^manualcar_.*_\d+$"))
application.add_handler(CallbackQueryHandler(handle_manualdfcar, pattern=r"^openpdf_"))

# ✅ المراكز والمتاجر
application.add_handler(CallbackQueryHandler(handle_service_centers, pattern=r"^service_\d+$"))
application.add_handler(CallbackQueryHandler(handle_branch_list, pattern=r"^branches_\d+$"))
application.add_handler(CallbackQueryHandler(handle_independent, pattern=r"^independent_\d+$"))
application.add_handler(CallbackQueryHandler(show_center_list, pattern=r"^show_centers_\d+$"))
application.add_handler(CallbackQueryHandler(show_store_list, pattern=r"^show_stores_\d+$"))
application.add_handler(CallbackQueryHandler(set_city, pattern=r"^setcity_.*_\d+$"))

# ✅ زر الإلغاء
application.add_handler(CallbackQueryHandler(handle_cancel, pattern=r"^cancel_"))

# ✅ زر غير نشط
application.add_handler(CallbackQueryHandler(
    lambda u, c: asyncio.create_task(u.callback_query.answer("🚫 هذا الزر غير نشط حالياً.")),
    pattern=r"^disabled$"
))

@app.api_route("/", methods=["GET", "HEAD"])
async def root():
    return {"message": "Bot is alive"}

@app.post("/webhook")
async def webhook_handler(request: Request):
    json_data = await request.json()

    # 🔎 لوق بسيط كل ما تيجي أبديت من تيليجرام
    logging.info(f"[WEBHOOK] وصل تحديث جديد من تيليجرام: keys={list(json_data.keys())}")

    update = Update.de_json(json_data, application.bot)
    await application.update_queue.put(update)
    return {"ok": True}

@app.on_event("startup")
async def on_startup():
    # 🔗 نبني رابط الـ Webhook بشكل مضمون
    base_url = os.getenv("RENDER_EXTERNAL_URL") or "https://chery-go-8a2z.onrender.com"

    # لو حطيت الدومين بدون بروتوكول نضيف https
    if not base_url.startswith("http"):
        base_url = "https://" + base_url.lstrip("/")

    # لو أحد كتبها أصلاً مع /webhook ما نكررها
    if base_url.endswith("/webhook"):
        webhook_url = base_url
    else:
        webhook_url = base_url.rstrip("/") + "/webhook"

    try:
        response = requests.get(
            f"https://api.telegram.org/bot{API_TOKEN}/setWebhook",
            params={"url": webhook_url},
            timeout=10,
        )
        logging.info(f"🔗 Webhook set to {webhook_url} status={response.status_code} body={response.text}")
    except Exception as e:
        logging.error(f"❌ Failed to set webhook: {e}")

    await application.initialize()
    await application.start()

        # ✅ تفعيل JobQueue (تنظيف الجلسات + health + النسخ الاحتياطي اليومي + keepalive)
    if application.job_queue:
        application.job_queue.run_repeating(
            cleanup_old_sessions,
            interval=60 * 60,  # كل ساعة
            first=60           # أول تشغيل بعد 60 ثانية من الإقلاع
        )

        # نبضات صحية دورية داخل الذاكرة فقط
        application.job_queue.run_repeating(
            health_log_job,
            interval=60 * 10,  # كل 10 دقائق
            first=120
        )

        # 🔁 KEEPALIVE: طلب داخلي للخدمة كل 5 دقائق لإبقائها مستيقظة
        try:
            application.job_queue.run_repeating(
                keepalive_ping,
                interval=60 * 5,   # كل 5 دقائق
                first=180,         # أول تشغيل بعد 3 دقائق من الإقلاع
                name="render_keepalive",
            )
        except Exception as e:
            logging.error(f"[KEEPALIVE] ❌ فشل جدولة keepalive: {e}")

        # نسخ احتياطي يومي للبيانات الساعة 4 فجراً بتوقيت السعودية
        try:
            saudi_tz = timezone(timedelta(hours=3))
            application.job_queue.run_daily(
                daily_backup_job,
                time=time(hour=4, minute=0, tzinfo=saudi_tz),
                name="daily_excel_backup",
            )
        except Exception as e:
            logging.error(f"[BACKUP] ❌ فشل جدولة النسخ الاحتياطي اليومي: {e}")

        print("✅ JobQueue تم تشغيلها")
    else:
        print("⚠️ job_queue غير مفعلة أو غير جاهزة")